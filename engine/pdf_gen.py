import streamlit as st
import pandas as pd
import numpy as np
import datetime
import base64
import os
import tempfile
import io
from fpdf import FPDF


class PDFReport(FPDF):
    def footer(self):
        self.set_y(-15)
        self.set_font("Arial", "I", 8)
        self.set_text_color(150, 150, 150)
        self.multi_cell(0, 4, _s("Document genere par Logiflo.io. Recommandations a titre indicatif."), align="C")


def _s(text):
    if text is None: return ""
    text = str(text)
    REPL = {
        "\u2019":"'","\u2018":"'","\u201c":'"',"\u201d":'"',
        "\u2013":"-","\u2014":"-","\u2026":"...","\u20ac":"EUR",
        "\u00a9":"(c)","\u00ae":"(R)","\u2122":"TM","\u2022":"-",
        "\u25cf":"-","\u2192":"->","\u2190":"<-","\u21d2":"=>",
        "\u2713":"OK","\u2717":"X","\u00b0":"deg","\u00b1":"+/-",
        "\u00d7":"x","\u00f7":"/","\u2248":"~","\u2260":"!=",
        "\u2264":"<=","\u2265":">=","\u26a0":"[!]","\u2139":"[i]",
        "\u2605":"*","\u2606":"*",
    }
    for ch, rep in REPL.items():
        text = text.replace(ch, rep)
    result = []
    for c in text:
        if ord(c) > 255:
            try:    result.append(c.encode('ascii', errors='strict').decode('ascii'))
            except: result.append('?')
        else:
            result.append(c)
    return ''.join(result)


def _asc(text): return _s(text)
def _clean_pdf(text): return _s(str(text).replace("**", ""))


def predict_ruptures(df, seuil_rupture=0, lang="fr"):
    if df is None or len(df) == 0: return []
    alertes = []
    cols_conso = [c for c in ["conso_an4","conso_an3","conso_an2","conso_an1"] if c in df.columns]
    if not cols_conso: return []
    try:
        for _, row in df.iterrows():
            ref   = str(row.get("reference", "?"))
            stock = float(row.get("quantite", 0))
            if stock <= seuil_rupture: continue
            consos = []
            for c in cols_conso:
                v = row.get(c, 0)
                try:
                    v = float(v)
                    if v > 0: consos.append(v / 52)
                except Exception: pass
            if not consos: continue
            conso_hebdo = sum(consos) / len(consos)
            if conso_hebdo <= 0: continue
            semaines = stock / conso_hebdo
            if semaines <= 4:
                urgence = "critique" if semaines <= 1 else ("urgent" if semaines <= 2 else "alerte")
                alertes.append({"reference":ref,"stock":stock,"conso_hebdo":round(conso_hebdo,1),"semaines":round(semaines,1),"urgence":urgence})
        alertes.sort(key=lambda x: x["semaines"])
        return alertes[:10]
    except Exception:
        return []


def format_predictions_pour_prompt(alertes, lang="fr"):
    if not alertes: return ""
    lines = []
    if lang == "en":
        lines.append("=== STOCKOUT PREDICTIONS (next 4 weeks) ===")
        for a in alertes:
            sem = a['semaines']
            if sem < 2:
                delai = f"{max(1, round(sem * 7))} days"
            else:
                delai = f"{round(sem)} weeks"
            lines.append(f"  {a['urgence'].upper()} -- {a['reference']}: {a['stock']:.0f} units, consumption {a['conso_hebdo']:.1f}/week, stockout in ~{delai}")
    else:
        lines.append("=== PREDICTIONS RUPTURE (4 prochaines semaines) ===")
        for a in alertes:
            sem = a['semaines']
            if sem < 2:
                delai = f"{max(1, round(sem * 7))} jours"
            else:
                delai = f"{round(sem)} semaines"
            lines.append(f"  {a['urgence'].upper()} -- {a['reference']}: {a['stock']:.0f} unites, conso {a['conso_hebdo']:.1f}/semaine, rupture dans ~{delai}")
    return "\n".join(lines)

def compute_alerte_bfr(df, ca_annuel_estime=None, lang="fr"):
    result = {"available": False, "texte": "", "capital_liberatable": 0}
    if df is None or len(df) == 0: return result
    try:
        if "valeur_totale" in df.columns:
            capital_total = df["valeur_totale"].sum()
        elif "quantite" in df.columns and "prix_unitaire" in df.columns:
            capital_total = (df["quantite"] * df["prix_unitaire"]).sum()
        else:
            return result
        if capital_total <= 0: return result

        capital_dormant = 0
        capital_surstock = 0
        if "Statut" in df.columns and "valeur_totale" in df.columns:
            mask_dorm = df["Statut"].str.contains("Dormant", na=False)
            capital_dormant = df.loc[mask_dorm, "valeur_totale"].sum()
            mask_surs = df["Statut"].str.contains("Surstock", na=False)
            capital_surstock = df.loc[mask_surs, "valeur_totale"].sum()

        capital_lib = capital_dormant + capital_surstock * 0.5
        cout_possession = round(capital_dormant * 0.20)
        bfr_jours = None
        if ca_annuel_estime and ca_annuel_estime > 0:
            bfr_jours = round((capital_total / ca_annuel_estime) * 365)

        result["available"] = True
        result["capital_liberatable"] = round(capital_lib)
        result["capital_dormant"] = round(capital_dormant)
        result["capital_surstock"] = round(capital_surstock)
        result["cout_possession"] = cout_possession
        result["bfr_jours"] = bfr_jours

        if lang == "en":
            result["texte"] = (f"BFR ALERT: {capital_lib:,.0f} EUR can be freed (dormant: {capital_dormant:,.0f} EUR + overstock: {capital_surstock:,.0f} EUR). Annual holding cost: ~{cout_possession:,.0f} EUR/year." + (f" WCR: ~{bfr_jours} days revenue." if bfr_jours else ""))
        else:
            result["texte"] = (f"ALERTE BFR : {capital_lib:,.0f} EUR liberables (dormant : {capital_dormant:,.0f} EUR + surstock : {capital_surstock:,.0f} EUR). Cout possession : ~{cout_possession:,.0f} EUR/an." + (f" BFR : ~{bfr_jours} jours de CA." if bfr_jours else ""))
    except Exception:
        pass
    return result


def render_prediction_rupture(df, lang="fr"):
    alertes = predict_ruptures(df, lang=lang)
    if not alertes: return
    _lbl = "Predictions de rupture -- 4 semaines" if lang=="fr" else "Stockout Predictions -- 4 weeks"
    st.markdown(f'<div style="font-size:11px;font-weight:700;color:#4A6080;letter-spacing:2px;text-transform:uppercase;margin-bottom:10px;margin-top:20px;">⚠️ {_lbl}</div>', unsafe_allow_html=True)
    for a in alertes:
        sem = a["semaines"]
        if sem < 2:
            delai = f"{max(1, round(sem * 7))} {'jours' if lang=='fr' else 'days'}"
        else:
            delai = f"{round(sem)} {'semaines' if lang=='fr' else 'weeks'}"
        _clr  = "#E8304A" if a["urgence"]=="critique" else ("#F39C12" if a["urgence"]=="urgent" else "#F59E0B")
        _bg   = "#FFF1F2" if a["urgence"]=="critique" else ("#FFFBEB" if a["urgence"]=="urgent" else "#FEFCE8")
        _icon = "🔴" if a["urgence"]=="critique" else ("🟠" if a["urgence"]=="urgent" else "🟡")
        if lang == "en":
            _msg = f"<strong>{a['reference']}</strong> -- {a['stock']:.0f} units -- consumption {a['conso_hebdo']:.1f}/week -- <strong>stockout in ~{delai}</strong>"
        else:
            _msg = f"<strong>{a['reference']}</strong> -- {a['stock']:.0f} unites -- conso {a['conso_hebdo']:.1f}/sem -- <strong>rupture dans ~{delai}</strong>"
        st.markdown(f'<div style="background:{_bg};border-left:4px solid {_clr};border-radius:8px;padding:10px 14px;margin-bottom:6px;display:flex;align-items:center;gap:10px;"><span style="font-size:18px;">{_icon}</span><div style="flex:1;font-size:12px;color:#0B2545;">{_msg}</div><div style="font-size:10px;font-weight:700;color:{_clr};text-transform:uppercase;">{a["urgence"]}</div></div>', unsafe_allow_html=True)


def tooltip_metric(label, value, unit="", delta=None, lang="fr"):
    KPI_TOOLTIPS = {
        "fr": {
            "Taux de service":("Pourcentage de commandes livrees sans rupture.","Benchmark : > 93% B2B"),
            "Capital immobilise":("Valeur totale du stock en EUR.","Norme : < 60 jours de CA"),
            "Ruptures":("References dont le stock est a zero.","Objectif : < 2% des references"),
            "Marge nette":("CA moins les couts d exploitation.","Norme PME transport : 6-10%"),
            "Trajets toxiques":("Trajets avec marge < 5% ou negative.","Alerte si > 20% du portefeuille"),
        },
        "en": {
            "Service Level":("Percentage of orders fulfilled without stockout.","Benchmark: > 93% B2B"),
            "Tied-up Capital":("Total stock value in EUR.","Norm: < 60 days revenue"),
            "Stockouts":("References at zero stock.","Target: < 2% of references"),
            "Net Margin":("Revenue minus operating costs.","SME transport norm: 6-10%"),
            "Toxic Routes":("Routes with margin < 5% or negative.","Alert if > 20% of portfolio"),
        }
    }
    tips = KPI_TOOLTIPS.get(lang, KPI_TOOLTIPS["fr"])
    tip = tips.get(label)
    if tip:
        def_txt, bench_txt = tip
        tooltip_html = f'<span style="font-size:11px;color:#4A6080;border-bottom:1.5px dashed #4A6080;cursor:help;" title="{def_txt} | {bench_txt}">{label}</span>'
    else:
        tooltip_html = f'<span style="font-size:11px;color:#4A6080;">{label}</span>'

    if isinstance(value, float) and abs(value) >= 1000: val_str = f"{value:,.0f}"
    elif isinstance(value, float): val_str = f"{value:.1f}%"
    else: val_str = str(value)

    color = "#00C896"
    if delta is not None:
        color = "#00C896" if delta >= 0 else "#E8304A"

    st.markdown(f'<div style="background:white;border:1px solid #E2E8F0;border-top:3px solid {color};border-radius:12px;padding:14px 16px;text-align:center;">{tooltip_html}<div style="font-family:Syne,sans-serif;font-size:26px;font-weight:800;color:{color};margin-top:4px;line-height:1;">{val_str}<span style="font-size:13px;font-weight:400;color:#4A6080;">{unit}</span></div></div>', unsafe_allow_html=True)


def generate_exemple_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    ws_stock = wb.active
    ws_stock.title = "Stock"
    headers_stock = ["Reference","Designation","Fournisseur","Prix_Achat","Prix_Vente","Stock","Conso_2023","Conso_2024","Conso_2025","Categorie","Date_Entree"]
    for ci, h in enumerate(headers_stock, 1):
        cell = ws_stock.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B2545")
        cell.alignment = Alignment(horizontal="center")
    data_stock = [
        ("REF-001","T-shirt Coton Bio","FournisseurA",8.5,22.0,45,120,135,98,"Textile","15/01/2025"),
        ("REF-002","Pantalon Slim","FournisseurB",15.0,45.0,12,80,75,62,"Textile","20/01/2025"),
        ("REF-003","Chaussures Sport","FournisseurC",25.0,89.0,0,60,55,48,"Chaussures","10/02/2025"),
        ("REF-004","Veste Laine","FournisseurA",35.0,120.0,3,40,38,29,"Textile","05/03/2025"),
        ("REF-005","Echarpe","FournisseurD",18.0,65.0,28,55,50,44,"Accessoires","12/03/2025"),
        ("REF-006","Sac a Dos","FournisseurB",12.0,38.0,0,90,95,88,"Accessoires","08/04/2025"),
        ("REF-007","Bonnet Laine","FournisseurA",5.5,18.0,67,200,185,170,"Accessoires","02/01/2025"),
        ("REF-008","Pull Cachemire","FournisseurD",45.0,150.0,8,30,25,0,"Textile","18/02/2025"),
        ("REF-009","Jean Denim","FournisseurC",20.0,65.0,34,110,105,98,"Textile","25/01/2025"),
        ("REF-010","Chemise Oxford","FournisseurB",18.0,55.0,15,70,68,62,"Textile","30/01/2025"),
    ]
    for ri, row in enumerate(data_stock, 2):
        for ci, val in enumerate(row, 1):
            ws_stock.cell(row=ri, column=ci, value=val)

    ws_trans = wb.create_sheet("Transport")
    headers_trans = ["Client","Depart","Arrivee","Distance_km","CA_EUR","Cout_EUR","Poids_kg","Date_Livraison","Mode"]
    for ci, h in enumerate(headers_trans, 1):
        cell = ws_trans.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B2545")
        cell.alignment = Alignment(horizontal="center")
    data_trans = [
        ("Carrefour","Marseille","Lyon",315,450,385,1200,"05/04/2025","Routier"),
        ("Auchan","Paris","Bordeaux",580,820,710,2500,"06/04/2025","Routier"),
        ("Lidl","Lyon","Toulouse",295,420,368,900,"07/04/2025","Routier"),
        ("Leclerc","Bordeaux","Paris",580,750,680,1800,"08/04/2025","Routier"),
        ("Fnac","Paris","Lille",225,310,272,800,"08/04/2025","Routier"),
        ("Decathlon","Lyon","Paris",465,640,575,2200,"09/04/2025","Routier"),
        ("GlobeTrans","Marseille","Barcelona",820,1100,985,2200,"13/04/2025","Routier"),
        ("AlphaFret","Marseille","Turin",450,610,545,2800,"14/04/2025","Routier"),
        ("SudLog","Toulouse","Madrid",1050,1450,1280,3500,"15/04/2025","Routier"),
        ("NordTrans","Lille","Hamburg",680,920,810,2100,"15/04/2025","Routier"),
    ]
    for ri, row in enumerate(data_trans, 2):
        for ci, val in enumerate(row, 1):
            ws_trans.cell(row=ri, column=ci, value=val)

    for ws in [ws_stock, ws_trans]:
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_free_pdf(module, summary_text, kpis, labels):
    pdf = PDFReport()
    lang = st.session_state.get("language", "fr")
    pdf.add_page()
    pdf.set_fill_color(11,37,69); pdf.rect(0,0,210,297,'F')
    pdf.set_fill_color(0,200,150); pdf.rect(0,0,210,6,'F')
    pdf.set_y(90)
    pdf.set_text_color(255,255,255); pdf.set_font("Arial","B",34)
    pdf.cell(0,16,"LOGIFLO.IO",ln=True,align='C')
    pdf.set_font("Arial","",13); pdf.set_text_color(0,200,150)
    lbl_free = "Free Audit" if lang=="en" else "Audit Gratuit"
    pdf.cell(0,10,_s(f"[ {lbl_free} ]"),ln=True,align='C')
    pdf.ln(10)
    pdf.set_text_color(200,220,255); pdf.set_font("Arial","",12)
    pdf.cell(0,8,_s(f"Date : {datetime.date.today().strftime('%d/%m/%Y')}"),ln=True,align='C')
    pdf.cell(0,8,_s(f"Module : {'STOCK' if module=='stock' else 'TRANSPORT'}"),ln=True,align='C')
    pdf.set_fill_color(0,200,150); pdf.rect(0,291,210,6,'F')

    pdf.add_page()
    pdf.set_fill_color(11,37,69); pdf.rect(0,0,210,18,'F')
    pdf.set_y(4); pdf.set_text_color(255,255,255); pdf.set_font("Arial","B",11)
    pdf.cell(0,10,_s("YOUR AUDIT RESULTS" if lang=="en" else "RESULTATS DE VOTRE AUDIT"),ln=True,align='C')
    pdf.ln(8)

    if kpis and labels:
        n = min(len(kpis), len(labels), 3)
        card_w = 56; total_w = n*card_w+(n-1)*8; start_x = (210-total_w)/2; card_y = pdf.get_y()
        for i in range(n):
            cx = start_x + i*(card_w+8)
            pdf.set_fill_color(240,244,248); pdf.rect(cx,card_y,card_w,34,'F')
            pdf.set_fill_color(0,168,122); pdf.rect(cx,card_y,card_w,3,'F')
            pdf.set_xy(cx+2,card_y+5); pdf.set_font("Arial","",7); pdf.set_text_color(74,96,128)
            pdf.cell(card_w-4,6,_s(labels[i]).upper()[:22],align='C')
            pdf.set_xy(cx+2,card_y+13); pdf.set_font("Arial","B",15); pdf.set_text_color(11,37,69)
            val = kpis[i]
            if isinstance(val,float) and abs(val)>=1000: vs=_s(f"{val:,.0f}")
            elif isinstance(val,float): vs=_s(f"{val:.1f}%")
            else: vs=_s(str(val))
            pdf.cell(card_w-4,10,vs,align='C')
        pdf.ln(42)

    pdf.set_font("Arial","B",11); pdf.set_text_color(11,37,69)
    pdf.cell(0,8,_s("RAPID DIAGNOSIS" if lang=="en" else "DIAGNOSTIC RAPIDE"),ln=True)
    pdf.set_draw_color(0,200,150); pdf.line(10,pdf.get_y(),200,pdf.get_y()); pdf.ln(4)
    lines_done = 0
    for line in summary_text.split("\n"):
        line = line.strip()
        if not line: pdf.ln(2); continue
        if pdf.get_y() > 258: break
        if line.startswith("### "):
            pdf.set_font("Arial","B",10); pdf.set_text_color(0,168,122)
            pdf.cell(0,7,_s(line[4:].upper()),ln=True); lines_done+=1
        elif lines_done < 25:
            pdf.set_font("Arial","",9); pdf.set_text_color(40,40,40)
            pdf.set_x(10); pdf.multi_cell(190,5,_s(line.replace("**",""))); lines_done+=1

    pdf.set_y(260)
    pdf.set_fill_color(240,244,248); pdf.rect(10,pdf.get_y(),190,28,'F')
    pdf.set_fill_color(0,200,150); pdf.rect(10,pdf.get_y(),3,28,'F')
    pdf.set_xy(16,pdf.get_y()+5); pdf.set_font("Arial","B",10); pdf.set_text_color(11,37,69)
    cta1 = "Get the full audit with history, charts and scoring." if lang=="en" else "Obtenez l audit complet avec historique, graphiques et scoring."
    pdf.multi_cell(184,6,_s(cta1))
    pdf.set_x(16); pdf.set_font("Arial","",9); pdf.set_text_color(74,96,128)
    pdf.cell(0,6,"logiflo-io.streamlit.app  |  contact@logiflo.io",ln=True)

    try:
        raw = pdf.output(dest='S')
        if isinstance(raw, bytes): return raw
        return raw.encode('latin-1', errors='replace')
    except Exception:
        return b""


def generate_expert_pdf(title, content, figs=None, kpis=None, labels=None, module="stock"):
    if kpis is None:   kpis = []
    if labels is None: labels = []
    pdf = PDFReport()
    lang = st.session_state.get("language", "fr")

    # PAGE 1 COUVERTURE
    pdf.add_page()
    pdf.set_fill_color(11,37,69); pdf.rect(0,0,210,297,'F')
    pdf.set_fill_color(0,200,150); pdf.rect(0,0,210,6,'F')
    pdf.set_y(80)
    pdf.set_text_color(255,255,255); pdf.set_font("Arial","B",38)
    pdf.cell(0,18,"LOGIFLO.IO",ln=True,align='C')
    pdf.set_font("Arial","",14); pdf.set_text_color(0,200,150)
    pdf.cell(0,10,"[ Logistics Intelligence ]",ln=True,align='C')
    pdf.ln(8)
    pdf.set_draw_color(0,200,150); pdf.set_line_width(0.8)
    pdf.line(40,pdf.get_y(),170,pdf.get_y()); pdf.ln(10)
    pdf.set_text_color(255,255,255); pdf.set_font("Arial","B",22)
    pdf.multi_cell(0,12,_s(title),align='C'); pdf.ln(4)
    pdf.set_font("Arial","",12); pdf.set_text_color(180,200,220)
    pdf.cell(0,8,_s(f"Date : {datetime.date.today().strftime('%d/%m/%Y')}"),ln=True,align='C')
    pdf.cell(0,8,_s("CONFIDENTIAL" if lang=="en" else "CONFIDENTIEL"),ln=True,align='C')
    pdf.set_fill_color(0,200,150); pdf.rect(0,291,210,6,'F')

    # PAGE 2 SYNTHESE
    pdf.add_page()
    pdf.set_fill_color(11,37,69); pdf.rect(0,0,210,18,'F')
    pdf.set_y(4); pdf.set_text_color(255,255,255); pdf.set_font("Arial","B",11)
    pdf.cell(0,10,_s("EXECUTIVE SUMMARY" if lang=="en" else "SYNTHESE EXECUTIVE"),ln=True,align='C')
    pdf.ln(8)
    pdf.set_text_color(11,37,69); pdf.set_font("Arial","B",16)
    pdf.cell(0,10,_s("Key Indicators" if lang=="en" else "Indicateurs Cles"),ln=True,align='L')
    pdf.set_draw_color(0,200,150); pdf.set_line_width(0.6)
    pdf.line(10,pdf.get_y(),200,pdf.get_y()); pdf.ln(8)

    if kpis and labels:
        n = min(len(kpis),len(labels),3)
        card_w = 56; total_w = n*card_w+(n-1)*8; start_x = (210-total_w)/2
        card_colors = [(0,168,122),(0,168,122),(232,48,74)]
        card_y = pdf.get_y()
        for i in range(n):
            cx = start_x + i*(card_w+8)
            pdf.set_fill_color(240,244,248); pdf.rect(cx,card_y,card_w,38,'F')
            r,g,b = card_colors[i] if i < len(card_colors) else (0,168,122)
            pdf.set_fill_color(r,g,b); pdf.rect(cx,card_y,card_w,3,'F')
            pdf.set_xy(cx+2,card_y+5); pdf.set_font("Arial","",7); pdf.set_text_color(74,96,128)
            pdf.cell(card_w-4,6,_asc(labels[i]).upper()[:22],align='C')
            pdf.set_xy(cx+2,card_y+14); pdf.set_font("Arial","B",18); pdf.set_text_color(r,g,b)
            val = kpis[i]
            if isinstance(val,float) and abs(val)>=1000: val_str=_s(f"{val:,.0f}")
            elif isinstance(val,float) and abs(val)<=100: val_str=_s(f"{val:.1f}%")
            else: val_str=_s(str(int(val)) if isinstance(val,float) else str(val))
            pdf.cell(card_w-4,12,val_str,align='C')
        pdf.ln(46)

    try:
        from engine.scoring import compute_logiflo_score
        _score_pdf = compute_logiflo_score(module=module,df=None,kpis=kpis,labels=labels,sector_key="generique",lang=lang)
        _details_pdf = _score_pdf.get("details", {})
    except Exception:
        _details_pdf = {}

    if _details_pdf:
        for _dim_label, _sv in list(_details_pdf.items())[:3]:
            _sv = int(_sv) if _sv else 0
            rc,gc,bc = (0,168,122) if _sv>=70 else (243,156,18) if _sv>=40 else (232,48,74)
            _lx=10; _bx=72; _sx=183; _row_y=pdf.get_y()
            pdf.set_font("Arial","",8); pdf.set_text_color(74,96,128)
            pdf.set_xy(_lx,_row_y); pdf.cell(60,7,_s(str(_dim_label))[:32],align='L')
            _bar_y = _row_y+1
            pdf.set_fill_color(225,232,240); pdf.rect(_bx,_bar_y,108,5,'F')
            pdf.set_fill_color(rc,gc,bc)
            _fill = int((_sv/100)*108) if _sv>0 else 0
            if _fill>0: pdf.rect(_bx,_bar_y,_fill,5,'F')
            pdf.set_font("Arial","B",8); pdf.set_text_color(rc,gc,bc)
            pdf.set_xy(_sx,_row_y); pdf.cell(22,7,f"{_sv}/100",align='R')
            pdf.ln(8)

    # PAGE 3 GRAPHIQUES
    if figs:
        pdf.add_page()
        pdf.set_fill_color(11,37,69); pdf.rect(0,0,210,18,'F')
        pdf.set_y(4); pdf.set_text_color(255,255,255); pdf.set_font("Arial","B",11)
        pdf.cell(0,10,_s("CHARTS & VISUALIZATIONS" if lang=="en" else "GRAPHIQUES & VISUALISATIONS"),ln=True,align='C')
        pdf.ln(6)
        for fig in figs:
            _tp = None
            try:
                import uuid
                _tp = os.path.join(tempfile.gettempdir(), f"lgf_{uuid.uuid4().hex}.png")
                _ok = False
                try:
                    _b = fig.to_image(format="png", width=740, height=320)
                    if _b and len(_b) > 200:
                        with open(_tp,"wb") as _f: _f.write(_b)
                        _ok = True
                except Exception:
                    pass
                if not _ok:
                    try:
                        fig.write_image(_tp, format="png", width=740, height=320)
                        if os.path.exists(_tp) and os.path.getsize(_tp) > 200: _ok = True
                    except Exception:
                        pass
                if _ok and os.path.exists(_tp):
                    _space = 297 - pdf.get_y() - 20
                    _img_h = 84; _img_w = 176
                    if _space < _img_h + 10:
                        pdf.add_page(); pdf.ln(5)
                    _margin_x = (210 - _img_w) / 2
                    pdf.image(_tp, x=_margin_x, y=pdf.get_y(), w=_img_w)
                    pdf.ln(_img_h + 6)
            except Exception:
                pass
            finally:
                if _tp:
                    try: os.unlink(_tp)
                    except: pass

    # PAGE 4 ANALYSE IA
    pdf.add_page()
    pdf.set_fill_color(11,37,69); pdf.rect(0,0,210,18,'F')
    pdf.set_y(4); pdf.set_text_color(255,255,255); pdf.set_font("Arial","B",11)
    pdf.cell(0,10,_s("AI ANALYSIS & RECOMMENDATIONS" if lang=="en" else "ANALYSE IA & RECOMMANDATIONS"),ln=True,align='C')
    pdf.ln(8)

    content_r = (content.replace("\u2019","'").replace("\u2018","'").replace("\u201c",'"').replace("\u201d",'"').replace("\u20ac","EUR").replace("\u2022","-").replace("\u2013","-").replace("\u2014","-"))
    _in_scoring = False
    for line in content_r.split('\n'):
        line = line.strip()
        if 'SCORING' in line.upper() and line.startswith('###'):
            _in_scoring = True
        if _in_scoring and not line.startswith('###'):
            continue
        if _in_scoring and line.startswith('###') and 'SCORING' not in line.upper():
            _in_scoring = False
        if not line:
            pdf.ln(2); continue
        if line.startswith('### '):
            if pdf.get_y() > 250: pdf.add_page(); pdf.ln(4)
            t = _asc(line[4:])
            pdf.ln(5); _yh = pdf.get_y()
            pdf.set_fill_color(240,244,248); pdf.rect(10,_yh,190,10,'F')
            pdf.set_fill_color(0,200,150); pdf.rect(10,_yh,3,10,'F')
            pdf.set_font("Arial","B",10); pdf.set_text_color(11,37,69)
            pdf.set_x(16); pdf.cell(184,10,_s(t).upper(),ln=True); pdf.ln(4)
        elif line.startswith(('- ','* ')):
            _bt = _s(line[2:].replace("**",""))
            if pdf.get_y() + 8 > 272: pdf.add_page(); pdf.ln(4)
            pdf.set_font("Arial","",10); pdf.set_text_color(40,40,40)
            pdf.set_x(14); pdf.cell(5,6,"-"); pdf.set_x(19)
            pdf.multi_cell(181,6,_bt)
        else:
            _est = max(1, len(line)//45+1) * 6 + 4
            if pdf.get_y() + _est > 272: pdf.add_page(); pdf.ln(4)
            pdf.set_font("Arial","",10); pdf.set_text_color(40,40,40)
            pdf.set_x(10); pdf.multi_cell(190,6,_s(line.replace("**","")))

    # PAGE 5 CTA
    pdf.add_page()
    pdf.set_fill_color(11,37,69); pdf.rect(0,0,210,297,'F')
    pdf.set_fill_color(0,200,150); pdf.rect(0,0,210,6,'F'); pdf.rect(0,291,210,6,'F')
    pdf.set_y(85)
    pdf.set_text_color(0,200,150); pdf.set_font("Arial","B",32)
    pdf.cell(0,16,"LOGIFLO.IO",ln=True,align='C')
    pdf.ln(6)
    pdf.set_draw_color(0,200,150); pdf.set_line_width(0.6)
    pdf.line(50,pdf.get_y(),160,pdf.get_y()); pdf.ln(12)
    if lang == "en":
        cta_lines = [("This report was generated by LOGIFLO.IO",True,255),("","",200),("Designed by a field logistics professional.",False,200),("Not by a consultant.",False,200),("","",200),("To go further :",True,255),("contact@logiflo.io",False,150),("logiflo-io.streamlit.app",False,150)]
    else:
        cta_lines = [("Ce rapport a ete genere par LOGIFLO.IO",True,255),("","",200),("Concu par un logisticien terrain.",False,200),("Pas par un consultant.",False,200),("","",200),("Pour aller plus loin :",True,255),("contact@logiflo.io",False,150),("logiflo-io.streamlit.app",False,150)]
    for (txt,bold,br) in cta_lines:
        if not txt: pdf.ln(5); continue
        pdf.set_font("Arial","B" if bold else "",12 if bold else 11)
        pdf.set_text_color(br,br,br)
        pdf.cell(0,9,_s(txt),ln=True,align='C')

    try:
        raw = pdf.output(dest='S')
        if isinstance(raw, bytes): return raw
        return raw.encode('latin-1', errors='replace')
    except Exception:
        return b""
