# Public app - no Streamlit auth required

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import datetime
import re
import tempfile
import os
import difflib
import unicodedata
import math
import time
import requests
import concurrent.futures
import base64
import json
import io
from fpdf import FPDF
from openai import OpenAI
import gspread
try:
    from supabase import create_client as _supa_create
    from supabase import Client as _supa_Client
except Exception:
    _supa_create = None
    _supa_Client = None
from google.oauth2.service_account import Credentials

st.set_page_config(page_title="LOGIFLO.IO | Control Tower", layout="wide", page_icon="🏢")

# =========================================
# 0. INIT
# =========================================
client   = OpenAI(api_key=st.secrets.get("OPENAI_API_KEY", ""))
ORS_API_KEY = st.secrets.get("ORS_API_KEY", "")
SHEET_ID    = st.secrets.get("GOOGLE_SHEET_ID", "")

# =========================================
# SECTORAL BENCHMARKS DATABASE
# =========================================

# ══════════════════════════════════════════════════════════════════
# NEWS SECTORIELLES — Google RSS + NewsAPI fallback + cache Supabase
# ══════════════════════════════════════════════════════════════════
import xml.etree.ElementTree as _ET_news

NEWS_QUERIES = {
    "transport_routier":       "transport routier PME carburant CNR France actualite",
    "transport_maritime":      "fret maritime conteneur Marseille logistique actualite",
    "transport_aerien_intl":   "fret aerien cargo IATA logistique international",
    "transport_routier_eu":    "transport routier international Europe logistique",
    "stock_pharma":            "rupture medicaments supply chain pharmaceutique France",
    "stock_agroalim":          "logistique agroalimentaire froid chaine approvisionnement",
    "stock_distribution":      "distribution logistique entrepot stock rupture",
    "stock_industrie":         "logistique industrielle supply chain production",
    "stock_retail":            "logistique retail e-commerce stock gestion",
    "stock_btp":               "logistique chantier BTP materiau construction approvisionnement",
    "supply_chain_maghreb":    "logistique Maroc transport Casablanca supply chain",
    "supply_chain_afrique":    "logistique Afrique Abidjan Dakar transport",
    "transport_maritime_intl": "maritime shipping container freight port international",
    "generique":               "supply chain logistique France actualite innovation",
}

def fetch_news_google_rss(query, lang="fr", country="FR", max_results=5):
    """Récupère les news via Google RSS — sans clé API."""
    try:
        import urllib.parse
        url = (f"https://news.google.com/rss/search"
               f"?q={urllib.parse.quote(query)}"
               f"&hl={lang}-{country}&gl={country}&ceid={country}:{lang}")
        r = requests.get(url, timeout=8,
                         headers={"User-Agent": "Mozilla/5.0 Logiflo/1.0"})
        root = _ET_news.fromstring(r.content)
        items = root.findall(".//item")[:max_results]
        articles = []
        for item in items:
            title = item.findtext("title","").split(" - ")[0].strip()
            link  = item.findtext("link","")
            date  = item.findtext("pubDate","")[:16]
            if title and link:
                articles.append({"title":title,"link":link,"date":date})
        return articles
    except Exception:
        return []

def fetch_news_newsapi(query, lang="fr", max_results=5):
    """Récupère les news via NewsAPI.org — fallback si RSS vide."""
    try:
        key = st.secrets.get("NEWSAPI_KEY","")
        if not key:
            return []
        url = "https://newsapi.org/v2/everything"
        params = {
            "q":        query,
            "language": lang,
            "sortBy":   "publishedAt",
            "pageSize": max_results,
            "apiKey":   key,
        }
        r = requests.get(url, params=params, timeout=8)
        data = r.json()
        articles = []
        for a in data.get("articles",[])[:max_results]:
            articles.append({
                "title": a.get("title","").split(" - ")[0][:80],
                "link":  a.get("url",""),
                "date":  (a.get("publishedAt","") or "")[:10],
            })
        return articles
    except Exception:
        return []

def get_sector_news(sector_key, lang="fr"):
    """
    Retourne les 5 dernières news pour un secteur.
    Cache 4h pour limiter les appels API.
    Essaie d'abord Supabase cache, puis Google RSS, puis NewsAPI.
    """
    # 1. Vérifier le cache Supabase
    sb = get_supabase()
    if sb:
        try:
            resp = (sb.table("news_cache")
                      .select("articles,fetched_at")
                      .eq("sector_key", sector_key)
                      .order("fetched_at", desc=True)
                      .limit(1)
                      .execute())
            if resp.data:
                import datetime as _dt_n
                fetched = _dt_n.datetime.fromisoformat(
                    resp.data[0]["fetched_at"].replace("Z",""))
                age_h = (datetime.datetime.utcnow() - fetched).total_seconds() / 3600
                if age_h < 4:
                    return resp.data[0]["articles"] or []
        except Exception:
            pass

    # 2. Récupérer depuis Google RSS
    query = NEWS_QUERIES.get(sector_key, NEWS_QUERIES["generique"])
    country = "MA" if "maghreb" in sector_key else (
              "CI" if "afrique" in sector_key else "FR")
    articles = fetch_news_google_rss(query, lang=lang, country=country)

    # 3. Fallback NewsAPI si RSS vide
    if not articles:
        articles = fetch_news_newsapi(query, lang=lang)

    # 4. Sauvegarder dans le cache Supabase
    if sb and articles:
        try:
            sb.table("news_cache").insert({
                "sector_key": sector_key,
                "articles":   articles,
                "fetched_at": datetime.datetime.now().isoformat(),
            }).execute()
        except Exception:
            pass

    return articles

def render_news_widget(sector_key, lang="fr"):
    """News horizontales défilantes toutes les 5 secondes, clic vers article original."""
    news = get_sector_news(sector_key, lang)
    if not news:
        return

    _lbl = "Actualités sectorielles" if lang=="fr" else "Sector News"
    st.markdown(f"""
    <div style="font-size:11px;font-weight:700;color:#4A6080;
                letter-spacing:2px;text-transform:uppercase;
                margin-bottom:10px;margin-top:16px;">
        📰 {_lbl}
    </div>""", unsafe_allow_html=True)

    _articles = [a for a in news[:6] if a.get("title","").strip() and a.get("link","")]

    if not _articles:
        return

    # Construire les cards en HTML — défilement automatique JS toutes les 5s
    _cards_html = ""
    for _art in _articles:
        _title  = str(_art.get("title",""))[:90].replace("'","&#39;").replace('"','&quot;')
        _link   = str(_art.get("link",""))
        _date   = str(_art.get("date",""))[:10]
        _src    = _link.split("/")[2].replace("www.","") if "/" in _link else ""
        # Aperçu = les 60 premiers caractères du titre reformulés comme accroche
        _apercu = _title[:60] + "…" if len(_title)>60 else _title
        _cards_html += f"""
        <div class="nw-card">
            <a href="{_link}" target="_blank" style="text-decoration:none;display:block;height:100%;">
                <div class="nw-cat">📰 {_src}</div>
                <div class="nw-title">{_title}</div>
                <div class="nw-apercu">{_apercu}</div>
                <div class="nw-footer">
                    <span class="nw-date">{_date}</span>
                    <span class="nw-arrow">Lire →</span>
                </div>
            </a>
        </div>"""

    _uid = abs(hash(sector_key)) % 99999
    st.markdown(f"""
    <style>
    #nwrap_{_uid} {{
        position:relative;overflow:hidden;border-radius:12px;
        background:#F8FAFC;border:1px solid #E2E8F0;
    }}
    #ntrack_{_uid} {{
        display:flex;transition:transform 0.6s cubic-bezier(.4,0,.2,1);
        will-change:transform;
    }}
    .nw-card {{
        min-width:100%;box-sizing:border-box;
        padding:18px 20px;
        background:white;border-radius:12px;
        border-top:3px solid #00C896;
        cursor:pointer;
    }}
    .nw-card:hover {{ background:#F0FDF9; }}
    .nw-cat {{font-size:10px;font-weight:700;color:#00C896;text-transform:uppercase;
               letter-spacing:1px;margin-bottom:6px;}}
    .nw-title {{font-size:14px;font-weight:700;color:#0B2545;line-height:1.4;margin-bottom:6px;}}
    .nw-apercu {{font-size:12px;color:#4A6080;line-height:1.5;margin-bottom:10px;
                  display:-webkit-box;-webkit-line-clamp:2;
                  -webkit-box-orient:vertical;overflow:hidden;}}
    .nw-footer {{display:flex;justify-content:space-between;align-items:center;
                  border-top:1px solid #F0F4F8;padding-top:8px;}}
    .nw-date {{font-size:10px;color:#8FA3BC;}}
    .nw-arrow {{font-size:11px;font-weight:700;color:#00C896;}}
    .nw-dots {{display:flex;gap:5px;justify-content:center;padding:8px 0 4px;}}
    .nw-dot {{width:6px;height:6px;border-radius:50%;background:#E2E8F0;
               transition:background 0.3s;cursor:pointer;}}
    .nw-dot.active {{background:#00C896;}}
    </style>
    <div id="nwrap_{_uid}">
        <div id="ntrack_{_uid}">{_cards_html}</div>
    </div>
    <div class="nw-dots" id="ndots_{_uid}">
        {"".join(f'<div class="nw-dot{" active" if i==0 else ""}" onclick="nwGo_{_uid}({i})"></div>' for i in range(len(_articles)))}
    </div>
    <script>
    (function(){{
        var cur_{_uid} = 0;
        var n_{_uid}   = {len(_articles)};
        var track_{_uid} = document.getElementById("ntrack_{_uid}");
        var dots_{_uid}  = document.querySelectorAll("#ndots_{_uid} .nw-dot");
        function nwSet_{_uid}(i){{
            cur_{_uid} = i;
            track_{_uid}.style.transform = "translateX(-" + (i*100) + "%)";
            dots_{_uid}.forEach(function(d,j){{
                d.classList.toggle("active", j===i);
            }});
        }}
        window["nwGo_{_uid}"] = function(i){{ nwSet_{_uid}(i); }};
        setInterval(function(){{
            nwSet_{_uid}((cur_{_uid}+1) % n_{_uid});
        }}, 5000);
    }})();
    </script>
    """, unsafe_allow_html=True)



def get_user_plan(username):
    """Retourne le plan de l'utilisateur (depuis Supabase ou USERS_PLAN)."""
    try:
        _prefs = load_user_prefs(username)
        if _prefs and _prefs.get("plan"):
            return _prefs["plan"]
    except Exception:
        pass
    return USERS_PLAN.get(username, "starter")

def can_access(feature, username=None):
    """Vérifie si l'utilisateur peut accéder à une feature selon son plan."""
    if username is None:
        username = st.session_state.get("current_user","")
    plan = get_user_plan(username)
    limits = PLAN_LIMITS.get(plan, PLAN_LIMITS["starter"])
    return bool(limits.get(feature, False))

def show_lock(feature, username=None):
    """Affiche un cadenas avec CTA upgrade si feature non accessible."""
    if username is None:
        username = st.session_state.get("current_user","")
    plan = get_user_plan(username)
    lang = st.session_state.get("language","fr")

    # Trouver le plan requis
    required = "business"
    for p, l in PLAN_LIMITS.items():
        if l.get(feature):
            required = p
            break

    rp = PLAN_LIMITS.get(required, PLAN_LIMITS["business"])
    labels_fr = {
        "terrain":        "Profil Terrain",
        "prediction":     "Prédiction rupture 4 semaines",
        "bfr":            "Alerte BFR",
        "benchmarks":     "Benchmarks sectoriels",
        "news":           "Actualités sectorielles",
        "scoring_detail": "Scoring détaillé",
        "api":            "Accès API",
    }
    labels_en = {
        "terrain":        "Terrain Profile",
        "prediction":     "4-week stockout prediction",
        "bfr":            "BFR / WCR Alert",
        "benchmarks":     "Sector benchmarks",
        "news":           "Sector news",
        "scoring_detail": "Detailed scoring",
        "api":            "API access",
    }
    _labels = labels_en if lang == "en" else labels_fr
    _feat_lbl = _labels.get(feature, feature)

    if lang == "en":
        _msg = f"🔒 **{_feat_lbl}** — available from the **{rp['label']}** plan."
        _cta = "Upgrade → contact@logiflo.io"
    else:
        _msg = f"🔒 **{_feat_lbl}** — disponible à partir du plan **{rp['label']}**."
        _cta = "Upgrader → contact@logiflo.io"

    st.markdown(f"""
    <div style="border:1px solid {rp['color']}40;border-left:3px solid {rp['color']};
                border-radius:8px;padding:12px 16px;background:{rp['bg']};
                margin:8px 0;">
        <div style="font-size:13px;color:#0B2545;margin-bottom:4px;">{_msg}</div>
        <div style="font-size:11px;color:#4A6080;">{_cta}</div>
    </div>
    """, unsafe_allow_html=True)

def audit_counter_sidebar(username, plan):
    """Affiche le compteur d'audits mensuel pour les plans Starter."""
    if PLAN_LIMITS.get(plan,{}).get("audits_mois") is None:
        return True  # illimité

    lang = st.session_state.get("language","fr")
    max_a = PLAN_LIMITS[plan]["audits_mois"]
    # Ne charger les archives que si Supabase est disponible
    # pour éviter de bloquer le démarrage
    try:
        _arc = load_archives_from_sheets(username)
        if _arc is None or _arc.empty:
            used = 0
        else:
            import datetime as _dtc
            this_month = datetime.datetime.now().month
            this_year  = datetime.datetime.now().year
            if "date" in _arc.columns:
                _arc["_m"] = pd.to_datetime(_arc["date"], format="%d/%m/%Y", errors="coerce").dt.month
                _arc["_y"] = pd.to_datetime(_arc["date"], format="%d/%m/%Y", errors="coerce").dt.year
                used = len(_arc[(_arc["_m"]==this_month) & (_arc["_y"]==this_year)])
            elif "created_at" in _arc.columns:
                _arc["_dt2"] = pd.to_datetime(_arc["created_at"], errors="coerce")
                used = len(_arc[(_arc["_dt2"].dt.month==this_month) & (_arc["_dt2"].dt.year==this_year)])
            else:
                used = 0
    except Exception:
        used = 0

    color = "#E8304A" if used >= max_a else ("#F39C12" if used >= max_a*0.6 else "#00C896")
    pct = min(int((used/max_a)*100), 100)

    lbl = f"Audits ce mois" if lang=="fr" else "Audits this month"
    st.sidebar.markdown(f"""
    <div style="padding:8px 12px;background:rgba(255,255,255,0.05);
                border-radius:8px;margin-bottom:8px;">
        <div style="font-size:10px;color:rgba(255,255,255,0.5);margin-bottom:3px;">
            {lbl}
        </div>
        <div style="font-size:16px;font-weight:800;color:{color};
                    font-family:Syne,sans-serif;">{used}/{max_a}</div>
        <div style="height:3px;background:rgba(255,255,255,0.1);
                    border-radius:99px;margin-top:4px;overflow:hidden;">
            <div style="height:100%;width:{pct}%;background:{color};
                        border-radius:99px;"></div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    if used >= max_a:
        lbl_q = "Quota atteint" if lang=="fr" else "Quota reached"
        st.sidebar.warning(lbl_q)
        return False
    return True



# ══════════════════════════════════════════════════════════════════
# BENCHMARKS SECTORIELS
# ══════════════════════════════════════════════════════════════════
SECTORAL_DB = {
    "transport_routier": {
        "keywords": ["routier","camion","transport","trajet","trajet","fret","livraison","tournee","chauffeur","vehicule","traction","semiremorque","porteur","messagerie"],
        "fr": """BENCHMARKS TRANSPORT ROUTIER CNR 2026:
- Cout/km longue distance articule gazole: 1,85-2,10 EUR/km
- Cout/km regional porteur: 1,40-1,65 EUR/km
- Part carburant: ~26,5% du cout total
- Marge nette saine PME transport: 6-10% | Alerte < 6% | Perte < 0%
- Taux de remplissage sain: > 80%
- Temps de conduite max: 9h/jour | 56h/semaine (reglementation EU)""",
        "en": """CNR 2026 ROAD TRANSPORT BENCHMARKS:
- Long-haul articulated cost/km: 1.85-2.10 EUR/km
- Regional rigid truck cost/km: 1.40-1.65 EUR/km
- Fuel share: ~26.5% of total cost
- Healthy net margin SME transport: 6-10% | Alert < 6% | Loss < 0%
- Load factor target: > 80%"""
    },
    "transport_maritime": {
        "keywords": ["maritime","port","navire","conteneur","teu","fcl","lcl","bateaux","shipping","fos","havre","marseille"],
        "fr": """BENCHMARKS TRANSPORT MARITIME 2026:
- Transit Europe-Asie via Suez: 28-35 jours
- Demurrage moyen: 120-180 EUR/jour/conteneur
- Marge transitaire maritime: 15-25%
- Taux de ponctualite armateurs: 55-70%""",
        "en": """MARITIME TRANSPORT BENCHMARKS 2026:
- Europe-Asia transit via Suez: 28-35 days
- Avg demurrage: 120-180 EUR/day/container
- Freight forwarder margin: 15-25%
- Carrier on-time rate: 55-70%"""
    },
    "transport_aerien": {
        "keywords": ["aerien","air","cargo","fret aerien","awb","airline","avion","roissy","cdg","orly"],
        "fr": """BENCHMARKS FRET AERIEN 2026:
- Taux fret Europe-Asie: 2,20-3,80 EUR/kg
- Europe-Amerique du Nord: 2,80-4,50 EUR/kg
- Surcharge carburant FSC: 25-40% du tarif de base
- Marge transitaire aerien: 20-35%""",
        "en": """AIR FREIGHT BENCHMARKS 2026:
- Europe-Asia rate: 2.20-3.80 EUR/kg
- Europe-North America: 2.80-4.50 EUR/kg
- Fuel surcharge FSC: 25-40% of base rate
- Freight forwarder margin air: 20-35%"""
    },
    "stock_industrie": {
        "keywords": ["industrie","manufacturing","usine","production","piece","composant","matiere","cable","machine","outil"],
        "fr": """BENCHMARKS STOCK INDUSTRIEL 2026:
- Taux de service cible: > 97%
- Couverture stock saine: 1-3 mois
- Stock dormant: alerte si > 5% du capital
- Cout possession stock: 18-25% valeur/an
- BFR cible: < 45 jours de CA""",
        "en": """INDUSTRIAL STOCK BENCHMARKS 2026:
- Target service level: > 97%
- Healthy stock coverage: 1-3 months
- Dormant stock: alert if > 5% of capital
- Holding cost: 18-25% value/year
- Target WCR: < 45 days revenue"""
    },
    "stock_distribution": {
        "keywords": ["distribution","grossiste","negoce","commerce","grossiste","import","export","vente"],
        "fr": """BENCHMARKS STOCK DISTRIBUTION 2026:
- Taux de service cible: > 95%
- Couverture stock saine: 1-2 mois
- Rotation stock annuelle saine: > 6 fois
- Stock dormant: alerte si > 8% du capital
- Cout possession: 20-28% valeur/an""",
        "en": """DISTRIBUTION STOCK BENCHMARKS 2026:
- Target service level: > 95%
- Healthy stock coverage: 1-2 months
- Healthy annual inventory turns: > 6x
- Dormant stock: alert if > 8% of capital
- Holding cost: 20-28% value/year"""
    },
    "stock_pharma": {
        "keywords": ["pharma","medicament","lot","dluo","dlc","peremption","sante","medicament","molecule","laboratoire"],
        "fr": """BENCHMARKS STOCK PHARMACEUTIQUE 2026:
- Taux de service: > 99% (critique)
- Gestion des lots FEFO obligatoire
- Alertes peremption: > 6 mois avant DLC
- Temperature stockage: tracabilite obligatoire
- Stock dormant: < 2% des references""",
        "en": """PHARMACEUTICAL STOCK BENCHMARKS 2026:
- Service level: > 99% (critical)
- FEFO lot management mandatory
- Expiry alerts: > 6 months before expiry
- Storage temperature: traceability mandatory
- Dormant stock: < 2% of references"""
    },
    "stock_retail": {
        "keywords": ["retail","magasin","boutique","fashion","mode","soldes","enseigne","rayon","commerce de detail"],
        "fr": """BENCHMARKS STOCK RETAIL / E-COMMERCE 2026:
- Taux de service: > 96%
- Rotation stock fashion: > 4 fois/an
- Stock dormant: alerte si > 6 mois sans mouvement
- Taux de retour e-commerce sain: < 20%
- BFR cible: < 60 jours de CA""",
        "en": """RETAIL / E-COMMERCE STOCK BENCHMARKS 2026:
- Service level: > 96%
- Fashion stock turns: > 4x/year
- Dormant stock: alert if > 6 months no movement
- E-commerce return rate: < 20%
- Target WCR: < 60 days revenue"""
    },
    "stock_agroalim": {
        "keywords": ["alimentaire","agroalimentaire","dlc","dluo","frais","surgele","epicerie","boisson","conserve"],
        "fr": """BENCHMARKS STOCK AGROALIMENTAIRE 2026:
- Taux de service: > 96%
- FEFO obligatoire (First Expired First Out)
- Couverture stock frais: < 5 jours
- Couverture stock sec: < 30 jours
- Perte / casse cible: < 1,5% du CA""",
        "en": """FOOD & BEVERAGE STOCK BENCHMARKS 2026:
- Service level: > 96%
- FEFO mandatory (First Expired First Out)
- Fresh stock coverage: < 5 days
- Dry stock coverage: < 30 days
- Waste / breakage target: < 1.5% of revenue"""
    },
    "stock_btp": {
        "keywords": ["btp","chantier","construction","materiaux","beton","acier","menuiserie","electricite","plomberie"],
        "fr": """BENCHMARKS STOCK BTP / CHANTIER 2026:
- Taux de service cible: > 95%
- Couverture stock: 2-4 semaines chantier
- Perte et casse sur chantier: < 3%
- Gestion FIFO imperativo
- Stock dormant: alerte si > 4 mois""",
        "en": """CONSTRUCTION / SITE STOCK BENCHMARKS 2026:
- Target service level: > 95%
- Stock coverage: 2-4 weeks per site
- Site waste and breakage: < 3%
- FIFO management mandatory
- Dormant stock: alert if > 4 months"""
    },
    "transport_maritime_intl": {
        "keywords": ["container","conteneur","teu","fcl","lcl","bl","vessel","shipping","freight","port","ocean"],
        "fr": """BENCHMARKS TRANSPORT MARITIME INTERNATIONAL 2026 (IRU/UNCTAD/Drewry):
- Taux fret 20' Europe-Asie: 800-2500 USD/TEU
- Corridors Mediterranee-Afrique Ouest: 1200-2800 USD/TEU
- Demurrage Marseille/Le Havre: 120-180 EUR/jour/conteneur
- Ponctualite armateurs 2026: 55-70%
- Transit Europe-Asie via Suez: 28-35j | Cape: 38-45j""",
        "en": """INTERNATIONAL MARITIME BENCHMARKS 2026:
- 20' container Europe-Asia: 800-2500 USD/TEU
- Mediterranean-West Africa: 1200-2800 USD/TEU
- Demurrage Marseille/Le Havre: 120-180 EUR/day/container
- Carrier on-time 2026: 55-70%
- Europe-Asia via Suez: 28-35d | Cape: 38-45d"""
    },
    "transport_aerien_intl": {
        "keywords": ["airfreight","air cargo","awb","iata","airline","express","dhl","fedex","ups"],
        "fr": """BENCHMARKS FRET AERIEN INTERNATIONAL 2026 (IATA):
- Taux Europe-Asie: 2,20-3,80 EUR/kg
- Europe-Amerique du Nord: 2,80-4,50 EUR/kg
- Europe-Afrique: 2,50-4,20 EUR/kg
- Surcharge carburant FSC: 25-40%
- Marge transitaire aerien: 20-35%""",
        "en": """INTERNATIONAL AIR FREIGHT BENCHMARKS 2026 (IATA):
- Europe-Asia: 2.20-3.80 EUR/kg
- Europe-North America: 2.80-4.50 EUR/kg
- Europe-Africa: 2.50-4.20 EUR/kg
- Fuel surcharge FSC: 25-40%
- Freight forwarder margin: 20-35%"""
    },
    "transport_routier_eu": {
        "keywords": ["international","europe","cross-border","export","import","incoterm","douane"],
        "fr": """BENCHMARKS TRANSPORT ROUTIER EUROPEEN 2026 (IRU/Eurostat):
- Taux France-Espagne (FTL 33T): 1800-2400 EUR/trajet
- Taux France-Allemagne: 1600-2200 EUR/trajet
- Taux France-Maroc (via ferry): 3200-4500 EUR/trajet
- Cout/km EU: 1,85-2,30 EUR/km
- Surcoút EuroVignette/peages: +8-12%""",
        "en": """EUROPEAN ROAD TRANSPORT BENCHMARKS 2026 (IRU/Eurostat):
- France-Spain FTL rate: 1800-2400 EUR/trip
- France-Germany: 1600-2200 EUR/trip
- France-Morocco (via ferry): 3200-4500 EUR/trip
- FTL cost/km Europe: 1.85-2.30 EUR/km
- EuroVignette/tolls surcharge: +8-12%"""
    },
    "supply_chain_maghreb": {
        "keywords": ["maroc","morocco","algerie","tunisie","casablanca","rabat","maghreb"],
        "fr": """BENCHMARKS SUPPLY CHAIN MAGHREB 2026:
- LPI Maroc: 3,2/5 (Banque Mondiale)
- Transport Casablanca-Agadir: 1800-2500 MAD/trajet
- Taux service PME distribution Maroc: 85-92%
- Delai dedouanement: 3-7 jours
- BFR PME marocaine: 60-90 jours de CA""",
        "en": """MAGHREB SUPPLY CHAIN BENCHMARKS 2026:
- Morocco LPI: 3.2/5 (World Bank)
- Transport Casablanca-Agadir: 1800-2500 MAD/trip
- SME service level Morocco: 85-92%
- Customs clearance: 3-7 days
- WCR SME Morocco: 60-90 days revenue"""
    },
    "supply_chain_afrique": {
        "keywords": ["afrique","africa","cote d'ivoire","ivory coast","senegal","abidjan","dakar","ghana","nigeria"],
        "fr": """BENCHMARKS SUPPLY CHAIN AFRIQUE 2026 (Banque Mondiale/CEDEAO):
- LPI Cote d'Ivoire: 3,1/5 | Senegal: 2,9/5
- Transport Abidjan-Bouake: 180000-250000 FCFA
- Taux service distribution: 75-88%
- Delai dedouanement: 5-12 jours
- Taux pertes transit: 2-5%""",
        "en": """SUB-SAHARAN AFRICA SUPPLY CHAIN BENCHMARKS 2026:
- Ivory Coast LPI: 3.1/5 | Senegal: 2.9/5
- Abidjan-Bouake transport: 180000-250000 FCFA
- Distribution service level: 75-88%
- Customs clearance: 5-12 days
- Transit loss rate: 2-5%"""
    },
    "generique": {
        "keywords": [],
        "fr": """BENCHMARKS GENERIQUES SUPPLY CHAIN 2026:
- Taux de service B2B minimum: > 93% | B2C: > 96%
- Cout possession stock: 18-28% valeur/an
- Rotation stock annuelle saine: > 4 fois/an
- Stock dormant: alerte si > 10% sans mouvement
- Marge transport saine: > 6%
- BFR cible: < 60 jours de CA""",
        "en": """GENERIC SUPPLY CHAIN BENCHMARKS 2026:
- B2B service level minimum: > 93% | B2C: > 96%
- Stock holding cost: 18-28% value/year
- Healthy annual turns: > 4x/year
- Dormant stock: alert if > 10% no movement
- Healthy transport margin: > 6%
- Target WCR: < 60 days revenue"""
    }
}

def detect_sector(df=None, module="stock", mode_detected=None):
    """Detecte le secteur pertinent selon le fichier, le module et le contexte geo."""
    if module == "transport":
        if mode_detected:
            m = str(mode_detected).lower()
            if "maritime" in m or "sea" in m or "ocean" in m:
                # Vérifier si international
                if df is not None:
                    all_t = " ".join([str(v).lower() for v in df.values.flatten()[:100]])
                    intl_kw = ["rotterdam","anvers","hambourg","barcelona","algeciras",
                               "tanger","abidjan","dakar","casablanca","shanghai","container"]
                    if any(k in all_t for k in intl_kw):
                        return "transport_maritime_intl"
                return "transport_maritime"
            if "aerien" in m or "air" in m or "cargo" in m:
                return "transport_aerien_intl"
            # Vérifier si routier européen
            if df is not None:
                all_t = " ".join([str(v).lower() for v in df.values.flatten()[:100]])
                eu_kw = ["espagne","spain","allemagne","germany","italie","italy",
                         "belgique","netherlands","maroc","morocco","export","import"]
                if any(k in all_t for k in eu_kw):
                    return "transport_routier_eu"
        return "transport_routier"
    if df is not None:
        all_text = " ".join([str(c).lower() for c in df.columns])
        if len(df) > 0:
            all_text += " " + " ".join(df.iloc[:,0].astype(str).str.lower().head(30).tolist())
            # Aussi analyser les valeurs pour la détection géo
            all_vals = " ".join([str(v).lower() for v in df.values.flatten()[:200]])
            all_text += " " + all_vals
        scores = {}
        for sk, sd in SECTORAL_DB.items():
            if sk in ("generique","transport_maritime_intl","transport_aerien_intl",
                       "transport_routier_eu","supply_chain_maghreb","supply_chain_afrique"):
                continue
            hits = sum(1 for kw in sd["keywords"] if kw in all_text)
            if hits >= 2: scores[sk] = hits
        if scores: return max(scores, key=scores.get)
        # Détection géographique pour supply chain internationale
        maghreb_kw = ["maroc","morocco","algerie","casablanca","rabat","tanger","tunis"]
        afrique_kw = ["abidjan","dakar","cote d'ivoire","ivory coast","accra","ghana","nigeria"]
        if any(k in all_text for k in afrique_kw): return "supply_chain_afrique"
        if any(k in all_text for k in maghreb_kw): return "supply_chain_maghreb"
    return "generique"

def get_sector_benchmarks(sector_key, lang="fr"):
    """Retourne les benchmarks du secteur dans la bonne langue."""
    s = SECTORAL_DB.get(sector_key, SECTORAL_DB["generique"])
    return s.get(lang, s.get("fr",""))



# USERS_DB chargé depuis st.secrets pour ne pas exposer les credentials
def _load_users():
    try:
        raw = st.secrets.get("USERS_DB", {})
        if isinstance(raw, dict) and raw:
            return dict(raw)
    except Exception:
        pass
    # Fallback local uniquement pour dev — retirer en prod
    return {
        "eric":"logiflo2026","admin":"admin123","demo_client1":"audit2026",
        "demo_client2":"test2026","jury":"pitch2026","partenaire":"partner2026","test":"test123",
    }
USERS_DB = _load_users()

# =========================================
# 0.1 TRADUCTIONS
# =========================================
T = {
    "fr": {
        "nav_dashboard":"Tableau de bord","nav_compte":"Mon Compte","nav_workspace":"Espace de Travail","nav_archives":"Archives",
        "nav_params":"Paramètres","nav_legal":"Informations Légales","nav_logout":"Déconnexion",
        "home_title":"LOGIFLO.IO",
        "home_sub":"Plateforme d'Intelligence Logistique et d'Optimisation Financière",
        "home_stock":"AUDIT STOCKS","home_transport":"AUDIT TRANSPORT",
        "home_access":"DEMANDER UN ACCÈS PRIVÉ",
        "login_id":"Identifiant","login_pw":"Mot de passe","login_btn":"Connexion",
        "login_err":"Identifiants incorrects.","login_back":"← Retour",
        "profile_title":"Sélectionnez votre Espace de Travail",
        "profile_sub":"L'interface s'adaptera à vos habilitations.",
        "profile_mgr":"PROFIL MANAGER (Stratégie & Finance)",
        "profile_ops":"PROFIL TERRAIN (Action Opérationnelle)",
        "stock_title":"📦 Audit Financier des Stocks",
        "stock_import":"📥 Importation Sécurisée",
        "stock_import_sub":"Déposez votre fichier d'inventaire (CSV ou Excel).<br>Le <b>Smart Ingester™ V4</b> détecte automatiquement vos colonnes, même avec des noms atypiques.<br><span style='color:#00A87A;font-weight:600;'>✓ Prix optionnel &nbsp; ✓ Historique optionnel &nbsp; ✓ Tous formats</span>",
        "stock_kpi_capital":"Capital Immobilisé","stock_kpi_articles":"Articles en Stock",
        "stock_kpi_service":"Taux de Service","stock_kpi_rupture":"Articles en Rupture",
        "stock_btn_ia":"GÉNÉRER L'AUDIT FINANCIER (IA)","stock_btn_ia_terrain":"GÉNÉRER L'AUDIT IA",
        "stock_btn_save":"💾 Sauvegarder","stock_btn_dl":"📥 Télécharger le Rapport (PDF)",
        "stock_badge_no_price":"📊 Mode opérationnel — analyse sans prix",
        "stock_badge_conso":"📈 Historique de consommation détecté",
        "stock_badge_no_conso":"⚠️ Pas d'historique — couverture non calculable",
        "stock_saved":"✅ Sauvegardé !","stock_save_err":"⚠️ Connexion Supabase (PostgreSQL) absente.",
        "stock_urgent":"🚨 Priorités immédiates","stock_full":"📋 Stock complet",
        "stock_no_rupture":"✅ Aucun article en rupture.",
        "trans_title":"🚚 Audit de Rentabilité Transport",
        "trans_import":"🌍 Importation des Flux de Transport",
        "trans_import_sub":"Déposez votre fichier TMS ou Excel. Le moteur <b>ORS</b> calcule les distances routières réelles.<br><span style='color:#00A87A;font-weight:600;'>✓ Maritime &nbsp; ✓ Aérien &nbsp; ✓ Routier &nbsp; ✓ Ferroviaire</span>",
        "trans_kpi_marge":"Marge Nette Globale","trans_kpi_taux":"Taux de Rentabilité",
        "trans_kpi_fuite":"🚨 Fuite de Marge","trans_kpi_sain":"✅ Réseau",
        "trans_btn_ia":"GÉNÉRER L'AUDIT DE RENTABILITÉ (IA)",
        "trans_btn_save":"💾 Sauvegarder","trans_btn_dl":"📥 Télécharger le Rapport (PDF)",
        "trans_tab_top":"🎯 Top 15 — Pires trajets","trans_tab_all":"🗺️ Vue d'ensemble",
        "trans_ca_miss":"💡 CA manquant — estimé à marge 15%.",
        "trans_no_cost":"🚨 Colonne 'Coût' introuvable.",
        "trans_top15_title":"Top 15 trajets les plus déficitaires",
        "trans_scatter_title":"Vue d'ensemble — Rentabilité vs CA par trajet",
        "trans_seuil_zero":"Seuil zéro","trans_seuil_alert":"Seuil alerte 10%",
        "trans_detail":"Détail des trajets en alerte",
        "trans_col_client":"Client / Trajet","trans_col_ca":"CA (€)",
        "trans_col_co":"Coût (€)","trans_col_marge":"Marge (€)","trans_col_pct":"Marge (%)",
        "arch_title":"🗄️ Archives & Historique",
        "arch_empty":"Aucun audit archivé. Générez votre premier audit depuis l'Espace de Travail.",
        "arch_dl":"📥 PDF","arch_filter":"Filtrer","arch_filter_all":"Tous",
        "arch_show":"audit(s) affiché(s)","arch_resume":"📋 Résumé IA",
        "step_read":"Lecture du fichier...","step_detect":"Détection des colonnes...",
        "step_calc":"Calcul des indicateurs...","step_ia":"Analyse IA en cours...",
        "step_report":"Génération du rapport...","step_geo":"Géocodage des villes...",
        "step_dist":"Calcul des distances ORS...","step_mode":"Détection du mode de transport...",
        "pdf_title_stock":"AUDIT STRATEGIQUE DES STOCKS",
        "pdf_title_trans":"AUDIT FINANCIER TRANSPORT",
        "pdf_confidential":"CONFIDENTIEL","pdf_strategic":"AUDIT STRATEGIQUE",
        "pdf_report":"RAPPORT D ANALYSE","pdf_date":"Date",
        "pdf_footer":"Document genere par Logiflo.io. Recommandations a titre indicatif.",
        "pdf_cta":"Ce rapport a ete genere par LOGIFLO.IO\nConcu par un logisticien terrain — pas par un consultant.\nPour aller plus loin : contact@logiflo.io | logiflo-io.streamlit.app",
        "mode_detected":"— analyse adaptée activée",
        "change_profile":"Changer de profil","active_profile":"Profil Actif",
        "params_title":"⚙️ Configuration des Seuils",
        "params_alert":"Seuil d'Alerte","params_rupture":"Seuil de Rupture Critique",
        "params_km":"Seuil Rentabilité EUR/KM",
        "contact_title":"Demande d'Accès Réservé",
        "contact_name":"Nom & Prénom","contact_email":"Email Professionnel",
        "contact_company":"Entreprise","contact_volume":"Volume géré :",
        "contact_issue":"Enjeu prioritaire :","contact_btn":"Transmettre",
        "contact_ok":"Demande transmise. Notre équipe vous contactera sous 24h.",
        "vol1":"Moins de 10M EUR","vol2":"De 10M à 50M EUR","vol3":"Plus de 50M EUR",
        "iss1":"Optimisation BFR (Stocks)","iss2":"Réduction coûts Transport","iss3":"Global Supply Chain",
    },
    "en": {
        "nav_dashboard":"Dashboard","nav_compte":"My Account","nav_workspace":"Workspace","nav_archives":"Archives",
        "nav_params":"Settings","nav_legal":"Legal Information","nav_logout":"Log out",
        "home_title":"LOGIFLO.IO",
        "home_sub":"Logistics Intelligence & Financial Optimization Platform",
        "home_stock":"STOCK AUDIT","home_transport":"TRANSPORT AUDIT",
        "home_access":"REQUEST PRIVATE ACCESS",
        "login_id":"Username","login_pw":"Password","login_btn":"Sign in",
        "login_err":"Incorrect credentials.","login_back":"← Back",
        "profile_title":"Select your Workspace",
        "profile_sub":"The interface will adapt to your permissions.",
        "profile_mgr":"MANAGER PROFILE (Strategy & Finance)",
        "profile_ops":"OPERATIONS PROFILE (Field Action)",
        "stock_title":"📦 Stock Financial Audit",
        "stock_import":"📥 Secure Import",
        "stock_import_sub":"Drop your inventory file (CSV or Excel).<br>The <b>Smart Ingester™ V4</b> automatically detects your columns, even with unusual names.<br><span style='color:#00A87A;font-weight:600;'>✓ Price optional &nbsp; ✓ History optional &nbsp; ✓ All formats</span>",
        "stock_kpi_capital":"Tied-up Capital","stock_kpi_articles":"Items in Stock",
        "stock_kpi_service":"Service Level","stock_kpi_rupture":"Stock-outs",
        "stock_btn_ia":"GENERATE FINANCIAL AUDIT (AI)","stock_btn_ia_terrain":"GENERATE AI AUDIT",
        "stock_btn_save":"💾 Save","stock_btn_dl":"📥 Download Report (PDF)",
        "stock_badge_no_price":"📊 Operational mode — analysis without prices",
        "stock_badge_conso":"📈 Consumption history detected",
        "stock_badge_no_conso":"⚠️ No history — coverage not calculable",
        "stock_saved":"✅ Saved!","stock_save_err":"⚠️ Google Sheets connection unavailable.",
        "stock_urgent":"🚨 Immediate Priorities","stock_full":"📋 Full Inventory",
        "stock_no_rupture":"✅ No stock-outs detected.",
        "trans_title":"🚚 Transport Profitability Audit",
        "trans_import":"🌍 Import Transport Flows",
        "trans_import_sub":"Drop your TMS or Excel file. The <b>ORS</b> engine computes real road distances.<br><span style='color:#00A87A;font-weight:600;'>✓ Maritime &nbsp; ✓ Air &nbsp; ✓ Road &nbsp; ✓ Rail</span>",
        "trans_kpi_marge":"Total Net Margin","trans_kpi_taux":"Profitability Rate",
        "trans_kpi_fuite":"🚨 Margin Leak","trans_kpi_sain":"✅ Network",
        "trans_btn_ia":"GENERATE PROFITABILITY AUDIT (AI)",
        "trans_btn_save":"💾 Save","trans_btn_dl":"📥 Download Report (PDF)",
        "trans_tab_top":"🎯 Top 15 — Worst routes","trans_tab_all":"🗺️ Overview",
        "trans_ca_miss":"💡 Revenue missing — estimated at 15% margin.",
        "trans_no_cost":"🚨 'Cost' column not found.",
        "trans_top15_title":"Top 15 most unprofitable routes",
        "trans_scatter_title":"Overview — Profitability vs Revenue per route",
        "trans_seuil_zero":"Break-even","trans_seuil_alert":"Alert threshold 10%",
        "trans_detail":"Underperforming routes — detail",
        "trans_col_client":"Client / Route","trans_col_ca":"Revenue (€)",
        "trans_col_co":"Cost (€)","trans_col_marge":"Margin (€)","trans_col_pct":"Margin (%)",
        "arch_title":"🗄️ Archives & History",
        "arch_empty":"No saved audits yet. Generate your first audit from the Workspace.",
        "arch_dl":"📥 PDF","arch_filter":"Filter","arch_filter_all":"All",
        "arch_show":"audit(s) shown","arch_resume":"📋 AI Summary",
        "step_read":"Reading file...","step_detect":"Detecting columns...",
        "step_calc":"Computing indicators...","step_ia":"AI analysis in progress...",
        "step_report":"Generating report...","step_geo":"Geocoding cities...",
        "step_dist":"Computing ORS distances...","step_mode":"Detecting transport mode...",
        "pdf_title_stock":"STRATEGIC STOCK AUDIT",
        "pdf_title_trans":"TRANSPORT FINANCIAL AUDIT",
        "pdf_confidential":"CONFIDENTIAL","pdf_strategic":"STRATEGIC AUDIT",
        "pdf_report":"ANALYSIS REPORT","pdf_date":"Date",
        "pdf_footer":"Generated by Logiflo.io. Recommendations are indicative only.",
        "pdf_cta":"This report was generated by LOGIFLO.IO\nDesigned by a field logistics professional — not a consultant.\nTo go further: contact@logiflo.io | logiflo-io.streamlit.app",
        "mode_detected":"— adapted analysis activated",
        "change_profile":"Change profile","active_profile":"Active Profile",
        "params_title":"⚙️ Threshold Configuration",
        "params_alert":"Alert Threshold","params_rupture":"Critical Stock-out Threshold",
        "params_km":"Profitability Threshold EUR/KM",
        "contact_title":"Request Private Access",
        "contact_name":"Full Name","contact_email":"Professional Email",
        "contact_company":"Company","contact_volume":"Managed volume:",
        "contact_issue":"Main challenge:","contact_btn":"Submit",
        "contact_ok":"Request submitted. Our team will contact you within 24h.",
        "vol1":"Less than 10M EUR","vol2":"10M to 50M EUR","vol3":"More than 50M EUR",
        "iss1":"BFR Optimization (Stock)","iss2":"Transport Cost Reduction","iss3":"Global Supply Chain",
    }
}

def _(key):
    lang = st.session_state.get("language","fr")
    return T.get(lang,T["fr"]).get(key, T["fr"].get(key, key))

# =========================================
# 0.2 SUPABASE (remplace Google Sheets)
# =========================================
def get_supabase():
    """Client Supabase — fallback silencieux si non configuré."""
    try:
        url = st.secrets.get("SUPABASE_URL","")
        key = st.secrets.get("SUPABASE_KEY","")
        if url and key and _supa_create:
            return _supa_create(url, key)
    except Exception:
        pass
    return None

def _supa_set_user(client, username):
    """Injecte le user_id pour le Row Level Security."""
    try:
        client.postgrest.auth(st.secrets.get("SUPABASE_KEY",""))
    except Exception:
        pass

def save_audit_to_sheets(username, module, nb_lignes, kpis, labels,
                          resume_ia, pdf_bytes):
    """Sauvegarde un audit dans Supabase (compatible ancien nom)."""
    sb = get_supabase()
    if not sb:
        return _save_audit_sheets_fallback(username, module, nb_lignes,
                                           kpis, labels, resume_ia, pdf_bytes)
    try:
        now = datetime.datetime.now()
        pdf_b64 = base64.b64encode(pdf_bytes).decode("utf-8") if pdf_bytes else ""
        sb.table("audits").insert({
            "user_id":     username,
            "created_at":  now.isoformat(),
            "module":      module,
            "nb_lignes":   int(nb_lignes),
            "kpi_1":       round(float(kpis[0]),2) if len(kpis)>0 else 0,
            "kpi_2":       round(float(kpis[1]),2) if len(kpis)>1 else 0,
            "kpi_3":       round(float(kpis[2]),2) if len(kpis)>2 else 0,
            "kpi_label_1": labels[0] if len(labels)>0 else "",
            "kpi_label_2": labels[1] if len(labels)>1 else "",
            "kpi_label_3": labels[2] if len(labels)>2 else "",
            "resume_ia":   (resume_ia or "")[:800],
            "pdf_base64":  pdf_b64,
        }).execute()
        return True
    except Exception as _e_supa:
        # Log silencieux — ne pas tomber sur Sheets si Supabase répond mais échoue
        # Uniquement si Supabase est absent (sb=None), le fallback Sheets est utilisé
        pass
    return False

def load_archives_from_sheets(username):
    """Charge les archives depuis Supabase (compatible ancien nom)."""
    sb = get_supabase()
    if not sb:
        return _load_archives_sheets_fallback(username)
    try:
        resp = (sb.table("audits")
                  .select("*")
                  .eq("user_id", username)
                  .order("created_at", desc=False)
                  .execute())
        if not resp.data:
            return pd.DataFrame()
        df = pd.DataFrame(resp.data)
        # Normaliser les colonnes pour compatibilité avec le reste du code
        if "created_at" in df.columns:
            df["date"]  = pd.to_datetime(df["created_at"]).dt.strftime("%d/%m/%Y")
            df["heure"] = pd.to_datetime(df["created_at"]).dt.strftime("%H:%M")
        return df
    except Exception:
        return _load_archives_sheets_fallback(username)

# ── Prefs utilisateur ─────────────────────────────────────────────
def load_user_prefs(username):
    """Charge les préférences utilisateur depuis Supabase."""
    sb = get_supabase()
    if not sb:
        return {}
    try:
        resp = (sb.table("user_prefs")
                  .select("*")
                  .eq("user_id", username)
                  .execute())
        if resp.data:
            return resp.data[0]
        return {}
    except Exception:
        return {}

def save_user_prefs(username, prefs_dict):
    """Sauvegarde ou met à jour les prefs utilisateur."""
    sb = get_supabase()
    if not sb:
        return False
    try:
        prefs_dict["user_id"] = username
        prefs_dict["updated_at"] = datetime.datetime.now().isoformat()
        sb.table("user_prefs").upsert(prefs_dict).execute()
        return True
    except Exception:
        return False

# ── Fallbacks Google Sheets (si Supabase indisponible) ────────────
def _save_audit_sheets_fallback(username, module, nb_lignes, kpis,
                                 labels, resume_ia, pdf_bytes):
    try:
        gc_f = get_gsheet_client()
        if not gc_f or not SHEET_ID: return False
        sh_f = gc_f.open_by_key(SHEET_ID)
        try: ws_f = sh_f.worksheet(username)
        except: ws_f = sh_f.add_worksheet(title=username,rows=1000,cols=12)
        now_f = datetime.datetime.now()
        ws_f.append_row([now_f.strftime("%d/%m/%Y"),now_f.strftime("%H:%M"),
            module,nb_lignes,
            round(kpis[0],2) if len(kpis)>0 else "",
            round(kpis[1],2) if len(kpis)>1 else "",
            round(kpis[2],2) if len(kpis)>2 else "",
            labels[0] if len(labels)>0 else "",
            labels[1] if len(labels)>1 else "",
            labels[2] if len(labels)>2 else "",
            (resume_ia or "")[:800],
            base64.b64encode(pdf_bytes).decode("utf-8") if pdf_bytes else ""])
        return True
    except: return False

def _load_archives_sheets_fallback(username):
    try:
        gc_f = get_gsheet_client()
        if not gc_f or not SHEET_ID: return None
        sh_f = gc_f.open_by_key(SHEET_ID)
        try: ws_f = sh_f.worksheet(username)
        except: return pd.DataFrame()
        records = ws_f.get_all_records()
        return pd.DataFrame(records) if records else pd.DataFrame()
    except: return None

def get_gsheet_client():
    try:
        creds=Credentials.from_service_account_info(
            dict(st.secrets["gcp_service_account"]),
            scopes=["https://www.googleapis.com/auth/spreadsheets",
                    "https://www.googleapis.com/auth/drive"])
        return gspread.authorize(creds)
    except: return None

def get_historique_audits(username, module, n=6, current_kpis=None, current_labels=None):
    """
    Charge les n derniers audits du même module depuis Google Sheets.
    Retourne un dict avec les tendances calculées ou None si pas d'historique.
    """
    try:
        df = load_archives_from_sheets(username)
        if df is None or df.empty:
            return None
        # Filtrer par module
        df = df[df["module"] == module].copy()
        if len(df) < 2:
            return None  # pas assez d'historique pour une tendance

        # Convertir les colonnes numériques
        for col in ["kpi_1","kpi_2","kpi_3"]:
            df[col] = pd.to_numeric(df[col], errors="coerce")

        # Trier par date/heure descendant, prendre les n derniers
        try:
            df["_dt"] = pd.to_datetime(df["date"] + " " + df["heure"],
                                        format="%d/%m/%Y %H:%M", errors="coerce")
            df = df.sort_values("_dt", ascending=False)
        except:
            pass  # si parse date échoue, on garde l'ordre sheets

        recent = df.head(n).iloc[::-1]  # remettre chronologique

        # Construire l'historique depuis les archives
        history = []
        for _hr, row in recent.iterrows():
            entry = {
                "date":   row.get("date","?"),
                "kpi_1":  row.get("kpi_1", 0),
                "kpi_2":  row.get("kpi_2", 0),
                "kpi_3":  row.get("kpi_3", 0),
                "label_1":row.get("kpi_label_1","KPI1"),
                "label_2":row.get("kpi_label_2","KPI2"),
                "label_3":row.get("kpi_label_3","KPI3"),
                "resume": str(row.get("resume_ia",""))[:400],
            }
            history.append(entry)

        # Ajouter l'audit ACTUEL (non encore sauvegardé) comme dernier point
        if current_kpis and len(current_kpis) >= 2:
            import datetime as _dt
            current_labels_safe = current_labels or ["KPI1","KPI2","KPI3"]
            history.append({
                "date":    _dt.date.today().strftime("%d/%m/%Y"),
                "kpi_1":  float(current_kpis[0]) if len(current_kpis) > 0 else 0,
                "kpi_2":  float(current_kpis[1]) if len(current_kpis) > 1 else 0,
                "kpi_3":  float(current_kpis[2]) if len(current_kpis) > 2 else 0,
                "label_1": current_labels_safe[0] if len(current_labels_safe) > 0 else "KPI1",
                "label_2": current_labels_safe[1] if len(current_labels_safe) > 1 else "KPI2",
                "label_3": current_labels_safe[2] if len(current_labels_safe) > 2 else "KPI3",
                "resume":  "",
            })

        # Si un audit courant est fourni, l'ajouter même si historique vide
        if current_kpis and len(current_kpis) >= 2:
            import datetime as _dt2
            _current_labels_safe = current_labels or ["KPI1","KPI2","KPI3"]
            history.append({
                "date":    _dt2.date.today().strftime("%d/%m/%Y"),
                "kpi_1":  float(current_kpis[0]) if len(current_kpis) > 0 else 0,
                "kpi_2":  float(current_kpis[1]) if len(current_kpis) > 1 else 0,
                "kpi_3":  float(current_kpis[2]) if len(current_kpis) > 2 else 0,
                "label_1": _current_labels_safe[0] if len(_current_labels_safe) > 0 else "KPI1",
                "label_2": _current_labels_safe[1] if len(_current_labels_safe) > 1 else "KPI2",
                "label_3": _current_labels_safe[2] if len(_current_labels_safe) > 2 else "KPI3",
                "resume":  "",
            })

        if len(history) < 2:
            return None

        # Calculer les tendances entre le plus ancien et le plus récent
        first = history[0]
        last  = history[-1]

        def delta_pct(new, old):
            try:
                new, old = float(new), float(old)
                if old == 0: return None
                return round((new - old) / abs(old) * 100, 1)
            except:
                return None

        def tendance_label(d, lang="fr", invert=False):
            """invert=True : une baisse est bonne (ex: nb ruptures)"""
            if d is None: return ""
            if lang == "en":
                if invert:
                    return f"{'▼ improving' if d < 0 else '▲ worsening'} ({abs(d):.1f}%)"
                return f"{'▲ up' if d > 0 else '▼ down'} ({abs(d):.1f}%)"
            else:
                if invert:
                    return f"{'▼ en amelioration' if d < 0 else '▲ en degradation'} ({abs(d):.1f}%)"
                return f"{'▲ hausse' if d > 0 else '▼ baisse'} ({abs(d):.1f}%)"

        d1 = delta_pct(last["kpi_1"], first["kpi_1"])
        d2 = delta_pct(last["kpi_2"], first["kpi_2"])
        d3 = delta_pct(last["kpi_3"], first["kpi_3"])

        return {
            "history":   history,
            "n_audits":  len(history),
            "first_date":first["date"],
            "last_date": last["date"],
            "delta_1":   d1,
            "delta_2":   d2,
            "delta_3":   d3,
        }
    except Exception:
        return None


def format_historique_pour_prompt(hist, module, lang="fr"):
    """
    Formate l'historique en texte structuré pour injection dans le prompt IA.
    Adapte le vocabulaire selon le module et la langue.
    """
    if not hist:
        return ""

    h = hist["history"]
    n = hist["n_audits"]

    if lang == "en":
        lines = [f"\n=== HISTORICAL TREND — last {n} audits ==="]
        lines.append(f"Period: {hist['first_date']} → {hist['last_date']}")
        lines.append("")

        for i, entry in enumerate(h):
            tag = "CURRENT" if i == len(h)-1 else f"Audit {i+1}"
            l1 = entry["label_1"][:20]
            l2 = entry["label_2"][:20]
            l3 = entry["label_3"][:20]
            lines.append(f"[{tag} — {entry['date']}]")
            lines.append(f"  {l1}: {entry['kpi_1']:.1f} | {l2}: {entry['kpi_2']:.1f} | {l3}: {entry['kpi_3']:.1f}")
            if entry["resume"] and i == len(h)-1:
                pass  # ne pas répéter le résumé actuel

        lines.append("")
        lines.append("COMPUTED TRENDS (first → last):")

        d1,d2,d3 = hist["delta_1"], hist["delta_2"], hist["delta_3"]

        if module == "transport":
            if d1 is not None:
                direction = "improving" if d1 > 0 else "declining"
                lines.append(f"  Net margin: {direction} ({d1:+.1f}%)")
            if d2 is not None:
                direction = "improving" if d2 > 0 else "declining"
                lines.append(f"  Profitability rate: {direction} ({d2:+.1f}%)")
            if d3 is not None:
                direction = "improving" if d3 < 0 else "worsening"
                lines.append(f"  Toxic routes: {direction} ({d3:+.1f}%)")
            lines.append("")
            lines.append("INSTRUCTIONS FOR THIS ANALYSIS:")
            lines.append("- Compare current figures to this historical trend")
            lines.append("- If margin is declining: identify root cause (new routes? lost client? fuel?)")
            lines.append("- If toxic routes increasing: flag as structural risk, not anomaly")
            lines.append("- Mention explicit trend in your PROFITABILITY AUDIT section")
            lines.append("- If trend reversal detected: highlight it as a positive signal")

        elif module in ("stock","terrain"):
            if d1 is not None:
                lines.append(f"  Capital/Items: {d1:+.1f}% vs first audit")
            if d2 is not None:
                direction = "improving" if d2 > 0 else "declining"
                lines.append(f"  Service level: {direction} ({d2:+.1f}%)")
            if d3 is not None:
                direction = "worsening" if d3 > 0 else "improving"
                lines.append(f"  Stock-outs: {direction} ({d3:+.1f}%)")
            lines.append("")
            lines.append("INSTRUCTIONS FOR THIS ANALYSIS:")
            lines.append("- Compare current figures to this historical trend")
            lines.append("- If service level declining: flag as urgent priority")
            lines.append("- If stock-outs increasing: identify if structural or seasonal")
            lines.append("- If dormant stock growing: estimate cash impact over trend period")
            lines.append("- Mention trend explicitly in your OPERATIONAL DIAGNOSIS section")

        lines.append("=== END HISTORICAL DATA ===\n")

    else:
        lines = [f"\n=== TENDANCE HISTORIQUE — {n} derniers audits ==="]
        lines.append(f"Periode : {hist['first_date']} -> {hist['last_date']}")
        lines.append("")

        for i, entry in enumerate(h):
            tag = "ACTUEL" if i == len(h)-1 else f"Audit {i+1}"
            l1 = entry["label_1"][:25]
            l2 = entry["label_2"][:25]
            l3 = entry["label_3"][:25]
            lines.append(f"[{tag} — {entry['date']}]")
            lines.append(f"  {l1}: {entry['kpi_1']:.1f} | {l2}: {entry['kpi_2']:.1f} | {l3}: {entry['kpi_3']:.1f}")

        lines.append("")
        lines.append("TENDANCES CALCULEES (premier -> dernier audit) :")

        d1,d2,d3 = hist["delta_1"], hist["delta_2"], hist["delta_3"]

        if module == "transport":
            if d1 is not None:
                sens = "en hausse" if d1 > 0 else "en baisse"
                lines.append(f"  Marge nette : {sens} ({d1:+.1f}%)")
            if d2 is not None:
                sens = "en hausse" if d2 > 0 else "en baisse"
                lines.append(f"  Taux de rentabilite : {sens} ({d2:+.1f}%)")
            if d3 is not None:
                sens = "en hausse" if d3 > 0 else "en baisse"
                lines.append(f"  Trajets toxiques : {sens} ({d3:+.1f}%)")
            lines.append("")
            lines.append("INSTRUCTIONS POUR CETTE ANALYSE :")
            lines.append("- Compare les chiffres actuels a cette tendance historique")
            lines.append("- Si la marge baisse : identifie la cause racine (nouveaux trajets? perte client? carburant?)")
            lines.append("- Si les trajets toxiques augmentent : signal de risque structurel, pas une anomalie")
            lines.append("- Mentionne explicitement la tendance dans ta section AUDIT DE RENTABILITE")
            lines.append("- Si retournement de tendance : le signaler comme signal positif")
            lines.append("- Si un client disparait entre deux audits : le nommer et analyser l impact")

        elif module == "stock":
            if d1 is not None:
                sens = "en hausse" if d1 > 0 else "en baisse"
                lines.append(f"  Capital/Articles : {sens} ({d1:+.1f}%)")
            if d2 is not None:
                sens = "en amelioration" if d2 > 0 else "en degradation"
                lines.append(f"  Taux de service : {sens} ({d2:+.1f}%)")
            if d3 is not None:
                sens = "en hausse" if d3 > 0 else "en baisse"
                lines.append(f"  Ruptures : {sens} ({d3:+.1f}%)")
            lines.append("")
            lines.append("INSTRUCTIONS POUR CETTE ANALYSE :")
            lines.append("- Compare les chiffres actuels a cette tendance historique")
            lines.append("- Si taux de service en baisse : priorite urgente dans ton plan d action")
            lines.append("- Si ruptures croissantes : identifier si structurel ou saisonnier")
            lines.append("- Si stock dormant augmente : estimer l impact cash sur la periode")
            lines.append("- Mentionne la tendance dans ton DIAGNOSTIC OPERATIONNEL")

        elif module == "terrain":
            if d2 is not None:
                sens = "meilleure" if d2 > 0 else "moins bonne"
                lines.append(f"  Disponibilite : {sens} ({d2:+.1f}%)")
            if d3 is not None:
                sens = "plus" if d3 > 0 else "moins"
                lines.append(f"  Articles a reapprovisionner : {sens} ({d3:+.1f}%)")
            lines.append("")
            lines.append("INSTRUCTIONS POUR CETTE ANALYSE :")
            lines.append("- Dis si la situation s ameliore ou se degrade par rapport aux semaines precedentes")
            lines.append("- Nomme les articles qui etaient deja en rupture la derniere fois")
            lines.append("- Signale si un article ne bouge pas depuis plusieurs audits consecutifs")

        lines.append("=== FIN DONNEES HISTORIQUES ===\n")

    return "\n".join(lines)


# =========================================
# 0.3 PROMPTS IA BILINGUES
# =========================================
def get_prompt_stock():
    lang=st.session_state.get("language","fr")
    if lang=="en":
        return """You are a Senior Financial Auditor and Supply Chain Director for Logiflo.io.
RESPOND ENTIRELY IN ENGLISH.

RULES on data:
- If prices available: full financial analysis (tied-up capital, dormant stock, cash trap)
- If NO prices: pure operational analysis (rotation, velocity, stock-outs in quantities)
- If consumption history available: calculate coverage in months and 3-year trend
- If NO consumption: flag BLIND SPOT, give sector benchmark (2-4 months healthy coverage)
- If historical audit data present: MANDATORY trend integration into diagnosis

Mandatory structure:

### OPERATIONAL DIAGNOSIS
Service level and rotation. Name the 3 most critical references with exact figures.
If historical data: compare to previous audit, state clearly improving or worsening.

### FINANCIAL DIAGNOSIS AND DORMANT STOCK
Analyze tied-up capital vs sector norms.
CRITICAL RULE ON DORMANT STOCK: a reference is dormant ONLY if a consumption/sales column
exists AND shows zero consumption. If no consumption data exists in the file, it is IMPOSSIBLE
to identify dormant items — say "no consumption data: rotation cannot be calculated".
Never label a reference as dormant based solely on absence of prices or sales data.

### IMMEDIATE ACTION PLAN (TOP 3)
3 concrete actionable recommendations.
If historical data: prioritize recurring issues across multiple audits.
Potential impact: High/Medium/Low | Execution difficulty: 1 to 5

### LOGIFLO SCORE
- Stock Performance and Rotation: /100
- Stock-out Risk: /100
- Supply Chain Resilience: /100

ABSOLUTE RULES:
1. LANGUAGE: use simple direct language only. FORBIDDEN: 'en deca', 'insofar as', 'notwithstanding'. Use: 'below', 'but', 'however' instead.
2. TONE: use a senior consultant's tone — explicitly congratulate what is good,
   firmly signal what is concerning, propose concrete improvement axes.
   Positive example: "your 97% service level is excellent and exceeds the sector benchmark."
   Negative example: "your 75% service level is insufficient — here are the 3 priority actions."
2. NUMBERS: never invent ANY amount, percentage, or quantity not present in the provided data.
3. DORMANT STOCK: only label as dormant with explicit zero-consumption data.
4. HISTORY: if no history: analyze normally without mentioning its absence or writing the section."""
    return """Tu es l'Auditeur Financier et Directeur Supply Chain Senior pour Logiflo.io.
REPONDS IMPERATIVEMENT EN FRANCAIS.

REGLE sur les donnees :
- Si prix disponibles : analyse financiere complete (capital immobilise, dormants, cash trap)
- Si PAS de prix : analyse operationnelle pure (rotation, velocite, ruptures en quantites)
- Si consommations disponibles : calcule couverture en mois et tendance sur 3 ans
- Si PAS de consommations : signale ANGLE MORT et donne mediane sectorielle (2-4 mois couverture saine)
- Si donnees historiques presentes : integre OBLIGATOIREMENT la tendance dans le diagnostic

Structure obligatoire :

### DIAGNOSTIC OPERATIONNEL
Taux de service et rotation. Nomme les 3 references critiques avec chiffres exacts.
Si historique : compare a l'audit precedent et indique si la situation s'ameliore ou se degrade.

### DIAGNOSTIC FINANCIER ET STOCKS DORMANTS
Si prix : capital immobilise, dormants, cash trap.
Si pas de prix : velocite par reference, articles a rotation nulle, risques caches.

### PLAN D'ACTION IMMEDIAT (TOP 3)
3 recommandations concretes et actionnables.
Si historique : priorise les problemes recurrents sur plusieurs audits consecutifs.
Impact potentiel : Fort/Moyen/Faible | Difficulte : 1 a 5

### SCORING LOGIFLO
- Performance et Rotation stock : /100
- Risque de rupture : /100
- Resilience supply chain : /100

REGLES ABSOLUES :
1. VOCABULAIRE : utilise uniquement un francais simple et direct. INTERDIT : 'en deca', 'neanmoins', 'cependant que', 'a cet egard'. Prefere : 'en dessous de', 'mais', 'pourtant'.
2. TON : adopte le ton d'un conseiller senior — felicite explicitement ce qui va bien,
   signale fermement ce qui est preoccupant, propose des axes concrets pour ameliorer.
   Exemple positif : "votre taux de service de 97% est excellent et depasse le benchmark sectoriel."
   Exemple negatif : "votre taux de service de 75% est insuffisant — voici les 3 actions prioritaires."
2. CHIFFRES : n'invente AUCUN montant, AUCUN pourcentage, AUCUNE quantite non presente dans les donnees.
3. DORMANTS : ne qualifier de dormant qu'avec une donnee de consommation explicite a zero.
4. HISTORIQUE : si pas d'historique : analyse normalement sans mentionner son absence ni ecrire la section."""

def get_prompt_terrain():
    lang=st.session_state.get("language","fr")
    if lang=="en":
        return """You are an experienced warehouse supervisor helping your team day-to-day.
RESPOND IN ENGLISH. Direct tone, short sentences. No jargon.

RULES on data:
- If no prices: quantities only
- If no consumption: say so clearly, observe what you can
- If consumption available: calculate coverage in weeks or months
- If historical data: clearly state better or worse than last time
- Always use real references from the file

Structure:

### What is urgent
Items to reorder today. Exact references, exact quantities.
If historical data: flag items already out of stock last time - recurring is a serious signal.

### What changed since last audit
INCLUDE ONLY IF HISTORICAL DATA IS AVAILABLE.
What improved: concrete list.
What got worse: list with one action each.
What is new: items appeared or disappeared from stock.

### What is sleeping
Items with no movement. For each: one concrete action.

### Your 3 actions this week
One line per action. Difficulty: Easy / Medium / Hard

### Summary
2 sentences max to brief your manager.
If historical data: end with "overall: improving / stable / worsening".

RULES: Concrete only. No invented figures. Talk like a colleague."""
    return """Tu es un chef magasinier experimente qui aide son equipe au quotidien.
REPONDS EN FRANCAIS. Ton direct, phrases courtes. Pas de jargon financier.

REGLE sur les donnees :
- Si pas de prix : parle en quantites uniquement
- Si pas de consommations : dis-le clairement et observe ce que tu peux quand meme
- Si consommations disponibles : calcule la couverture en semaines ou en mois
- Si historique disponible : dis clairement si c'est mieux ou moins bien qu'avant
- Cite toujours les vraies references du fichier (REF-001, ART-234, etc.)

Structure :

### Ce qui est urgent
Les articles a commander aujourd'hui. References exactes, quantites exactes.
Si historique : indique les articles qui etaient deja en rupture la derniere fois - si ca se repete c'est grave.

### Ce qui a change depuis le dernier audit
PRESENTE UNIQUEMENT SI HISTORIQUE DISPONIBLE.
Ce qui s'est ameliore : liste courte, concret.
Ce qui s'est degrade : liste courte, avec une action pour chaque point.
Ce qui est nouveau : articles apparus ou disparus du stock.

### Ce qui dort
Articles sans mouvement depuis longtemps. Pour chacun : que faire maintenant ?

### Tes 3 actions pour cette semaine
Une phrase par action. Difficulte : Facile / Moyen / Complique

### En resume
2 phrases max pour briefer ton chef en 30 secondes.
Si historique : termine par "situation globale : en amelioration / stable / en degradation".

REGLES : Concret uniquement. Pas de chiffres inventes. Parle comme a un collegue."""

def get_prompt_transport():
    lang=st.session_state.get("language","fr")
    if lang=="en":
        return """You are a Senior Transport and Supply Chain Strategy Auditor for Logiflo.io.
RESPOND ENTIRELY IN ENGLISH.
DO NOT JUST REPEAT THE DATA: deduce hidden problems and root causes.
If weight is missing: flag STRATEGIC BLIND SPOT.
Adapt vocabulary to detected mode:
- Maritime: TEU, container, demurrage, carrier, port, FCL/LCL
- Air: AWB, chargeable weight, vol/actual ratio, airline, air freight
- Road: FTL/LTL, cost/km, driver, lane, groupage, express
- Rail: wagon, slot, corridor, tonne-km

CNR benchmarks 2025-2026 (cite them in your analysis):
- Long-haul road articulated diesel: 1.85-2.10 EUR/km reference cost
- Regional road rigid truck: 1.40-1.65 EUR/km reference cost
- Fuel share: ~26.5% of total cost

NET MARGIN BENCHMARK — READ THIS BEFORE WRITING ANYTHING:
- Margin > 10%  → EXCELLENT. Above the healthy range. Say: "your X% margin is excellent."
- Margin 6-10%  → HEALTHY. Within the norm. Say: "your X% margin is within the healthy range."
- Margin < 6%   → ALERT. Below minimum. Say: "your X% margin is below the 6% minimum."
- Margin < 0%   → LOSS. Critical. Say: "your negative X% margin signals a critical situation."
NEVER say a margin above 10% is "concerning" or "below the norm". That is factually wrong.
NEVER say a margin of 15% is "below 6%". Read the number first.

Mandatory structure:

### PROFITABILITY AUDIT
State the verdict in ONE sentence using the benchmark rule above.
Then: identify the 3 worst routes/clients with their exact figures from the data.
Root cause hypothesis — not just a description of the numbers.
If historical data available: state improving or worsening with exact numbers.

### NETWORK DIAGNOSIS
Spatial coherence and operational efficiency.
Compare cost/km to CNR benchmarks - cite percentage gaps.
If weight available: load efficiency and cost per tonne.

### RATIONALIZATION PLAN (TOP 3)
3 mode-specific immediately actionable recommendations.
If historical data: start with previously recommended actions not yet implemented.
Cash Impact: High/Medium/Low | Execution difficulty: 1 to 5

### LOGIFLO SCORE
- Profitability and Transport Yield: /100
- Operational Efficiency: /100
- OPEX Control: /100

ABSOLUTE RULES:
1. LANGUAGE: use simple direct language only. FORBIDDEN: 'en deca', 'insofar as', 'notwithstanding'. Use: 'below', 'but', 'however' instead.
2. TONE: use a senior consultant's tone — explicitly congratulate what is good,
   firmly signal what is concerning, propose concrete improvement axes.
   Positive example: "your 97% service level is excellent and exceeds the sector benchmark."
   Negative example: "your 75% service level is insufficient — here are the 3 priority actions."
2. NUMBERS: never invent ANY amount, percentage, or quantity not present in the provided data.
3. DORMANT STOCK: only label as dormant with explicit zero-consumption data.
4. HISTORY: if no history: analyze normally without mentioning its absence or writing the section."""
    return """Tu es un Auditeur Senior en Strategie Transport et Supply Chain pour Logiflo.io.
REPONDS IMPERATIVEMENT EN FRANCAIS.
NE REPETE PAS LES DONNEES : deduis les problemes caches et les causes racines.
Si le poids est absent : signale ANGLE MORT STRATEGIQUE.
Adapte ton vocabulaire au mode detecte :
- Maritime : TEU, conteneur, demurrage, armateur, port, FCL/LCL
- Aerien : AWB, poids taxable, ratio vol/reel, compagnie, fret aerien
- Routier : FTL/LTL, cout/km, chauffeur, axe, messagerie, groupage
- Ferroviaire : wagon, sillon, corridor, tonne-km

Referentiels CNR 2025-2026 (cite-les dans ton analyse) :
- Longue distance articule gazole : 1,85-2,10 EUR/km de reference
- Regional porteur : 1,40-1,65 EUR/km
- Part carburant : ~26,5% du cout total

BENCHMARK MARGE NETTE — LIS CECI AVANT D'ECRIRE QUOI QUE CE SOIT :
- Marge > 10%  → EXCELLENTE. Au-dessus de la norme. Ecris : "votre marge de X% est excellente."
- Marge 6-10%  → SAINE. Dans la norme. Ecris : "votre marge de X% est dans la norme sectorielle."
- Marge < 6%   → ALERTE. En dessous du minimum. Ecris : "votre marge de X% est en dessous du seuil de 6%."
- Marge < 0%   → PERTE. Critique. Ecris : "votre marge negative de X% signale une situation critique."
NE DIS JAMAIS qu'une marge superieure a 10% est "preoccupante" ou "en dessous de la norme". C'est factuellement faux.
NE DIS JAMAIS qu'une marge de 15% est "en dessous de 6%". Lis le chiffre d'abord.

Structure obligatoire :

### AUDIT DE RENTABILITE
Commence par une phrase de verdict en utilisant la regle benchmark ci-dessus.
Ensuite : nomme les 3 pires trajets/clients avec leurs chiffres exacts issus des donnees.
Hypothese experte sur la cause racine — pas juste une description des chiffres.
Si historique disponible : indique si la marge s'ameliore ou se degrade avec des chiffres exacts.

### DIAGNOSTIC RESEAU
Coherence spatiale et efficacite operationnelle.
Compare le cout/km aux referentiels CNR - cite les ecarts en pourcentage.
Si poids disponible : analyse du taux de remplissage et du cout a la tonne.

### PLAN DE RATIONALISATION (TOP 3)
3 recommandations specifiques au mode detecte, actionnables immediatement.
Si historique : commence par les actions recommandees precedemment non encore mises en oeuvre.
Impact Cash : Fort/Moyen/Faible | Difficulte : 1 a 5

### SCORING LOGIFLO
- Rentabilite et Yield Transport : /100
- Efficacite Operationnelle : /100
- Maitrise des OPEX : /100

REGLES ABSOLUES :
1. VOCABULAIRE : utilise uniquement un francais simple et direct. INTERDIT : 'en deca', 'neanmoins', 'cependant que', 'a cet egard'. Prefere : 'en dessous de', 'mais', 'pourtant'.
2. TON : adopte le ton d'un conseiller senior — felicite explicitement ce qui va bien,
   signale fermement ce qui est preoccupant, propose des axes concrets pour ameliorer.
   Exemple positif : "votre taux de service de 97% est excellent et depasse le benchmark sectoriel."
   Exemple negatif : "votre taux de service de 75% est insuffisant — voici les 3 actions prioritaires."
2. CHIFFRES : n'invente AUCUN montant, AUCUN pourcentage, AUCUNE quantite non presente dans les donnees.
3. DORMANTS : ne qualifier de dormant qu'avec une donnee de consommation explicite a zero.
4. HISTORIQUE : si pas d'historique : analyse normalement sans mentionner son absence ni ecrire la section."""

# =========================================
# 1. SESSION STATE
# =========================================
for k,v in {
    "page":"accueil","module":"","auth":False,"current_user":None,
    "language":"fr",
    "df_stock_manager":None,"df_stock_terrain":None,"df_trans":None,"history_stock":[],"stock_view":"MANAGER",
    "seuil_bas":15,"seuil_rupture":0,"seuil_km":0,
    "geo_cache":{},"route_cache":{},"trans_mapping":None,"trans_filename":None,
    "analysis_stock_manager":None,"analysis_stock_terrain":None,"analysis_trans":None,
    "last_pdf":None,"last_kpis":[],"last_labels":[],"trans_mode_detected":None,"audit_gratuit_done":False,
}.items():
    if k not in st.session_state: st.session_state[k]=v

# =========================================
# 2. CSS
# =========================================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Sans:wght@300;400;500&display=swap');
:root{--navy:#0B2545;--navy2:#162D52;--green:#00C896;--green2:#00A87A;--slate:#4A6080;--light:#F0F4F8;--red:#E8304A;--orange:#f39c12;--white:#FFFFFF;}
html,body,[class*="css"]{font-family:'DM Sans',sans-serif;color:var(--navy);}
.block-container{padding-top:2rem!important;padding-bottom:2rem!important;max-width:95%!important;}
.kpi-card{background:var(--white);padding:24px;border-radius:12px;border:1px solid #e2e8f0;border-top:3px solid var(--green);box-shadow:0 4px 6px -1px rgba(0,0,0,0.05);transition:transform 0.2s;}
.kpi-card:hover{transform:translateY(-2px);}
.kpi-card h4{color:var(--slate)!important;font-family:'DM Sans',sans-serif!important;font-size:0.75rem!important;text-transform:uppercase;font-weight:600;letter-spacing:1.5px;margin-bottom:10px;}
.kpi-card h2{font-family:'Syne',sans-serif!important;font-size:2.2rem!important;font-weight:800!important;margin-top:0;line-height:1;letter-spacing:-1px;}
.kpi-card p{font-size:12px;color:var(--slate);margin-top:6px;}
div.stButton>button{border-radius:8px;font-family:'Syne',sans-serif;font-weight:700;background-color:var(--navy);color:#f8fafc;border:none;transition:0.3s;}
div.stButton>button:hover{background-color:var(--navy2);transform:translateY(-2px);}
[data-testid="stSidebar"]{background-color:var(--navy)!important;}
[data-testid="stSidebar"] *{color:#ffffff!important;font-size:1rem!important;}
[data-testid="stSidebar"] hr{border-color:#1e3a5f!important;}
.sidebar-logo{font-family:'Syne',sans-serif;font-size:26px;font-weight:800;color:white;letter-spacing:-0.5px;}
.sidebar-logo span{color:#00C896;}
.import-card{background:var(--white);padding:25px;border-radius:12px;border-left:6px solid var(--green);margin-bottom:20px;box-shadow:0 4px 6px -1px rgba(0,0,0,0.05);}
.import-card h3{margin-top:0;color:var(--navy);font-family:'Syne',sans-serif;font-size:1rem;}
.import-card p{color:var(--slate);font-size:14px;margin-bottom:0;line-height:1.5;}
.report-text{background:var(--light);padding:32px;border-radius:12px;border-left:6px solid var(--navy);line-height:1.8;}
.report-text h3{font-family:'Syne',sans-serif;font-size:1rem;font-weight:800;color:var(--navy);text-transform:uppercase;letter-spacing:1.5px;margin-top:28px;margin-bottom:10px;padding-bottom:6px;border-bottom:2px solid var(--green);}
.report-text h3:first-child{margin-top:0;}
.report-text p{color:#2d3748;font-size:14px;margin-bottom:8px;}
.report-text strong{color:var(--navy);}
.report-terrain{background:#f8fff8;padding:28px;border-radius:12px;border-left:6px solid var(--green);line-height:1.9;}
.report-terrain h3{font-family:'Syne',sans-serif;font-size:1rem;font-weight:700;color:var(--green2);margin-top:24px;margin-bottom:8px;}
.report-terrain h3:first-child{margin-top:0;}
.report-terrain p{color:#1a2e1a;font-size:15px;margin-bottom:6px;}
.mode-badge{display:inline-flex;align-items:center;gap:8px;background:rgba(0,200,150,0.1);border:1px solid rgba(0,200,150,0.3);color:var(--green2);font-size:13px;font-weight:600;padding:8px 16px;border-radius:8px;margin-bottom:16px;}
.sans-prix-badge{background:rgba(0,200,150,0.1);border:1px solid rgba(0,200,150,0.3);color:var(--green2);font-size:12px;font-weight:600;padding:4px 12px;border-radius:20px;display:inline-block;margin-bottom:12px;margin-right:8px;}
.archive-card{background:var(--white);border:1px solid #E2EAF4;border-radius:12px;padding:20px;margin-bottom:16px;border-left:4px solid var(--green);}
.archive-card h4{font-family:'Syne',sans-serif;font-size:14px;font-weight:700;color:var(--navy);margin-bottom:8px;}
.archive-kpi{display:inline-block;background:var(--light);border-radius:6px;padding:4px 10px;font-size:12px;font-weight:600;color:var(--navy);margin-right:8px;}
.big-emoji{font-size:70px;margin-bottom:10px;display:block;text-align:center;}
.legal-text{background:var(--white);padding:32px;border-radius:12px;border:1px solid #E2EAF4;line-height:1.9;}
.legal-text h2{font-family:'Syne',sans-serif;font-size:1.1rem;font-weight:800;color:var(--navy);margin-top:28px;margin-bottom:10px;padding-bottom:6px;border-bottom:2px solid var(--green);}
.legal-text h2:first-child{margin-top:0;}
.legal-text p,.legal-text li{color:#2d3748;font-size:14px;margin-bottom:6px;}
.legal-box{background:var(--light);border-left:4px solid var(--green);padding:16px 20px;border-radius:8px;margin:12px 0;}
.legal-box p{color:var(--navy)!important;font-weight:500;}
</style>
""",unsafe_allow_html=True)

# =========================================
# 3. HELPERS
# =========================================
def render_report(texte,mode="manager"):
    css="report-terrain" if mode=="terrain" else "report-text"
    lines=[]
    for line in texte.split('\n'):
        line=line.strip()
        if not line: continue
        if line.startswith('### '):
            lines.append(f"<h3>{line[4:].strip()}</h3>")
        else:
            line=re.sub(r'\*\*(.+?)\*\*',r'<strong>\1</strong>',line)
            prefix="• " if (line.startswith('- ') or line.startswith('* ')) else ""
            body=line[2:] if prefix else line
            lines.append(f"<p>{prefix}{body}</p>")
    return f'<div class="{css}">{"".join(lines)}</div>'

def nettoyer(t):
    t=str(t).lower()
    t=unicodedata.normalize('NFD',t).encode('ascii','ignore').decode("utf-8")
    return re.sub(r'[^a-z0-9]','',t)

class StepProgress:
    """Barre de progression simple — texte configurable, sans detail technique."""
    def __init__(self, steps, text=None):
        self._ph = st.empty()
        self._n  = max(len(steps), 1)
        self._i  = 0
        lang = st.session_state.get("language", "fr")
        if text:
            self._txt = text
        else:
            self._txt = "Computing..." if lang == "en" else "Calcul en cours..."
        self._ph.progress(0, text=self._txt)
    def step(self, label=None):
        self._i += 1
        pct = min(self._i / self._n, 1.0)
        self._ph.progress(pct, text=self._txt)
    def done(self):
        self._ph.empty()

# =========================================
# 4. SMART INGESTER V5
# =========================================
SYNONYMES={
    "reference":["reference","ref","article","code","codearticle","codeproduit","cdarticle",
                 "cdart","cdproduit","codemat","codematiere","nomarticle","nomproduit",
                 "sku","ean","ean13","upc","gtin","produit","designation","libelle",
                 "description","descproduit","descarticle","nom","item","itemcode",
                 "itemno","itemref","partnumber","partno","partref","refarticle",
                 "refproduit","refcommande","numero","numeroproduit","matricule",
                 "identifiant","id","cable","cablage","matiere","materiel","composant",
                 "piece","repere","nomenclature","famille","sousfamille","categorie",
                 "productcode","productref","dsg","desig","design","designat",
                 "articlecode","articleref","artcode","artref","artno","artnum"],
    "quantite":["quantite","qte","qty","qtstk","qte_stk","qtestk","stock","stk","stockactuel",
                "stockdispo","stockdisponible","stockreel","stockphysique","niveaustock",
                "qtestock","qtedispo","qtedisponible","qtereel","qtephysique",
                "volume","pieces","pcs","units","unit","unites","restant",
                "solde","soldedisponible","encours","inventaire","disponible",
                "existant","existants","present","metre","metres","meter","meters",
                "bobine","bobines","longueur","longueurstock","quantitedisponible",
                "quantitestock","quantiterestante","quantitepresente",
                "nbarticle","nbarticles","nbpieces","nbunites","nb","nbre","nombre",
                "qte_disponible","qt_stk","qtstck","qtstock"],
    "prix_unitaire":["prix","prixunitaire","prixachat","prixderevient","prixmoyen",
                     "prixmoyenpondere","pmp","pa","pu","pxu","px_u","price","unitprice",
                     "avgprice","cout","coutunitaire","coutachat","coutderevient","coutmoyen",
                     "cost","unitcost","avgcost","valeur","valeurunitaire","valeurachat",
                     "tarif","tarifunitaire","montantunitaire","achat","prixfournisseur",
                     "euro","eur","devise","prixbase","baseachat","priceeuro","priceeur"],
    "conso_an1":["conso2022","conso22","consommation2022","sorties2022","ventes2022",
                 "c2022","n3","nminus3","annee2022","a2022","quantite2022","qte2022","cso22","cso2022"],
    "conso_an2":["conso2023","conso23","consommation2023","sorties2023","ventes2023",
                 "c2023","n2","nminus2","annee2023","a2023","quantite2023","qte2023","cso23","cso2023"],
    "conso_an3":["conso2024","conso24","consommation2024","sorties2024","ventes2024",
                 "c2024","n1","nminus1","annee2024","a2024","quantite2024","qte2024","cso24","cso2024"],
    "conso_an4":["conso2025","conso25","consommation2025","sorties2025","ventes2025",
                 "c2025","n0","nactuel","annee2025","a2025","quantite2025","qte2025",
                 "cso25","cso2025","sortie2025","consoactuelle","consoencoursannee"],
    # Transport
    "ca":         ["ca","chiffreaffaires","revenue","revenu","facture","facturation",
                   "recette","vente","ventes","montantfacture","montantca","totalca",
                   "prixvente","prixdevente","tariflbp","tariflocation","tarifclient",
                   "turnover","sales","salesamount","invoiceamount","totalrevenue"],
    "co":         ["cout","couts","cost","costs","charge","charges","depense","depenses",
                   "coutrevient","coutderevient","coutachat","coutexploitation",
                   "coutprestation","coutservice","couttransport","frais","fraistransport",
                   "fraisexploitation","montantachat","totalcout","totalcouts",
                   "benefice","marge","margebrute","margenet","margennette","profit",
                   "resultat","solde","bilan","gain","perte"],
    "fournisseur":["fournisseur","supplier","vendor","fournisseurs","suppliers",
                   "prestataire","prestataires","acheteur","source","origine",
                   "partnername","vendorname","suppliername","sourcing"],
    "date_col":   ["date","dates","dateop","datetransaction","datemouvement",
                   "datecommande","datelivraison","datesortie","datentree","dateachat",
                   "datestock","period","periode","mois","month","annee","year",
                   "semaine","week","exercice","timestamp","datetime","jour","day"],
    "delai":      ["delai","delailivraison","leadtime","lt","lead","delaifournisseur",
                   "delaiapprovisionnement","delaireapprovisionnement","delaicommande",
                   "supplierleadtime","leadtimedays","leadtimeweeks"],
    "categorie":  ["categorie","categories","category","famille","familles","family",
                   "sousfamille","type","types","classe","classes","segment","gamme",
                   "rayon","departement","division","group","groupe"],
}

# Alias géographiques pour ports, aéroports, villes logistiques
GEO_ALIASES = {
    # Ports maritimes France
    "marseille":["marseille","fos","fos-sur-mer","fos sur mer","gpmm","port de marseille"],
    "le havre":  ["le havre","havre","lehavre","gpmh","port du havre"],
    "dunkerque": ["dunkerque","dunkerque","dunkirk","gpmd"],
    "nantes":    ["nantes","saint nazaire","nantes saint-nazaire","nantes-saint-nazaire"],
    "bordeaux":  ["bordeaux","le verdon","bassens","bordeaux atlantique"],
    "rouen":     ["rouen","grand port rouen"],
    # Aéroports France
    "cdg":       ["cdg","roissy","charles de gaulle","paris-cdg","paris cdg",
                  "roissy charles de gaulle","roissy-cdg","lfpg"],
    "orly":      ["orly","paris-orly","paris orly","lfpo"],
    "lyon":      ["lyon","saint-exupery","saint exupery","lyon saint-exupery","lfly"],
    "nice":      ["nice","nice cote d'azur","nice cote dazur","lfmn"],
    "marseille_aero":["marseille provence","mp","lfml","marseille-provence"],
    # Ports maritimes Europe
    "rotterdam": ["rotterdam","eurtm","port of rotterdam"],
    "anvers":    ["anvers","antwerp","antwerpen","port of antwerp"],
    "hambourg":  ["hambourg","hamburg","port of hamburg"],
    "genes":     ["genes","genova","genoa","port de genes","porto di genova"],
    "barcelone": ["barcelone","barcelona","port de barcelone"],
    "valence":   ["valence es","valencia","port of valencia"],
    "algeciras": ["algeciras","gibraltar","port of algeciras"],
    "tanger":    ["tanger","tanger med","tangier","tanger-med"],
    # Villes logistiques France
    "paris":     ["paris","ile-de-france","idf","region parisienne","rungis"],
    "lyon_v":    ["lyon","rhone-alpes","auvergne-rhone-alpes"],
    "toulouse":  ["toulouse","midi-pyrenees","occitanie"],
    "lille":     ["lille","nord","hauts-de-france"],
    "strasbourg":["strasbourg","alsace","grand est","bas-rhin"],
    # Pays (transport international)
    "france":    ["france","fr","fra"],
    "maroc":     ["maroc","morocco","ma","mar","casablanca","rabat","agadir","tanger"],
    "algerie":   ["algerie","algeria","dz","alger","oran","annaba"],
    "tunisie":   ["tunisie","tunisia","tn","tunis","sfax","sousse"],
    "cote d'ivoire":["cote d'ivoire","ivory coast","ci","civ","abidjan","bouake"],
    "senegal":   ["senegal","sn","sen","dakar"],
    "espagne":   ["espagne","spain","es","esp","madrid","barcelone","valence"],
    "italie":    ["italie","italy","it","ita","rome","milan","genes"],
    "allemagne": ["allemagne","germany","de","deu","hambourg","francfort","munich"],
    "belgique":  ["belgique","belgium","be","bel","anvers","bruxelles"],
    "pays-bas":  ["pays-bas","netherlands","nl","nld","rotterdam","amsterdam"],
    "royaume-uni":["royaume-uni","uk","gbr","gb","london","londres","manchester"],
    "chine":     ["chine","china","cn","chn","shanghai","shenzhen","guangzhou"],
    "usa":       ["usa","etats-unis","us","united states","new york","los angeles"],
}

def _normalize_geo(text):
    """Normalise un nom géographique vers sa forme standard."""
    if not text:
        return text
    t = str(text).lower().strip()
    for canonical, aliases in GEO_ALIASES.items():
        if t in aliases:
            return canonical
    return t

def detect_periode(df):
    """
    Détecte automatiquement la période depuis les colonnes de dates.
    Ne pose jamais de question à l'utilisateur.
    Retourne un dict avec trimestre, label, contexte saisonnier.
    """
    import datetime as _dt_p
    from dateutil import parser as _dparser

    mois_min, mois_max, annee = None, None, _dt_p.date.today().year

    # Chercher colonnes de dates
    date_cols = [c for c in df.columns
                 if any(k in str(c).lower() for k in
                        ["date","mois","month","periode","semaine","week",
                         "exercice","timestamp","datetime","jour","day"])]

    for col in date_cols:
        try:
            sample = df[col].dropna().head(50).astype(str)
            parsed = []
            for v in sample:
                try:
                    d = _dparser.parse(v, dayfirst=True)
                    parsed.append(d)
                except Exception:
                    pass
            if len(parsed) >= 3:
                mois_vals = [d.month for d in parsed]
                annee = max(d.year for d in parsed)
                mois_min = min(mois_vals)
                mois_max = max(mois_vals)
                break
        except Exception:
            continue

    # Si pas de dates trouvées → utiliser date du jour
    if mois_min is None:
        today = _dt_p.date.today()
        mois_min = mois_max = today.month
        annee = today.year

    # Déterminer le trimestre dominant
    if mois_max <= 3:
        trim = "T1"
    elif mois_max <= 6:
        trim = "T2"
    elif mois_max <= 9:
        trim = "T3"
    else:
        trim = "T4"

    labels_trim = {"T1":"Janv-Mars","T2":"Avr-Juin","T3":"Juil-Sept","T4":"Oct-Déc"}

    # Contexte saisonnier injecté dans le prompt IA
    saison = "standard"
    if mois_min >= 10 or (mois_max >= 10 and mois_min >= 9):
        saison = "pre_fetes"
    elif mois_min >= 6 and mois_max <= 9:
        saison = "ete"
    elif mois_max <= 2:
        saison = "post_fetes"
    elif mois_min >= 3 and mois_max <= 5:
        saison = "printemps"

    contextes_fr = {
        "pre_fetes":  "Periode pre-fetes (oct-dec). Un surstock est normal "
                      "en anticipation de Noel. Ne pas conclure a un probleme "
                      "sans verifier si c'est du stock de precaution delibere.",
        "ete":        "Periode estivale (jun-sept). Attention aux variations de "
                      "consommation liees aux conges. Certaines ruptures peuvent "
                      "etre temporaires et non structurelles.",
        "post_fetes": "Periode post-fetes (jan-fev). Les surstocks residuels de "
                      "Noel sont normaux. Distinguer stock dormant et stock de "
                      "fin de saison en cours de liquidation.",
        "printemps":  "Periode printemps (mar-mai). Debut de saison pour certains "
                      "secteurs (BTP, jardin, mode ete). Anticiper la montee en charge.",
        "standard":   ""
    }
    contextes_en = {
        "pre_fetes":  "Pre-holiday period (Oct-Dec). Overstock is normal in "
                      "anticipation of Christmas. Do not flag as a problem without "
                      "checking if it is deliberate precautionary stock.",
        "ete":        "Summer period (Jun-Sep). Watch for consumption variations "
                      "due to holidays. Some stockouts may be temporary.",
        "post_fetes": "Post-holiday period (Jan-Feb). Residual Christmas overstock "
                      "is normal. Distinguish dormant stock from end-of-season "
                      "liquidation stock.",
        "printemps":  "Spring period (Mar-May). Start of season for some sectors "
                      "(construction, garden, summer fashion). Anticipate ramp-up.",
        "standard":   ""
    }

    return {
        "trimestre": trim,
        "mois_min":  mois_min,
        "mois_max":  mois_max,
        "annee":     annee,
        "label":     f"{labels_trim[trim]} {annee}",
        "saison":    saison,
        "contexte_fr": contextes_fr[saison],
        "contexte_en": contextes_en[saison],
    }



def _levenshtein(s1,s2):
    if len(s1)<len(s2): return _levenshtein(s2,s1)
    if len(s2)==0: return len(s1)
    prev=list(range(len(s2)+1))
    for i,c1 in enumerate(s1):
        curr=[i+1]
        for j,c2 in enumerate(s2):
            curr.append(min(prev[j+1]+1,curr[j]+1,prev[j]+(c1!=c2)))
        prev=curr
    return prev[-1]

def _score_nom(propre,std):
    syns=SYNONYMES.get(std,[]);best=0
    try:
        from rapidfuzz import fuzz as _rfuzz
        for syn in syns:
            if propre==syn: return 100
            if len(syn)>=4 and propre.startswith(syn): best=max(best,95)
            if len(syn)>=3 and syn in propre: best=max(best,88)
            if len(propre)>=3 and propre in syn: best=max(best,82)
            # rapidfuzz — plus précis que difflib
            best=max(best,int(_rfuzz.ratio(propre,syn)))
            best=max(best,int(_rfuzz.partial_ratio(propre,syn)*0.9))
            best=max(best,int(_rfuzz.token_sort_ratio(propre,syn)*0.85))
    except ImportError:
        for syn in syns:
            if propre==syn: return 100
            if len(syn)>=4 and propre.startswith(syn): best=max(best,95)
            if len(syn)>=3 and syn in propre: best=max(best,88)
            if len(propre)>=3 and propre in syn: best=max(best,82)
            r=difflib.SequenceMatcher(None,propre,syn).ratio()
            best=max(best,int(r*85))
            if len(propre)>=3 and len(syn)>=3:
                dist=_levenshtein(propre,syn); ml=max(len(propre),len(syn))
                if ml>0: best=max(best,int((1-dist/ml)*78))
    year_bonus={"conso_an1":["2022","22"],"conso_an2":["2023","23"],
                "conso_an3":["2024","24"],"conso_an4":["2025","25"]}
    if std in year_bonus and any(y in propre for y in year_bonus[std]): best=max(best,85)
    return best

def _score_contenu(series,std):
    sample=series.dropna().head(50)
    if len(sample)==0: return 0
    cleaned=(sample.astype(str).str.replace(r'[€$£\s\xa0%]','',regex=True)
             .str.replace(',','.',regex=False).str.replace(r'[^\d.\-]','',regex=True))
    numeric=pd.to_numeric(cleaned,errors='coerce')
    pct_num=numeric.notna().mean(); vals=numeric.dropna()
    raw_text=sample.astype(str)
    avg_len=raw_text.str.len().mean()
    pct_alpha=raw_text.str.contains(r'[a-zA-Z]',na=False).mean()
    unique_r=sample.nunique()/len(sample)
    has_dec=(vals%1!=0).mean() if len(vals)>0 else 0
    pct_int=(vals%1==0).mean() if len(vals)>0 else 0
    pct_pos=(vals>=0).mean() if len(vals)>0 else 0
    pct_zero=(vals==0).mean() if len(vals)>0 else 0
    if std=="reference":
        score=0
        if pct_alpha>0.5: score+=40
        if unique_r>0.7: score+=25
        if 3<=avg_len<=50: score+=20
        if pct_num<0.5: score+=15
        if pct_num>0.9 and pct_alpha<0.1: score-=30
        return max(0,min(score,100))
    elif std=="quantite":
        if pct_num<0.6: return 10
        score=40
        if pct_int>0.85: score+=30
        elif pct_int>0.65: score+=15
        if pct_zero>0.05: score+=8
        if pct_pos>0.85: score+=8
        if has_dec>0.55: score-=20
        if pct_alpha>0.3: score-=25
        return max(0,min(score,100))
    elif std=="prix_unitaire":
        if pct_num<0.6: return 5
        score=35
        if has_dec>0.45: score+=30
        elif has_dec>0.25: score+=15
        if pct_zero<0.05: score+=12
        if pct_pos>0.85: score+=8
        if pct_int>0.95: score-=15
        if pct_alpha>0.3: score-=25
        return max(0,min(score,100))
    elif std in("conso_an1","conso_an2","conso_an3","conso_an4"):
        if pct_num<0.5: return 5
        score=30
        if pct_int>0.80: score+=25
        elif pct_int>0.60: score+=12
        if pct_zero>0.15: score+=15
        if pct_pos>0.5: score+=10
        if has_dec>0.5: score-=15
        if pct_alpha>0.3: score-=25
        return max(0,min(score,100))
    return 0

def smart_ingester_stock_ultime(df,client_ai=None):
    df=df.dropna(how='all').copy()
    # Supprimer les lignes où la référence est vide, None ou nan
    if "reference" in df.columns:
        df = df[df["reference"].astype(str).str.strip().str.lower().isin(
            ["nan","none","null","","n/a","-"]) == False].copy()
    df=df[df.apply(lambda r:r.astype(str).str.strip().ne('').any(),axis=1)]
    CIBLES=list(SYNONYMES.keys())
    propres={col:nettoyer(col) for col in df.columns}
    scores={std:{} for std in CIBLES}
    for col in df.columns:
        propre=propres[col]
        for std in CIBLES:
            sn=_score_nom(propre,std); sc=_score_contenu(df[col],std)
            if sn>=70: sf=int(sn*0.65+sc*0.35)
            elif sn>=45: sf=int(sn*0.55+sc*0.45)
            else: sf=int(sn*0.25+sc*0.75)
            scores[std][col]=min(sf,100)
    # Ajustement contextuel
    for col in df.columns:
        vals=pd.to_numeric(df[col].astype(str).str.replace(r'[^\d.,-]','',regex=True).str.replace(',','.'),errors='coerce').dropna()
        if len(vals)>5:
            if (vals%1==0).mean()>0.9 and vals.median()>10:
                scores["quantite"][col]=min(scores["quantite"][col]+10,100)
                scores["prix_unitaire"][col]=max(scores["prix_unitaire"][col]-8,0)
            if (vals%1!=0).mean()>0.5 and vals.median()<1000:
                scores["prix_unitaire"][col]=min(scores["prix_unitaire"][col]+10,100)
                scores["quantite"][col]=max(scores["quantite"][col]-8,0)
    trouvees={}; utilisees=set()
    ORDRE=["reference","quantite","prix_unitaire","conso_an4","conso_an3","conso_an2","conso_an1"]
    SEUILS={"reference":35,"quantite":55,"prix_unitaire":55,
            "conso_an4":55,"conso_an3":55,"conso_an2":55,"conso_an1":55}
    for std in ORDRE:
        seuil=SEUILS.get(std,55)
        candidats=[(col,scores[std][col]) for col in scores[std]
                   if col not in trouvees and scores[std][col]>=seuil]
        if not candidats: continue
        nom_forts=[(col,sc) for col,sc in candidats if _score_nom(propres[col],std)>=70]
        gagnant=max(nom_forts,key=lambda x:_score_nom(propres[x[0]],std))[0] if nom_forts else max(candidats,key=lambda x:x[1])[0]
        trouvees[gagnant]=std; utilisees.add(std)
    cols=list(df.columns)
    if "reference" not in utilisees:
        for c in cols:
            if c not in trouvees:
                s=df[c].dropna().head(20)
                if s.astype(str).str.contains(r'[a-zA-Z]',na=False).mean()>0.3 or s.nunique()/max(len(s),1)>0.6:
                    trouvees[c]="reference"; utilisees.add("reference"); break
    if "quantite" not in utilisees:
        for c in cols:
            if c not in trouvees:
                num=pd.to_numeric(df[c].astype(str).str.replace(r'[^\d.,-]','',regex=True).str.replace(',','.'),errors='coerce')
                if num.notna().mean()>0.6 and (num.dropna()%1==0).mean()>0.6:
                    trouvees[c]="quantite"; utilisees.add("quantite"); break
    critiques=[s for s in ["reference","quantite"] if s not in utilisees]
    if critiques and client_ai:
        titres=list(df.columns); sample_data=df.head(5).astype(str).to_dict(orient='list')
        prompt=f"""Logistics file. Columns: {titres}
Data (5 rows): {json.dumps(sample_data,ensure_ascii=False)[:3000]}
Missing concepts: {critiques}
Reply ONLY JSON: {{"concept": "exact_title"}} or null. Choose from: {titres}"""
        try:
            r=client_ai.chat.completions.create(model="gpt-4o-mini",
                messages=[{"role":"system","content":prompt}],temperature=0.0)
            raw=r.choices[0].message.content.strip().replace("```json","").replace("```","").strip()
            gpt_map=json.loads(raw)
            for std,col in gpt_map.items():
                if std in critiques and col in df.columns and col not in trouvees:
                    trouvees[col]=std; utilisees.add(std)
        except: pass
    df=df.rename(columns=trouvees)
    manq=[c for c in ["reference","quantite"] if c not in df.columns]
    if manq:
        return None,(f"Colonnes introuvables : {', '.join(manq)}.\nColonnes dans votre fichier : {list(df.columns[:10])}")
    df["quantite"]=pd.to_numeric(df["quantite"].astype(str).str.replace(r'[^\d.,-]','',regex=True).str.replace(',','.'),errors='coerce')
    df=df.dropna(subset=["quantite"]).copy()
    df=df[df["reference"].astype(str).str.strip().ne('')]
    df=df[~df["reference"].astype(str).str.lower().isin(['nan','none',''])]
    if "prix_unitaire" not in df.columns:
        df["prix_unitaire"]=0.0; df["_sans_prix"]=True
    else:
        df["prix_unitaire"]=pd.to_numeric(df["prix_unitaire"].astype(str).str.replace(r'[^\d.,-]','',regex=True).str.replace(',','.'),errors='coerce').fillna(0)
        df["_sans_prix"]=(df["prix_unitaire"]==0).all()
    has_conso=False; conso_cols=[]
    for c in ["conso_an1","conso_an2","conso_an3","conso_an4"]:
        if c in df.columns:
            df[c]=pd.to_numeric(df[c].astype(str).str.replace(r'[^\d.,-]','',regex=True).str.replace(',','.'),errors='coerce').fillna(0)
            conso_cols.append(c); has_conso=True
    df["_has_conso"]=has_conso
    df["_conso_moy"]=df[conso_cols].mean(axis=1) if has_conso else 0.0
    return df.copy(),"Succès"

# =========================================
# 5. AUTO MAP TRANSPORT
# =========================================
def auto_map_columns_with_ai(df):
    titres=list(df.columns)
    profil={col:{"exemples":list(df[col].dropna().astype(str).unique()[:5])} for col in titres}
    prompt=f"""Titres: {titres}\nDonnées: {json.dumps(profil,ensure_ascii=False)}
Associe à un titre EXACT. Si absent: null.
Concepts: "client","ca","co","dep","arr","dist","poids","mode".
JSON uniquement."""
    try:
        r=client.chat.completions.create(model="gpt-4o-mini",
            messages=[{"role":"system","content":prompt}],temperature=0.0)
        raw=r.choices[0].message.content.strip().replace("```json","").replace("```","").strip()
        return {k:v for k,v in json.loads(raw).items() if v in titres}
    except:
        return {"client":titres[0],"ca":titres[1] if len(titres)>1 else None,"co":None}

# =========================================
# 6. GÉNÉRATION IA
# =========================================
def generate_ai_analysis(data_summary, historique_txt="", df_raw=None,
                          sector_key=None, mode_detected=None):
    """
    Moteur IA V2 — prescriptif, sectoriel, historique garanti.
    Envoie : donnees enrichies + benchmarks sectoriels + historique + question precise.
    """
    lang   = st.session_state.get("language","fr")
    module = st.session_state.module
    view   = st.session_state.get("stock_view","MANAGER")

    # Detecter le secteur si pas fourni
    if not sector_key:
        sector_key = detect_sector(
            df=df_raw, module=module,
            mode_detected=mode_detected or (
                st.session_state.trans_mode_detected[0]
                if st.session_state.get("trans_mode_detected") else None
            )
        )

    # Benchmarks sectoriels
    benchmarks = get_sector_benchmarks(sector_key, lang)

    # Prompt systeme selon module et profil
    if module == "transport":
        sys_prompt = get_prompt_transport()
    elif view == "TERRAIN":
        sys_prompt = get_prompt_terrain()
    else:
        sys_prompt = get_prompt_stock()

    # ── Construction du message utilisateur enrichi ──────────────────────
    parts = []

    # 1. Données actuelles
    if lang == "en":
        parts.append(f"=== CURRENT AUDIT DATA ===\n{data_summary}")
    else:
        parts.append(f"=== DONNEES AUDIT ACTUEL ===\n{data_summary}")

    # 2. Benchmarks sectoriels — toujours présents
    if lang == "en":
        parts.append(f"=== SECTOR BENCHMARKS TO USE FOR COMPARISON ===\n{benchmarks}")
    else:
        parts.append(f"=== BENCHMARKS SECTORIELS A UTILISER POUR LA COMPARAISON ===\n{benchmarks}")

    # 3. Historique — garanti ou explicitement absent
    if historique_txt and historique_txt.strip():
        parts.append(historique_txt)
    else:
        if lang == "en":
            parts.append("=== HISTORY ===\nFirst audit for this user — no historical comparison available. Focus on present situation and benchmarks.")
        else:
            parts.append("=== HISTORIQUE ===\nPremier audit de cet utilisateur — pas de comparaison historique disponible. Concentre-toi sur la situation presente et les benchmarks.")

    # 4. Lignes brutes les plus significatives (si dataframe disponible)
    if df_raw is not None:
        try:
            key_data = _extract_key_rows(df_raw, module, lang)
            if key_data:
                parts.append(key_data)
        except Exception:
            pass

    # 4b. Saisonnalité détectée automatiquement
    try:
        _periode = detect_periode(df_raw) if df_raw is not None else None
        if _periode and _periode.get("contexte_fr" if lang=="fr" else "contexte_en"):
            _ctx = _periode.get("contexte_fr") if lang=="fr" else _periode.get("contexte_en")
            _lbl = _periode.get("label","")
            if lang == "en":
                parts.append(f"=== SEASONAL CONTEXT (auto-detected) ===\n"
                             f"Period: {_lbl}\n{_ctx}")
            else:
                parts.append(f"=== CONTEXTE SAISONNIER (detecte automatiquement) ===\n"
                             f"Periode : {_lbl}\n{_ctx}")
    except Exception:
        pass

    # 4c. Prédictions rupture (module stock uniquement)
    if module == "stock" and df_raw is not None:
        try:
            _alertes = predict_ruptures(df_raw, lang=lang)
            _pred_txt = format_predictions_pour_prompt(_alertes, lang)
            if _pred_txt:
                parts.append(_pred_txt)
        except Exception:
            pass

    # 4d. Alerte BFR (module stock uniquement)
    if module == "stock" and df_raw is not None:
        try:
            _bfr = compute_alerte_bfr(df_raw, lang=lang)
            if _bfr.get("available") and _bfr.get("texte"):
                if lang == "en":
                    parts.append(f"=== WORKING CAPITAL (BFR) ALERT ===\n{_bfr['texte']}")
                else:
                    parts.append(f"=== ALERTE BESOIN EN FONDS DE ROULEMENT (BFR) ===\n{_bfr['texte']}")
        except Exception:
            pass

    # 4e. Scoring pré-calculé (contexte pour l'IA)
    try:
        _kpis_ctx = st.session_state.get("last_kpis",[])
        _labels_ctx = st.session_state.get("last_labels",[])
        _score_ctx = compute_logiflo_score(
            module=module, df=df_raw,
            kpis=_kpis_ctx, labels=_labels_ctx,
            sector_key=sector_key or "generique", lang=lang
        )
        if _score_ctx.get("global",0) > 0:
            if lang == "en":
                parts.append(f"=== PRE-COMPUTED LOGIFLO SCORE ===\n"
                             f"Global score: {_score_ctx['global']}/100\n"
                             f"{_score_ctx.get('format_pdf','')}\n"
                             f"Comment this score in your ### LOGIFLO SCORE section. "
                             f"Explain what drives each dimension up or down.")
            else:
                parts.append(f"=== SCORING LOGIFLO PRE-CALCULE ===\n"
                             f"Score global : {_score_ctx['global']}/100\n"
                             f"{_score_ctx.get('format_pdf','')}\n"
                             f"Commente ce score dans ta section ### SCORING LOGIFLO. "
                             f"Explique ce qui tire chaque dimension vers le haut ou vers le bas.")
    except Exception:
        pass

    # 5. Instruction prescriptive — structure complète + 2 options décisionnaires
    if lang == "en":
        if module == "transport":
            parts.append("""=== YOUR INSTRUCTIONS ===
Using ALL data above, write a complete structured audit with the following sections.
Each section must be fully developed — not bullet points only, real analytical sentences.

### PROFITABILITY AUDIT
APPLY STRICTLY THE BENCHMARK DEFINED IN THE SYSTEM PROMPT.
Consultant tone: congratulate if margin > 6%, signal alert if < 6%, critical if < 0%.
Name the 3 worst routes with their EXACT figures from the file.
If names are "nan" or empty: write "missing client data in source file".

### NETWORK DIAGNOSIS
Analyze the cost/km vs CNR benchmarks for each route type. Assess spatial coherence,
empty return load ratio if visible. Cite exact benchmark figures in your comparison.

### WHAT TO DO - TOP PRIORITY
Start this section with the single most urgent action. One direct sentence: name the client
or route, the action, and the estimated cash recovery. Then one fallback if too risky.
The manager decides — give them exactly what they need to act today.

### RATIONALIZATION PLAN (1-2-3)
Now develop 3 strategic recommendations for the medium term. For each: specific action,
targeted client/route, expected impact in EUR, execution difficulty 1-5. These complement
the urgent action above and guide the manager beyond this week.

### LOGIFLO SCORE
- Profitability and Transport Yield: XX/100
- Operational Efficiency: XX/100
- OPEX Control: XX/100""")
        elif view == "TERRAIN":
            parts.append("""=== YOUR INSTRUCTIONS ===
Write as a warehouse supervisor who knows this floor. Direct language, no financial jargon.
Fully develop each section — real sentences, not just bullet points.

### WHAT IS URGENT
Name the references in stockout or near-stockout based on THE PROVIDED DATA.
For each: exact reference as it appears in the data, exact current stock level,
why it is critical. DO NOT invent order quantities — you do not have that information.
Say "stock: X units" but NEVER "order X units" unless historical consumption data is available.
If a reference was already in shortage at the last audit (if history provided), flag it explicitly.

### WHAT CHANGED SINCE LAST AUDIT
ONLY if historical data is explicitly provided in the context above.
If the history section says "First audit" or "no history available": SKIP THIS SECTION ENTIRELY,
do not even write the title.

### WHAT IS SLEEPING
References with no movement detected in this file. IMPORTANT: if this is a first audit,
do NOT say "for 6 months" or any duration — say "no movement detected in this file".
The actual duration is unknown without history. For each: exact reference, exact stock, proposed action.

### WHAT TO DO NOW
The most urgent action based on real data. Exact reference, exact stock level.
Do not invent order quantities. If budget tight: one concrete alternative.
The team decides — no ambiguity.

### YOUR 3 ACTIONS THIS WEEK
3 practical actions ranked by urgency. One concrete sentence each.
Difficulty: Easy / Medium / Hard. Based only on available data.

### SUMMARY
2-3 sentences to brief the manager in 30 seconds. End with overall situation: improving / stable / worsening.""")
        else:
            parts.append("""=== YOUR INSTRUCTIONS ===
Using ALL data above, write a complete structured audit. Fully develop each section
with analytical sentences — not bullet points only.

### OPERATIONAL DIAGNOSIS
Overall verdict first (one sentence). Then: service level vs sector benchmark (cite the exact
gap), rotation analysis, name the 3 most critical references with exact figures and why
they are critical.

### FINANCIAL DIAGNOSIS AND DORMANT STOCK
Analyze tied-up capital vs sector norms, identify dormant and overstock references with
exact values, quantify the cash trap risk. If no prices: analyze velocity and hidden risks.

### WHAT TO DO - TOP PRIORITY
Start this section with the single most urgent action. One direct sentence: name the reference,
the action, and the estimated impact in EUR or %. One fallback if budget is constrained.
The manager decides — give them exactly what they need to act today.

### IMMEDIATE ACTION PLAN (1-2-3)
Now develop 3 strategic recommendations for the medium term. For each: specific action,
targeted reference, expected impact, execution difficulty 1-5. These complement the urgent
action above and guide the manager on stock optimization beyond this week.

### LOGIFLO SCORE
- Stock Performance and Rotation: XX/100
- Stock-out Risk: XX/100
- Supply Chain Resilience: XX/100""")
    else:
        if module == "transport":
            parts.append("""=== TES INSTRUCTIONS ===
En utilisant TOUTES les donnees ci-dessus, redige un audit structure et complet.
Developpe chaque section avec de vraies phrases analytiques — pas seulement des puces.

### AUDIT DE RENTABILITE
APPLIQUE STRICTEMENT LE BENCHMARK DEFINI DANS LE SYSTEM PROMPT.
Ton de conseiller : felicite si marge > 6%, signale si marge < 6%, critique si marge < 0%.
Nomme les 3 pires trajets avec leurs chiffres EXACTS du fichier.
Si noms "nan" ou vides : ecris "donnee client manquante dans le fichier source".

### DIAGNOSTIC RESEAU
Analyse le cout/km vs referentiels CNR pour chaque type de trajet. Evalue la coherence
spatiale, le taux de retour a charge si visible. Cite les chiffres exacts des benchmarks.

### A FAIRE - PRIORITE ABSOLUE
Commence par l'action la plus urgente. Une phrase directe : nomme le client ou le trajet,
l'action concrete, et l'impact cash estime. Puis une alternative si trop risque commercialement.
C'est le dirigeant qui decide — donne-lui ce dont il a besoin pour agir aujourd'hui.

### PLAN DE RATIONALISATION (1-2-3)
Developpe maintenant 3 recommandations strategiques pour le moyen terme. Pour chacune :
action precise, client/trajet cible, impact attendu en EUR, difficulte d'execution 1 a 5.
Ces recommandations completent l'urgence ci-dessus et guident au-dela de cette semaine.

### SCORING LOGIFLO
- Rentabilite et Yield Transport : XX/100
- Efficacite Operationnelle : XX/100
- Maitrise des OPEX : XX/100""")
        elif view == "TERRAIN":
            parts.append("""=== TES INSTRUCTIONS ===
Ecris comme un chef de quai qui connait cet entrepot. Langage direct, pas de jargon financier.
Developpe chaque section avec de vraies phrases — pas seulement des puces.

### CE QUI EST URGENT
Nomme les references en rupture ou proches de la rupture selon les DONNEES FOURNIES.
Pour chacune : reference exacte telle qu'elle apparait dans les donnees, stock actuel exact,
pourquoi c'est critique. NE PAS inventer de quantites a commander — tu n'as pas cette information.
Dis "stock : X unites" mais JAMAIS "commander X unites" sauf si une consommation historique est disponible.
Si une reference etait deja en rupture au dernier audit (historique disponible), signale-le explicitement.

### CE QUI A CHANGE DEPUIS LE DERNIER AUDIT
SEULEMENT si des donnees d'historique sont explicitement fournies dans le contexte ci-dessus.
Si la section historique indique "Premier audit" ou "pas d'historique" : SAUTE COMPLETEMENT cette section,
n'ecris meme pas le titre.

### CE QUI DORT
References sans mouvement detectees dans ce fichier. IMPORTANT : si c'est un premier audit,
ne dis PAS "depuis 6 mois" ou toute duree — dis "aucun mouvement detecte dans ce fichier".
La duree reelle est inconnue sans historique. Pour chacune : reference exacte, stock exact, action proposee.

### A FAIRE MAINTENANT
L'action la plus urgente basee sur les donnees reelles. Reference exacte, stock exact.
Ne pas inventer de quantites de commande. Si budget serre : alternative concrète.
C'est l'equipe qui decide — pas d'ambiguite.

### TES 3 ACTIONS POUR CETTE SEMAINE
3 actions pratiques classees par urgence. Une phrase concrete chacune.
Difficulte : Facile / Moyen / Complique. Basees uniquement sur les donnees disponibles.

### EN RESUME
2-3 phrases pour briefer le responsable en 30 secondes. Termine par : situation globale : en amelioration / stable / en degradation.""")
        else:
            parts.append("""=== TES INSTRUCTIONS ===
En utilisant TOUTES les donnees ci-dessus, redige un audit structure et complet.
Developpe chaque section avec de vraies phrases analytiques — pas seulement des puces.

### DIAGNOSTIC OPERATIONNEL
Commence par ton verdict global (une phrase). Ensuite developpe : taux de service vs benchmark
sectoriel (cite l'ecart exact), analyse de la rotation, nomme les 3 references les plus critiques
avec leurs chiffres exacts et explique pourquoi elles sont critiques.

### DIAGNOSTIC FINANCIER ET STOCKS DORMANTS
Analyse le capital immobilise vs normes sectorielles.
REGLE CRITIQUE SUR LES DORMANTS : une reference est dormante UNIQUEMENT si une colonne
de consommation/ventes est presente ET que la consommation est nulle. Si aucune donnee
de consommation n'est disponible dans le fichier, il est IMPOSSIBLE d'identifier des dormants
— dis "absence de donnees de consommation : rotation non calculable".
Ne jamais qualifier une reference de dormante sur la seule base de l'absence de prix ou de ventes.

### A FAIRE - PRIORITE ABSOLUE
Commence par l'action la plus urgente. Une phrase directe : nomme la reference precise,
l'action concrete, et l'impact estime en EUR ou en pourcentage. Une alternative si budget contraint.
C'est le dirigeant qui decide — donne-lui ce dont il a besoin pour agir aujourd'hui.

### PLAN D'ACTION (1-2-3)
Developpe maintenant 3 recommandations strategiques pour le moyen terme. Pour chacune :
action precise, reference ciblee, impact attendu, difficulte d'execution 1 a 5. Ces recommandations
completent l'urgence ci-dessus et guident l'optimisation des stocks au-dela de cette semaine.

### SCORING LOGIFLO
- Performance et Rotation stock : XX/100
- Risque de rupture : XX/100
- Resilience supply chain : XX/100""")

    user_msg = "\n\n".join(parts)

    # ── Tentative 1 : OpenAI ────────────────────────────────────
    try:
        r = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": sys_prompt},
                {"role": "user",   "content": user_msg}
            ],
            temperature=0.35,
            max_tokens=2400,
            timeout=30
        )
        texte = r.choices[0].message.content
        try: return texte.encode('latin-1').decode('utf-8')
        except: return texte
    except Exception as _e_oai:
        pass  # Fallback silencieux

    # ── Tentative 2 : Gemini Flash (fallback gratuit) ────────────
    try:
        import google.generativeai as _genai
        _genai.configure(api_key=st.secrets.get("GEMINI_API_KEY",""))
        _gem = _genai.GenerativeModel("gemini-1.5-flash")
        _resp = _gem.generate_content(
            f"{sys_prompt}\n\n{user_msg}",
            generation_config=_genai.types.GenerationConfig(
                temperature=0.35,
                max_output_tokens=2400
            )
        )
        texte = _resp.text
        try: return texte.encode('latin-1').decode('utf-8')
        except: return texte
    except Exception as _e_gem:
        pass  # Mode dégradé

    # ── Mode dégradé : rapport structuré sans IA ─────────────────
    return _rapport_sans_ia(data_summary, sector_key or "generique", lang)


def _extract_key_rows(df, module, lang="fr"):
    """
    Extrait les lignes les plus significatives du dataframe brut.
    Envoie les anomalies et extremes — pas tout le fichier.
    """
    try:
        lines = []
        if lang == "en":
            lines.append("=== KEY DATA ROWS (worst performers + anomalies) ===")
        else:
            lines.append("=== LIGNES CLES DU FICHIER (pires performances + anomalies) ===")

        if module == "transport":
            if "Marge_Nette" in df.columns and "_CA" in df.columns:
                pires = df.nsmallest(5, "Marge_Nette")
                for _i, row in pires.iterrows():
                    client_col = df.columns[0]
                    lines.append(f"  - {row.get(client_col,'?')}: CA={row.get('_CA',0):.0f} EUR, Cout={row.get('_CO',0):.0f} EUR, Marge={row.get('Marge_Nette',0):.0f} EUR ({row.get('Rentabilite_%',0):.1f}%)")
        else:
            if "reference" in df.columns and "quantite" in df.columns:
                # Ruptures
                rupt = df[df["quantite"] <= 0]
                if len(rupt) > 0:
                    refs = rupt["reference"].astype(str).head(5).tolist()
                    label = "Stockouts" if lang=="en" else "Ruptures"
                    lines.append(f"  {label}: {', '.join(refs)}")
                # Dormants
                if "_conso_moy" in df.columns:
                    dorm = df[(df["quantite"] > 0) & (df["_conso_moy"] == 0)]
                    if len(dorm) > 0:
                        refs_d = dorm.nlargest(5,"quantite")["reference"].astype(str).tolist()
                        label2 = "Dormant (no consumption)" if lang=="en" else "Dormants (conso nulle)"
                        lines.append(f"  {label2}: {', '.join(refs_d)}")
                # Surstocks
                if "Couverture_mois" in df.columns:
                    surs = df[df["Couverture_mois"] > 6].nlargest(3,"valeur_totale") if "valeur_totale" in df.columns else df[df["Couverture_mois"] > 6].head(3)
                    if len(surs) > 0:
                        refs_s = surs["reference"].astype(str).tolist()
                        cov = surs["Couverture_mois"].apply(lambda x: f"{x:.0f}m" if x < 9999 else "inf").tolist()
                        label3 = "Overstock (>6 months coverage)" if lang=="en" else "Surstock (>6 mois couverture)"
                        lines.append(f"  {label3}: {', '.join([f'{r}({c})' for r,c in zip(refs_s,cov)])}")

        return "\n".join(lines) if len(lines) > 1 else ""
    except Exception:
        return ""


# =========================================
# 7. PDF
# =========================================
class PDFReport(FPDF):
    def footer(self):
        self.set_y(-15);self.set_font("Arial","I",8);self.set_text_color(150,150,150)
        footer_text=_("pdf_footer")
        self.multi_cell(0,4,_s(footer_text),align="C")

def _s(text):
    """
    Safe string pour fpdf — garantit 100% compatibilité latin-1.
    Remplace explicitement les caractères courants avant normalisation.
    """
    if text is None: return ""
    text = str(text)
    replacements = {
        "’":"'","‘":"'","“":'"',"”":'"',
        "–":"-","—":"-","…":"...","°":"deg",
        "€":"EUR","£":"GBP","©":"(c)","®":"(R)",
        "™":"TM","•":"-","‣":"-","●":"-",
        "→":"->","←":"<-","⇒":"=>","✓":"OK",
        "✔":"OK","✗":"X","✘":"X",
        "±":"+/-","×":"x","÷":"/",
        "≈":"~","≠":"!=","≤":"<=","≥":">=",
        "🔴":"[!]","🟠":"[!]","🟢":"[ok]",
        "📊":"","📈":"","📉":"",
        "⚠":"[!]","ℹ":"[i]","★":"*","☆":"*",
    }
    for char,repl in replacements.items():
        text = text.replace(char,repl)
    text = unicodedata.normalize('NFKD',text).encode('ASCII','ignore').decode('utf-8')
    try:
        text.encode('latin-1')
        return text
    except UnicodeEncodeError:
        return text.encode('latin-1',errors='ignore').decode('latin-1')

# Alias pour compatibilité
def _asc(text): return _s(text)

def _clean_pdf(text):
    """Nettoie le texte pour fpdf — utilise _s() pour garantie latin-1."""
    return _s(str(text).replace("**",""))



def compute_logiflo_score(module, df=None, kpis=None, labels=None,
                           sector_key="generique", lang="fr"):
    """
    Scoring Logiflo calculé en Python sur critères transparents.
    L'IA commente le score — elle ne le génère plus.
    Retourne dict avec scores par dimension et score global.
    """
    scores = {}
    details = {}

    if module == "stock":
        # Extraire les KPIs
        try:
            tx_service = float(kpis[1]) if kpis and len(kpis) > 1 else 0
            nb_ruptures = float(kpis[2]) if kpis and len(kpis) > 2 else 0
            nb_total = len(df) if df is not None and len(df) > 0 else 1
            taux_rupture = (nb_ruptures / nb_total) * 100 if nb_total > 0 else 0
        except Exception:
            tx_service = 0; taux_rupture = 0

        # Benchmarks cibles selon secteur
        target_service = {
            "stock_pharma":97, "stock_industrie":97, "stock_retail":96,
            "stock_distribution":95, "stock_agroalim":96, "stock_btp":95,
            "generique":93
        }.get(sector_key, 93)

        # DIMENSION 1 : Taux de service (40%)
        if tx_service >= target_service:
            s1 = 100
        elif tx_service >= target_service - 5:
            s1 = 80
        elif tx_service >= target_service - 10:
            s1 = 60
        elif tx_service >= 80:
            s1 = 40
        else:
            s1 = 20
        scores["service"] = s1
        d1_lbl = "Performance & Rotation stock" if lang == "en" else "Performance et Rotation stock"
        details[d1_lbl] = s1

        # DIMENSION 2 : Risque rupture (35%)
        if taux_rupture <= 1:
            s2 = 100
        elif taux_rupture <= 3:
            s2 = 80
        elif taux_rupture <= 5:
            s2 = 60
        elif taux_rupture <= 10:
            s2 = 40
        else:
            s2 = 20
        scores["rupture"] = s2
        d2_lbl = "Stock-out Risk" if lang == "en" else "Risque de rupture"
        details[d2_lbl] = s2

        # DIMENSION 3 : Résilience (25%) — dormants + surstocks
        if df is not None:
            try:
                nb_dorm = len(df[df["Statut"].str.contains("Dormant", na=False)]) if "Statut" in df.columns else 0
                nb_surs = len(df[df["Statut"].str.contains("Surstock", na=False)]) if "Statut" in df.columns else 0
                taux_anomalies = ((nb_dorm + nb_surs) / max(nb_total, 1)) * 100
                if taux_anomalies <= 5:
                    s3 = 100
                elif taux_anomalies <= 10:
                    s3 = 75
                elif taux_anomalies <= 20:
                    s3 = 50
                else:
                    s3 = 25
            except Exception:
                s3 = 70
        else:
            s3 = 70
        scores["resilience"] = s3
        d3_lbl = "Supply Chain Resilience" if lang == "en" else "Resilience supply chain"
        details[d3_lbl] = s3

        # Score global pondéré
        global_score = round(s1 * 0.40 + s2 * 0.35 + s3 * 0.25)

    elif module == "transport":
        try:
            marge_pct = float(kpis[1]) if kpis and len(kpis) > 1 else 0
            nb_tox = float(kpis[2]) if kpis and len(kpis) > 2 else 0
            nb_total = len(df) if df is not None and len(df) > 0 else 1
            taux_tox = (nb_tox / nb_total) * 100 if nb_total > 0 else 0
        except Exception:
            marge_pct = 0; taux_tox = 0

        # Coût/km si disponible
        cout_km = 0
        if df is not None and "_DS" in df.columns and "_CO" in df.columns:
            try:
                total_dist = df["_DS"].replace(0, 1).sum()
                total_cout = df["_CO"].sum()
                cout_km = total_cout / total_dist if total_dist > 0 else 0
            except Exception:
                pass

        # DIMENSION 1 : Rentabilité et Yield (40%)
        if marge_pct >= 10:
            s1 = 100
        elif marge_pct >= 8:
            s1 = 80
        elif marge_pct >= 6:
            s1 = 60
        elif marge_pct >= 4:
            s1 = 40
        elif marge_pct >= 0:
            s1 = 20
        else:
            s1 = 5
        scores["rentabilite"] = s1
        d1_lbl = "Profitability and Transport Yield" if lang == "en" else "Rentabilite et Yield Transport"
        details[d1_lbl] = s1

        # DIMENSION 2 : Efficacité opérationnelle (35%)
        if taux_tox <= 5:
            s2 = 100
        elif taux_tox <= 10:
            s2 = 75
        elif taux_tox <= 20:
            s2 = 50
        elif taux_tox <= 35:
            s2 = 30
        else:
            s2 = 10
        scores["efficacite"] = s2
        d2_lbl = "Operational Efficiency" if lang == "en" else "Efficacite Operationnelle"
        details[d2_lbl] = s2

        # DIMENSION 3 : Maîtrise OPEX / coût/km (25%)
        cnr_ref = 1.95  # CNR 2026 longue distance médiane
        if cout_km <= 0:
            s3 = 70  # pas de données km → score neutre
        elif cout_km <= cnr_ref:
            s3 = 100
        elif cout_km <= cnr_ref * 1.10:
            s3 = 80
        elif cout_km <= cnr_ref * 1.25:
            s3 = 60
        elif cout_km <= cnr_ref * 1.50:
            s3 = 35
        else:
            s3 = 15
        scores["opex"] = s3
        d3_lbl = "OPEX Control" if lang == "en" else "Maitrise des OPEX"
        details[d3_lbl] = s3

        global_score = round(s1 * 0.40 + s2 * 0.35 + s3 * 0.25)

    else:
        # Module générique
        scores = {"performance": 70}
        details = {"Performance": 70}
        global_score = 70

    return {
        "global":  global_score,
        "details": details,
        "scores":  scores,
        "format_pdf": "\n".join([f"- {k} : {v}/100" for k, v in details.items()])
    }



def predict_ruptures(df, seuil_rupture=0, lang="fr"):
    """
    Prédit les ruptures à 4 semaines par extrapolation linéaire.
    Nécessite au moins une colonne de consommation historique.
    Retourne une liste de dict avec les références à risque.
    """
    if df is None or len(df) == 0:
        return []

    alertes = []
    cols_conso = [c for c in ["conso_an4","conso_an3","conso_an2","conso_an1"]
                  if c in df.columns]

    if not cols_conso:
        return []

    try:
        for _ri, row in df.iterrows():
            ref = str(row.get("reference","?"))
            stock = float(row.get("quantite", 0))
            if stock <= seuil_rupture:
                continue  # déjà en rupture, pas besoin de prédire

            # Consommation hebdomadaire moyenne
            consos = []
            for c in cols_conso:
                v = row.get(c, 0)
                try:
                    v = float(v)
                    if v > 0:
                        consos.append(v / 52)  # annuel → hebdo
                except Exception:
                    pass

            if not consos:
                continue

            conso_hebdo = sum(consos) / len(consos)
            if conso_hebdo <= 0:
                continue

            # Semaines avant rupture
            semaines = stock / conso_hebdo

            if semaines <= 4:
                urgence = "critique" if semaines <= 1 else (
                          "urgent" if semaines <= 2 else "alerte")
                alertes.append({
                    "reference":   ref,
                    "stock":       stock,
                    "conso_hebdo": round(conso_hebdo, 1),
                    "semaines":    round(semaines, 1),
                    "urgence":     urgence,
                })

        # Trier par urgence (les plus proches de la rupture en premier)
        alertes.sort(key=lambda x: x["semaines"])
        return alertes[:10]  # top 10 max

    except Exception:
        return []


def format_predictions_pour_prompt(alertes, lang="fr"):
    """Formate les prédictions pour injection dans le prompt IA."""
    if not alertes:
        return ""
    lines = []
    if lang == "en":
        lines.append("=== STOCKOUT PREDICTIONS (next 4 weeks) ===")
        for a in alertes:
            lines.append(
                f"  {a['urgence'].upper()} — {a['reference']}: "
                f"{a['stock']:.0f} units left, "
                f"consumption {a['conso_hebdo']:.1f}/week, "
                f"stockout in ~{a['semaines']:.1f} weeks"
            )
    else:
        lines.append("=== PREDICTIONS RUPTURE (4 prochaines semaines) ===")
        for a in alertes:
            lines.append(
                f"  {a['urgence'].upper()} — {a['reference']}: "
                f"{a['stock']:.0f} unites restantes, "
                f"conso {a['conso_hebdo']:.1f}/semaine, "
                f"rupture dans ~{a['semaines']:.1f} semaines"
            )
    return "\n".join(lines)



def compute_alerte_bfr(df, ca_annuel_estime=None, lang="fr"):
    """
    Calcule le BFR stock et identifie le capital libérable.
    Retourne un dict avec les métriques BFR et le texte d'alerte.
    """
    result = {"available": False, "texte": "", "capital_liberatable": 0}

    if df is None or len(df) == 0:
        return result

    try:
        # Capital immobilisé total
        if "valeur_totale" in df.columns:
            capital_total = df["valeur_totale"].sum()
        elif "quantite" in df.columns and "prix_unitaire" in df.columns:
            capital_total = (df["quantite"] * df["prix_unitaire"]).sum()
        else:
            return result

        if capital_total <= 0:
            return result

        # Capital dormant (stock sans mouvement)
        capital_dormant = 0
        if "Statut" in df.columns and "valeur_totale" in df.columns:
            mask_dorm = df["Statut"].str.contains("Dormant", na=False)
            capital_dormant = df.loc[mask_dorm, "valeur_totale"].sum()
        elif "Statut" in df.columns and "quantite" in df.columns and "prix_unitaire" in df.columns:
            mask_dorm = df["Statut"].str.contains("Dormant", na=False)
            capital_dormant = (df.loc[mask_dorm,"quantite"] * df.loc[mask_dorm,"prix_unitaire"]).sum()

        # Capital surstock
        capital_surstock = 0
        if "Statut" in df.columns and "valeur_totale" in df.columns:
            mask_surs = df["Statut"].str.contains("Surstock", na=False)
            capital_surstock = df.loc[mask_surs, "valeur_totale"].sum()

        # Capital libérable (dormants + 50% surstock)
        capital_lib = capital_dormant + capital_surstock * 0.5

        # BFR en jours de CA (si CA disponible ou estimé)
        bfr_jours = None
        if ca_annuel_estime and ca_annuel_estime > 0:
            bfr_jours = round((capital_total / ca_annuel_estime) * 365)

        # Coût de possession (20% par an)
        cout_possession = round(capital_dormant * 0.20)

        result["available"] = True
        result["capital_liberatable"] = round(capital_lib)
        result["capital_dormant"] = round(capital_dormant)
        result["capital_surstock"] = round(capital_surstock)
        result["cout_possession"] = cout_possession
        result["bfr_jours"] = bfr_jours

        if lang == "en":
            result["texte"] = (
                f"BFR ALERT: {capital_lib:,.0f} EUR of capital can be freed "
                f"(dormant stock: {capital_dormant:,.0f} EUR + overstock: {capital_surstock:,.0f} EUR). "
                f"Annual holding cost of dormant stock: ~{cout_possession:,.0f} EUR/year (20% rate). "
                + (f"WCR: ~{bfr_jours} days of revenue." if bfr_jours else "")
            )
        else:
            result["texte"] = (
                f"ALERTE BFR : {capital_lib:,.0f} EUR de capital liberables "
                f"(stock dormant : {capital_dormant:,.0f} EUR + surstock : {capital_surstock:,.0f} EUR). "
                f"Cout de possession du stock dormant : ~{cout_possession:,.0f} EUR/an (taux 20%). "
                + (f"BFR : ~{bfr_jours} jours de CA." if bfr_jours else "")
            )

    except Exception:
        pass

    return result



def _rapport_sans_ia(data_summary, sector_key, lang="fr"):
    """Rapport de secours quand toutes les IA sont indisponibles."""
    benchmarks = get_sector_benchmarks(sector_key or "generique", lang)
    if lang == "en":
        return f"""### AUTOMATIC DIAGNOSIS

AI analysis is temporarily unavailable. Here are the raw computed metrics.

**Computed data:**
{data_summary}

**Sector benchmarks for reference:**
{benchmarks}

### WHAT TO DO - TOP PRIORITY
Compare your indicators to the benchmarks above.
Any negative gap above 5 points requires action this week.

*Full AI analysis available in a few minutes — relaunch the audit if needed.*"""
    else:
        return f"""### DIAGNOSTIC AUTOMATIQUE

L'analyse IA est temporairement indisponible. Voici les metriques calculees.

**Donnees calculees :**
{data_summary}

**Benchmarks sectoriels de reference :**
{benchmarks}

### A FAIRE - PRIORITE ABSOLUE
Comparez vos indicateurs aux benchmarks ci-dessus.
Tout ecart negatif de plus de 5 points merite une action cette semaine.

*Analyse IA complete disponible dans quelques minutes — relancez si necessaire.*"""



def render_prediction_rupture(df, lang="fr"):
    """
    Affiche les alertes de rupture prédites visuellement.
    S'insère entre les KPIs et l'analyse IA dans l'audit stock.
    """
    alertes = predict_ruptures(df, lang=lang)
    if not alertes:
        return

    _lbl = "Prédictions de rupture — 4 semaines" if lang=="fr" else "Stockout Predictions — 4 weeks"
    st.markdown(f"""
    <div style="font-size:11px;font-weight:700;color:#4A6080;
                letter-spacing:2px;text-transform:uppercase;margin-bottom:10px;
                margin-top:20px;">
        ⚠️ {_lbl}
    </div>
    """, unsafe_allow_html=True)

    for a in alertes:
        _clr  = "#E8304A" if a["urgence"]=="critique" else (
                "#F39C12" if a["urgence"]=="urgent" else "#F59E0B")
        _bg   = "#FFF1F2" if a["urgence"]=="critique" else (
                "#FFFBEB" if a["urgence"]=="urgent" else "#FEFCE8")
        _icon = "🔴" if a["urgence"]=="critique" else (
                "🟠" if a["urgence"]=="urgent" else "🟡")
        _semaines = a["semaines"]
        _ref   = str(a["reference"])
        _stock = f"{a['stock']:.0f}"
        _conso = f"{a['conso_hebdo']:.1f}"

        if lang=="en":
            _msg = (f"<strong>{_ref}</strong> — {_stock} units left · "
                    f"consumption {_conso}/week · "
                    f"<strong>stockout in ~{_semaines:.1f} weeks</strong>")
        else:
            _msg = (f"<strong>{_ref}</strong> — {_stock} unités restantes · "
                    f"conso {_conso}/sem · "
                    f"<strong>rupture dans ~{_semaines:.1f} semaines</strong>")

        st.markdown(f"""
        <div style="background:{_bg};border:1px solid {_clr}30;
                    border-left:4px solid {_clr};border-radius:8px;
                    padding:10px 14px;margin-bottom:6px;
                    display:flex;align-items:center;gap:10px;">
            <span style="font-size:18px;flex-shrink:0;">{_icon}</span>
            <div style="flex:1;font-size:12px;color:#0B2545;line-height:1.4;">
                {_msg}
            </div>
            <div style="font-size:10px;font-weight:700;color:{_clr};
                        text-transform:uppercase;letter-spacing:1px;
                        flex-shrink:0;">
                {a["urgence"]}
            </div>
        </div>
        """, unsafe_allow_html=True)



# ══════════════════════════════════════════════════════════════════
# TOOLTIPS SUR LES METRIQUES KPI
# ══════════════════════════════════════════════════════════════════
KPI_TOOLTIPS = {
    "fr": {
        "Taux de service":      ("Pourcentage de commandes livrées sans rupture.", "Benchmark : > 93% B2B · > 96% B2C"),
        "Taux Service":         ("Pourcentage de commandes livrées sans rupture.", "Benchmark : > 93% B2B · > 96% B2C"),
        "Capital immobilisé":   ("Valeur totale du stock en EUR. Un capital trop élevé pèse sur la trésorerie.", "Norme : < 60 jours de CA"),
        "Capital Immobilisé":   ("Valeur totale du stock en EUR.", "Norme : < 60 jours de CA"),
        "Ruptures":             ("Références dont le stock est à zéro ou sous le seuil d'alerte défini.", "Objectif : < 2% des références"),
        "Marge nette":          ("CA moins les coûts d'exploitation (carburant, chauffeur, péages).", "Norme PME transport : 6-10% · Excellent : > 10%"),
        "Marge Nette":          ("CA moins les coûts d'exploitation.", "Norme PME transport : 6-10%"),
        "Taux rentabilité":     ("Pourcentage de trajets avec une marge positive.", "Objectif : > 80% des trajets sains"),
        "Trajets toxiques":     ("Trajets avec une marge < 5% ou négative.", "Alerte si > 20% du portefeuille"),
        "BFR":                  ("Besoin en Fonds de Roulement : capital bloqué dans votre cycle d'exploitation.", "Cible : < 45 jours de CA"),
        "Score Logiflo":        ("Score composite sur 3 dimensions calculé côté Python.", "Vert ≥ 70 · Orange ≥ 40 · Rouge < 40"),
    },
    "en": {
        "Service Level":        ("Percentage of orders fulfilled without stockout.", "Benchmark: > 93% B2B · > 96% B2C"),
        "Tied-up Capital":      ("Total stock value in EUR. Excess capital strains cash flow.", "Norm: < 60 days revenue"),
        "Stockouts":            ("References at zero stock or below the defined alert threshold.", "Target: < 2% of references"),
        "Net Margin":           ("Revenue minus operating costs (fuel, driver, tolls).", "SME transport norm: 6-10% · Excellent: > 10%"),
        "Profitability Rate":   ("Percentage of routes with positive margin.", "Target: > 80% healthy routes"),
        "Toxic Routes":         ("Routes with margin < 5% or negative.", "Alert if > 20% of portfolio"),
        "WCR":                  ("Working Capital Requirement: capital locked in your operating cycle.", "Target: < 45 days revenue"),
        "Logiflo Score":        ("Composite score on 3 dimensions — calculated in Python.", "Green ≥ 70 · Orange ≥ 40 · Red < 40"),
    }
}

def tooltip_metric(label, value, unit="", delta=None, lang="fr"):
    """
    Affiche une métrique avec tooltip au survol (ligne pointillée).
    Wrapper autour de st.metric avec HTML enrichi.
    """
    tips = KPI_TOOLTIPS.get(lang, KPI_TOOLTIPS["fr"])
    tip = tips.get(label)

    if tip:
        def_txt, bench_txt = tip
        tooltip_html = f"""
        <div style="position:relative;display:inline-block;">
            <span style="font-size:11px;color:#4A6080;
                         border-bottom:1.5px dashed #4A6080;
                         cursor:help;"
                  title="{def_txt} | {bench_txt}">
                {label}
            </span>
        </div>
        """
    else:
        tooltip_html = f'<span style="font-size:11px;color:#4A6080;">{label}</span>'

    # Valeur formatée
    if isinstance(value, float):
        val_str = f"{value:,.1f}"
    elif isinstance(value, int):
        val_str = f"{value:,}"
    else:
        val_str = str(value)

    color = "#00C896"
    if delta is not None:
        color = "#00C896" if delta >= 0 else "#E8304A"

    st.markdown(f"""
    <div style="background:white;border:1px solid #E2E8F0;
                border-top:3px solid {color};border-radius:12px;
                padding:14px 16px;text-align:center;">
        {tooltip_html}
        <div style="font-family:Syne,sans-serif;font-size:26px;font-weight:800;
                    color:{color};margin-top:4px;line-height:1;">
            {val_str}
            <span style="font-size:13px;font-weight:400;color:#4A6080;">{unit}</span>
        </div>
    </div>
    """, unsafe_allow_html=True)



def generate_exemple_excel():
    """
    Génère un fichier Excel exemple téléchargeable depuis l'accueil.
    20 lignes réalistes — fonctionne avec Stock et Transport.
    """
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None

    wb = openpyxl.Workbook()

    # ── Onglet Stock ──────────────────────────────────────────────
    ws_stock = wb.active
    ws_stock.title = "Stock"

    headers_stock = ["Reference","Designation","Fournisseur","Prix_Achat",
                     "Prix_Vente","Stock","Conso_2023","Conso_2024","Conso_2025",
                     "Categorie","Date_Entree"]
    for ci, h in enumerate(headers_stock, 1):
        cell = ws_stock.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B2545")
        cell.alignment = Alignment(horizontal="center")

    data_stock = [
        ("REF-001","T-shirt Coton Bio","FournisseurA",8.5,22.0,45,120,135,98,"Textile","15/01/2025"),
        ("REF-002","Pantalon Slim","FournisseurB",15.0,45.0,12,80,75,62,"Textile","20/01/2025"),
        ("REF-003","Chaussures Sport","FournisseurC",25.0,89.0,0,60,55,48,"Chaussures","10/02/2025"),
        ("REF-004","Veste Laine","FournisseurA",35.0,120.0,3,40,38,29,"Textile","05/03/2025"),
        ("REF-005","Echarpe Cachemire","FournisseurD",18.0,65.0,28,55,50,44,"Accessoires","12/03/2025"),
        ("REF-006","Sac a Dos","FournisseurB",12.0,38.0,0,90,95,88,"Accessoires","08/04/2025"),
        ("REF-007","Bonnet Laine","FournisseurA",5.5,18.0,67,200,185,170,"Accessoires","02/01/2025"),
        ("REF-008","Pull Cachemire","FournisseurD",45.0,150.0,8,30,25,0,"Textile","18/02/2025"),
        ("REF-009","Jean Denim","FournisseurC",20.0,65.0,34,110,105,98,"Textile","25/01/2025"),
        ("REF-010","Chemise Oxford","FournisseurB",18.0,55.0,15,70,68,62,"Textile","30/01/2025"),
        ("REF-011","Robe Ete","FournisseurA",12.0,40.0,52,150,140,130,"Textile","10/04/2025"),
        ("REF-012","Short Sport","FournisseurC",8.0,25.0,75,180,175,160,"Sport","15/04/2025"),
        ("REF-013","Manteau Long","FournisseurD",65.0,220.0,2,20,18,15,"Textile","01/02/2025"),
        ("REF-014","Ceinture Cuir","FournisseurB",9.0,29.0,40,95,90,85,"Accessoires","20/02/2025"),
        ("REF-015","Gants Hiver","FournisseurA",7.0,22.0,22,130,125,110,"Accessoires","05/11/2024"),
        ("REF-016","Polo Coton","FournisseurC",10.0,32.0,0,100,95,80,"Textile","28/01/2025"),
        ("REF-017","Legging Sport","FournisseurB",12.0,38.0,88,160,155,148,"Sport","12/04/2025"),
        ("REF-018","Chaussettes x3","FournisseurA",3.5,12.0,150,300,290,270,"Accessoires","08/01/2025"),
        ("REF-019","Sweat Zip","FournisseurD",22.0,72.0,18,85,80,72,"Textile","22/02/2025"),
        ("REF-020","Casquette","FournisseurC",6.0,19.0,95,220,210,195,"Accessoires","15/01/2025"),
    ]
    for ri, row in enumerate(data_stock, 2):
        for ci, val in enumerate(row, 1):
            ws_stock.cell(row=ri, column=ci, value=val)

    # ── Onglet Transport ──────────────────────────────────────────
    ws_trans = wb.create_sheet("Transport")
    headers_trans = ["Client","Depart","Arrivee","Distance_km","CA_EUR",
                     "Cout_EUR","Poids_kg","Date_Livraison","Mode"]
    for ci, h in enumerate(headers_trans, 1):
        cell = ws_trans.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B2545")
        cell.alignment = Alignment(horizontal="center")

    data_trans = [
        ("Carrefour","Marseille","Lyon",315,450,385,1200,"05/04/2025","Routier"),
        ("Auchan","Paris","Bordeaux",580,820,710,2500,"06/04/2025","Routier"),
        ("Lidl","Lyon","Toulouse",295,420,368,900,"07/04/2025","Routier"),
        ("BHV","Marseille","Nice",200,280,245,600,"07/04/2025","Routier"),
        ("Leclerc","Bordeaux","Paris",580,750,680,1800,"08/04/2025","Routier"),
        ("Fnac","Paris","Lille",225,310,272,800,"08/04/2025","Routier"),
        ("Decathlon","Lyon","Paris",465,640,575,2200,"09/04/2025","Routier"),
        ("Ikea","Marseille","Montpellier",170,245,215,1500,"09/04/2025","Routier"),
        ("Zara","Paris","Strasbourg",490,680,595,1100,"10/04/2025","Routier"),
        ("H&M","Lille","Bruxelles",105,165,148,500,"10/04/2025","Routier"),
        ("Michelin","Clermont","Lyon",175,240,215,3000,"11/04/2025","Routier"),
        ("Renault","Paris","Le Havre",200,285,250,5000,"11/04/2025","Routier"),
        ("Total","Marseille","Fos",45,95,88,8000,"12/04/2025","Routier"),
        ("DHL","Roissy","Lyon",465,580,510,1800,"12/04/2025","Routier"),
        ("GlobeTrans","Marseille","Barcelona",820,1100,985,2200,"13/04/2025","Routier"),
        ("FalconDist","Paris","Amsterdam",510,720,640,1500,"13/04/2025","Routier"),
        ("BetaLog","Lyon","Geneve",155,195,188,900,"14/04/2025","Routier"),
        ("AlphaFret","Marseille","Turin",450,610,545,2800,"14/04/2025","Routier"),
        ("SudLog","Toulouse","Madrid",1050,1450,1280,3500,"15/04/2025","Routier"),
        ("NordTrans","Lille","Hamburg",680,920,810,2100,"15/04/2025","Routier"),
    ]
    for ri, row in enumerate(data_trans, 2):
        for ci, val in enumerate(row, 1):
            ws_trans.cell(row=ri, column=ci, value=val)

    # ── Onglet Guide ──────────────────────────────────────────────
    ws_guide = wb.create_sheet("Guide")
    ws_guide.cell(1,1,"GUIDE — Colonnes reconnues par Logiflo").font = Font(bold=True, size=14)
    guide_lines = [
        (3,"ONGLET STOCK",""),
        (4,"Reference","Code article ou SKU — requis"),
        (5,"Designation","Nom du produit — optionnel"),
        (6,"Fournisseur","Nom du fournisseur — optionnel"),
        (7,"Prix_Achat","Prix d'achat unitaire en EUR — pour le calcul BFR"),
        (8,"Prix_Vente","Prix de vente unitaire en EUR — pour la marge"),
        (9,"Stock","Quantité en stock actuelle — requis"),
        (10,"Conso_2023/2024/2025","Consommation annuelle — pour la rotation et la prédiction rupture"),
        (11,"Categorie","Famille ou catégorie produit — optionnel"),
        (12,"Date_Entree","Date d'entrée en stock — pour la saisonnalité"),
        (14,"ONGLET TRANSPORT",""),
        (15,"Client","Nom du client ou destinataire — requis"),
        (16,"Depart","Ville ou code de départ — requis"),
        (17,"Arrivee","Ville ou code d'arrivée — requis"),
        (18,"Distance_km","Distance en km — calculée automatiquement si absent"),
        (19,"CA_EUR","Chiffre d'affaires du trajet en EUR — requis"),
        (20,"Cout_EUR","Coût du trajet en EUR — requis"),
        (21,"Poids_kg","Poids total chargé — optionnel"),
        (22,"Date_Livraison","Date de livraison — pour la saisonnalité"),
        (23,"Mode","Routier / Maritime / Aerien / Ferroviaire — détecté auto"),
    ]
    for row_num, label, desc in guide_lines:
        ws_guide.cell(row_num, 1, label).font = Font(bold=(not desc))
        if desc:
            ws_guide.cell(row_num, 2, desc)

    # Largeurs colonnes
    for ws in [ws_stock, ws_trans, ws_guide]:
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_free_pdf(module, summary_text, kpis, labels):
    """PDF ultra-leger audit gratuit — 2 pages, sans graphiques."""
    pdf = PDFReport()
    lang = st.session_state.get("language","fr")

    # Page 1 — Couverture
    pdf.add_page()
    pdf.set_fill_color(11,37,69); pdf.rect(0,0,210,297,'F')
    pdf.set_fill_color(0,200,150); pdf.rect(0,0,210,6,'F')
    pdf.set_y(90)
    pdf.set_text_color(255,255,255); pdf.set_font("Arial","B",34)
    pdf.cell(0,16,"LOGIFLO.IO",ln=True,align='C')
    pdf.set_font("Arial","",13); pdf.set_text_color(0,200,150)
    lbl_free = "Free Audit" if lang=="en" else "Audit Gratuit"
    pdf.cell(0,10,_s(f"[ {lbl_free} ]"),ln=True,align='C')
    pdf.ln(10)
    pdf.set_draw_color(0,200,150); pdf.set_line_width(0.6)
    pdf.line(50,pdf.get_y(),160,pdf.get_y()); pdf.ln(10)
    pdf.set_text_color(200,220,255); pdf.set_font("Arial","",12)
    pdf.cell(0,8,_s(f"Date : {datetime.date.today().strftime('%d/%m/%Y')}"),ln=True,align='C')
    mod_label = "STOCK" if module=="stock" else "TRANSPORT"
    pdf.cell(0,8,_s(f"Module : {mod_label}"),ln=True,align='C')
    pdf.set_fill_color(0,200,150); pdf.rect(0,291,210,6,'F')

    # Page 2 — Resultats
    pdf.add_page()
    pdf.set_fill_color(11,37,69); pdf.rect(0,0,210,18,'F')
    pdf.set_y(4); pdf.set_text_color(255,255,255); pdf.set_font("Arial","B",11)
    res_label = "YOUR AUDIT RESULTS" if lang=="en" else "RESULTATS DE VOTRE AUDIT"
    pdf.cell(0,10,_s(res_label),ln=True,align='C'); pdf.ln(8)

    # KPI cards
    if kpis and labels:
        n=min(len(kpis),len(labels),3)
        card_w=56; total_w=n*card_w+(n-1)*8; start_x=(210-total_w)/2
        card_y=pdf.get_y()
        for i in range(n):
            cx=start_x+i*(card_w+8)
            pdf.set_fill_color(240,244,248); pdf.rect(cx,card_y,card_w,34,'F')
            pdf.set_fill_color(0,168,122); pdf.rect(cx,card_y,card_w,3,'F')
            pdf.set_xy(cx+2,card_y+5)
            pdf.set_font("Arial","",7); pdf.set_text_color(74,96,128)
            pdf.cell(card_w-4,6,_s(labels[i]).upper()[:22],align='C')
            pdf.set_xy(cx+2,card_y+13)
            pdf.set_font("Arial","B",15); pdf.set_text_color(11,37,69)
            val=kpis[i]
            if isinstance(val,float) and abs(val)>=1000: vs=_s(f"{val:,.0f}")
            elif isinstance(val,float): vs=_s(f"{val:.1f}%")
            else: vs=_s(str(val))
            pdf.cell(card_w-4,10,vs,align='C')
        pdf.ln(42)

    # Analyse rapide
    pdf.set_font("Arial","B",11); pdf.set_text_color(11,37,69)
    diag_label = "RAPID DIAGNOSIS" if lang=="en" else "DIAGNOSTIC RAPIDE"
    pdf.cell(0,8,_s(diag_label),ln=True)
    pdf.set_draw_color(0,200,150); pdf.line(10,pdf.get_y(),200,pdf.get_y()); pdf.ln(4)
    lines_done=0
    for line in summary_text.split("\n"):
        line=line.strip()
        if not line: pdf.ln(2); continue
        if pdf.get_y()>258: break
        if line.startswith("### "):
            pdf.set_font("Arial","B",10); pdf.set_text_color(0,168,122)
            pdf.cell(0,7,_s(line[4:].upper()),ln=True); lines_done+=1
        elif lines_done < 25:
            pdf.set_font("Arial","",9); pdf.set_text_color(40,40,40)
            pdf.set_x(10); pdf.multi_cell(190,5,_s(line.replace("**",""))); lines_done+=1

    # CTA
    pdf.set_y(260)
    pdf.set_fill_color(240,244,248); pdf.rect(10,pdf.get_y(),190,28,'F')
    pdf.set_fill_color(0,200,150); pdf.rect(10,pdf.get_y(),3,28,'F')
    pdf.set_xy(16,pdf.get_y()+5)
    pdf.set_font("Arial","B",10); pdf.set_text_color(11,37,69)
    cta1 = "Get the full audit with history, charts and scoring." if lang=="en" else "Obtenez l'audit complet avec historique, graphiques et scoring."
    pdf.multi_cell(184,6,_s(cta1))
    pdf.set_x(16); pdf.set_font("Arial","",9); pdf.set_text_color(74,96,128)
    pdf.cell(0,6,"logiflo-io.streamlit.app  |  contact@logiflo.io",ln=True)

    return pdf.output(dest='S').encode('latin-1',errors='replace')


def generate_expert_pdf(title, content, figs=None, kpis=None, labels=None, module="stock"):
    """
    PDF structuré 5 pages :
    P1 Couverture | P2 Synthese executive | P3 Graphiques | P4 Analyse IA | P5 CTA
    """
    if kpis is None: kpis = []
    if labels is None: labels = []
    pdf = PDFReport()
    lang = st.session_state.get("language","fr")

    # ── PAGE 1 : COUVERTURE ──────────────────────────────────────
    pdf.add_page()
    pdf.set_fill_color(11,37,69); pdf.rect(0,0,210,297,'F')
    pdf.set_fill_color(0,200,150); pdf.rect(0,0,210,6,'F')
    pdf.set_y(80)
    pdf.set_text_color(255,255,255)
    pdf.set_font("Arial","B",38); pdf.cell(0,18,"LOGIFLO.IO",ln=True,align='C')
    pdf.set_font("Arial","",14); pdf.set_text_color(0,200,150)
    pdf.cell(0,10,"[ Logistics Intelligence ]",ln=True,align='C')
    pdf.ln(8)
    pdf.set_draw_color(0,200,150); pdf.set_line_width(0.8)
    pdf.line(40,pdf.get_y(),170,pdf.get_y()); pdf.ln(10)
    # Nom entreprise et numéro rapport
    _cpdf = st.session_state.get("company_name","")
    _updf = st.session_state.get("current_user","")
    _rnum = 0
    try:
        _ap = load_archives_from_sheets(_updf)
        if _ap is not None and not _ap.empty and "module" in _ap.columns:
            _rnum = len(_ap[_ap["module"]==module]) + 1
    except Exception:
        pass

    pdf.set_text_color(255,255,255); pdf.set_font("Arial","B",22)
    pdf.multi_cell(0,12,_s(title),align='C'); pdf.ln(4)
    if _cpdf:
        pdf.set_font("Arial","",14); pdf.set_text_color(0,200,150)
        pdf.cell(0,9,_s(_cpdf),ln=True,align='C')
    if _rnum > 0:
        pdf.set_font("Arial","",11); pdf.set_text_color(180,200,220)
        pdf.cell(0,7,_s(f"Rapport #{_rnum}"),ln=True,align='C')
    pdf.ln(4)
    pdf.set_font("Arial","",12); pdf.set_text_color(180,200,220)
    conf = "CONFIDENTIAL" if lang=="en" else "CONFIDENTIEL"
    pdf.cell(0,8,_s(f"Date : {datetime.date.today().strftime('%d/%m/%Y')}"),ln=True,align='C')
    pdf.cell(0,8,_s(conf),ln=True,align='C')
    pdf.set_fill_color(0,200,150); pdf.rect(0,291,210,6,'F')

    # ── PAGE 2 : SYNTHESE EXECUTIVE ──────────────────────────────
    pdf.add_page()
    pdf.set_fill_color(11,37,69); pdf.rect(0,0,210,18,'F')
    pdf.set_y(4); pdf.set_text_color(255,255,255); pdf.set_font("Arial","B",11)
    h2 = "EXECUTIVE SUMMARY" if lang=="en" else "SYNTHESE EXECUTIVE"
    pdf.cell(0,10,_s(h2),ln=True,align='C'); pdf.ln(8)

    # Titre
    pdf.set_text_color(11,37,69); pdf.set_font("Arial","B",16)
    kpi_title = "Key Indicators" if lang=="en" else "Indicateurs Cles"
    pdf.cell(0,10,_s(kpi_title),ln=True,align='L')
    pdf.set_draw_color(0,200,150); pdf.set_line_width(0.6)
    pdf.line(10,pdf.get_y(),200,pdf.get_y()); pdf.ln(8)

    # KPI cards
    if kpis and labels:
        n = min(len(kpis),len(labels),3)
        card_w = 56
        total_w = n*card_w+(n-1)*8
        start_x = (210-total_w)/2
        card_colors = [(0,168,122),(0,168,122),(232,48,74)]
        card_y = pdf.get_y()
        for i in range(n):
            cx = start_x + i*(card_w+8)
            pdf.set_fill_color(240,244,248); pdf.rect(cx,card_y,card_w,38,'F')
            r,g,b = card_colors[i] if i < len(card_colors) else (0,168,122)
            pdf.set_fill_color(r,g,b); pdf.rect(cx,card_y,card_w,3,'F')
            # Label
            pdf.set_xy(cx+2, card_y+5)
            pdf.set_font("Arial","",7); pdf.set_text_color(74,96,128)
            pdf.cell(card_w-4,6,_asc(labels[i]).upper()[:22],align='C')
            # Valeur
            pdf.set_xy(cx+2, card_y+14)
            pdf.set_font("Arial","B",18)
            pdf.set_text_color(r,g,b)
            val = kpis[i]
            if isinstance(val,float) and abs(val)>=1000:
                val_str = _s(f"{val:,.0f}")
            elif isinstance(val,float) and abs(val)<=100:
                val_str = _s(f"{val:.1f}%")
            else:
                val_str = _s(str(int(val)) if isinstance(val,float) else str(val))
            pdf.cell(card_w-4,12,val_str,align='C')
        pdf.ln(46)

    # Scoring barres (sans titre redondant)
    scoring_lines=[]
    in_sc=False
    for line in content.split('\n'):
        ls=line.strip()
        if 'SCORING' in ls.upper():
            in_sc=True; continue
        if in_sc:
            if ls.startswith('###') or ls.startswith('---'): break
            if ls and ('/' in ls or ':' in ls): scoring_lines.append(ls)
    # Scoring calculé côté Python — plus fiable que la regex
    try:
        _score_pdf = compute_logiflo_score(
            module=module,
            df=None,
            kpis=kpis,
            labels=labels,
            sector_key="generique",
            lang=lang
        )
        _details_pdf = _score_pdf.get("details", {})
        if not _details_pdf and scoring_lines:
            # Fallback regex si compute n'a pas de données
            import re as _re2
            for sl in scoring_lines[:3]:
                sv=0
                nums=_re2.findall(r'(\d+)\s*/\s*100',sl)
                if nums: sv=int(nums[0])
                label_sc=_re2.sub(r'\s*[:\-]\s*\d+/100.*$','',sl).strip().lstrip('-').strip()
                _details_pdf[label_sc] = sv
    except Exception:
        _details_pdf = {}
        import re as _re2
        for sl in scoring_lines[:3]:
            sv=0
            nums=_re2.findall(r'(\d+)\s*/\s*100',sl)
            if nums: sv=int(nums[0])
            label_sc=_re2.sub(r'\s*[:\-]\s*\d+/100.*$','',sl).strip().lstrip('-').strip()
            _details_pdf[label_sc] = sv

    if _details_pdf:
        for _dim_label, _sv in list(_details_pdf.items())[:3]:
            _sv = int(_sv) if _sv else 0
            rc,gc,bc=(0,168,122) if _sv>=70 else (243,156,18) if _sv>=40 else (232,48,74)
            _lx=10; _bx=72; _sx=183; _row_y=pdf.get_y()
            pdf.set_font("Arial","",8); pdf.set_text_color(74,96,128)
            pdf.set_xy(_lx,_row_y); pdf.cell(60,7,_s(str(_dim_label))[:32],align='L')
            _bar_y=_row_y+1
            pdf.set_fill_color(225,232,240); pdf.rect(_bx,_bar_y,108,5,'F')
            pdf.set_fill_color(rc,gc,bc)
            _fill=int((_sv/100)*108) if _sv>0 else 0
            if _fill>0: pdf.rect(_bx,_bar_y,_fill,5,'F')
            pdf.set_font("Arial","B",8); pdf.set_text_color(rc,gc,bc)
            pdf.set_xy(_sx,_row_y); pdf.cell(22,7,f"{_sv}/100",align='R')
            pdf.ln(8)
    else:
        pdf.set_font("Arial","I",10); pdf.set_text_color(74,96,128)
        no_sc = "Generate AI analysis to see scoring." if lang=="en" else "Generez l'analyse IA pour voir le scoring."
        pdf.cell(0,8,_s(no_sc),ln=True)

    # ── PAGE 3 : GRAPHIQUES ───────────────────────────────────────
    if figs:
        pdf.add_page()
        pdf.set_fill_color(11,37,69); pdf.rect(0,0,210,18,'F')
        pdf.set_y(4); pdf.set_text_color(255,255,255); pdf.set_font("Arial","B",11)
        ch_label = "CHARTS & VISUALIZATIONS" if lang=="en" else "GRAPHIQUES & VISUALISATIONS"
        pdf.cell(0,10,_s(ch_label),ln=True,align='C'); pdf.ln(6)
        for fig in figs:
            _tp = None
            try:
                import uuid
                _tp = os.path.join(tempfile.gettempdir(), f"lgf_{uuid.uuid4().hex}.png")
                _ok = False
                # Export image — taille réduite pour tenir sur la page
                try:
                    _b = fig.to_image(format="png", width=740, height=320)
                    if _b and len(_b) > 200:
                        with open(_tp,"wb") as _f: _f.write(_b)
                        _ok = True
                except Exception:
                    pass
                if not _ok:
                    try:
                        fig.write_image(_tp, format="png", width=740, height=320)
                        if os.path.exists(_tp) and os.path.getsize(_tp) > 200:
                            _ok = True
                    except Exception:
                        pass
                if _ok and os.path.exists(_tp):
                    # Calculer l'espace dispo et adapter
                    _space = 297 - pdf.get_y() - 20  # espace vertical restant
                    _img_h = 84  # hauteur image en mm (320px @ ~72dpi)
                    _img_w = 176  # largeur image en mm (740px)
                    if _space < _img_h + 10:
                        # Plus assez de place verticalement → nouvelle page
                        pdf.add_page(); pdf.ln(5)
                    # Centrer horizontalement
                    _margin_x = (210 - _img_w) / 2
                    pdf.image(_tp, x=_margin_x, y=pdf.get_y(), w=_img_w)
                    pdf.ln(_img_h + 6)
            except Exception:
                pass
            finally:
                if _tp:
                    try: os.unlink(_tp)
                    except: pass

    # ── LISTING PIRES TRAJETS (transport uniquement, dans page 3) ──
    if module == "transport" and figs:
        # Extraire les pires trajets depuis le contenu IA
        _pires_lines = []
        _in_pires = False
        for _lp in content.split("\n"):
            _lp_s = _lp.strip()
            if any(k in _lp_s.upper() for k in ["PIRE","WORST","DEFICITAIRE","LOSS","TOXIC","RENTABILIT"]):
                if _lp_s.startswith("###"):
                    _in_pires = True
                    continue
            if _in_pires:
                if _lp_s.startswith("###"):
                    break
                if _lp_s.startswith(("-","*","•")) and _lp_s:
                    _pires_lines.append(_lp_s[1:].strip().replace("**",""))
                elif _lp_s and not _lp_s.startswith("#") and len(_pires_lines) < 8:
                    _pires_lines.append(_lp_s.replace("**",""))

        if _pires_lines:
            if pdf.get_y() > 240:
                pdf.add_page(); pdf.ln(5)
            pdf.set_font("Arial","B",10); pdf.set_text_color(11,37,69)
            _pl_lbl = "Top routes to address" if lang=="en" else "Trajets prioritaires à traiter"
            pdf.ln(4)
            pdf.set_fill_color(240,244,248); pdf.rect(10,pdf.get_y(),190,10,"F")
            pdf.set_fill_color(232,48,74); pdf.rect(10,pdf.get_y(),3,10,"F")
            pdf.set_x(16); pdf.cell(184,10,_s(_pl_lbl.upper()),ln=True); pdf.ln(3)
            for _pl in _pires_lines[:6]:
                if pdf.get_y() > 272: break
                pdf.set_font("Arial","",9); pdf.set_text_color(40,40,40)
                pdf.set_x(14); pdf.cell(4,5,"-"); pdf.set_x(18)
                pdf.multi_cell(182,5,_s(_pl[:120]))

    # ── PRÉDICTIONS RUPTURE (si données disponibles, module stock) ──
    if module == "stock":
        try:
            _preds_pdf = st.session_state.get("_last_predictions", [])
            if not _preds_pdf:
                # Recalculer depuis les KPIs si pas en session
                pass
            if _preds_pdf:
                if pdf.get_y() > 240:
                    pdf.add_page(); pdf.ln(5)
                pdf.set_font("Arial","B",10); pdf.set_text_color(11,37,69)
                _pred_lbl = "Stockout Predictions (next 4 weeks)" if lang=="en" else "Predictions de rupture (4 prochaines semaines)"
                pdf.ln(4)
                _ypr = pdf.get_y()
                pdf.set_fill_color(232,48,74); pdf.rect(10,_ypr,3,10,"F")
                pdf.set_fill_color(255,241,242); pdf.rect(13,_ypr,187,10,"F")
                pdf.set_x(17); pdf.cell(183,10,_s(_pred_lbl.upper()),ln=True); pdf.ln(3)
                for _pr in _preds_pdf[:6]:
                    if pdf.get_y() > 272: break
                    _urgence = str(_pr.get("urgence","")).upper()
                    _ref_pr  = str(_pr.get("reference",""))
                    _stk_pr  = int(_pr.get("stock",0))
                    _sem_pr  = float(_pr.get("semaines",0))
                    _conso_pr= float(_pr.get("conso_hebdo",0))
                    pdf.set_font("Arial","",9); pdf.set_text_color(40,40,40)
                    _txt_pr = f"[{_urgence}] {_ref_pr} — {_stk_pr} unites, conso {_conso_pr:.1f}/sem, rupture dans ~{_sem_pr:.1f} sem"
                    if lang=="en":
                        _txt_pr = f"[{_urgence}] {_ref_pr} — {_stk_pr} units, consumption {_conso_pr:.1f}/wk, stockout in ~{_sem_pr:.1f}wk"
                    pdf.set_x(14); pdf.cell(4,5,"⚠"); pdf.set_x(18)
                    pdf.multi_cell(182,5,_s(_txt_pr))
        except Exception:
            pass

    # ── PAGE 4 : ANALYSE IA ───────────────────────────────────────
    pdf.add_page()
    pdf.set_fill_color(11,37,69); pdf.rect(0,0,210,18,'F')
    pdf.set_y(4); pdf.set_text_color(255,255,255); pdf.set_font("Arial","B",11)
    ai_label = "AI ANALYSIS & RECOMMENDATIONS" if lang=="en" else "ANALYSE IA & RECOMMANDATIONS"
    pdf.cell(0,10,_s(ai_label),ln=True,align='C'); pdf.ln(8)

    content_r=(content.replace("\u2019","'").replace("\u2018","'")
                      .replace("\u201c",'"').replace("\u201d",'"')
                      .replace("\u20ac","EUR").replace("\u2022","-")
                      .replace("\u2013","-").replace("\u2014","-"))
    # Scoring toujours en p2 uniquement — on le saute en p4 pour tous les modules
    # Pour transport : on saute le scoring pour tenir sur 5 pages
    _force_skip_scoring = (module == "transport")
    skip_scoring=False
    _in_scoring=False
    for line in content_r.split('\n'):
        line=line.strip()
        if 'SCORING' in line.upper() and line.startswith('###'):
            _in_scoring=True; skip_scoring=True
        if _force_skip_scoring and _in_scoring:
            continue
        if skip_scoring and _in_scoring and not line.startswith('###'):
            continue
        if _in_scoring and line.startswith('###') and 'SCORING' not in line.upper():
            _in_scoring=False; skip_scoring=False
        if not line:
            pdf.ln(2); continue
        if line.startswith('### '):
            # Toujours laisser au moins 30mm après un titre (pour le contenu suivant)
            if pdf.get_y() > 250:
                pdf.add_page(); pdf.ln(4)
            t=_asc(line[4:])
            pdf.ln(5)
            _yh = pdf.get_y()
            pdf.set_fill_color(240,244,248); pdf.rect(10,_yh,190,10,'F')
            pdf.set_fill_color(0,200,150); pdf.rect(10,_yh,3,10,'F')
            pdf.set_font("Arial","B",10); pdf.set_text_color(11,37,69)
            pdf.set_x(16); pdf.cell(184,10,_s(t).upper(),ln=True); pdf.ln(4)
        elif line.startswith(('- ','* ')):
            _bt_text = _s(line[2:].replace("**",""))
            _bt_lines = max(1, len(_bt_text) // 42 + 1)
            if pdf.get_y() + _bt_lines * 6 + 2 > 272:
                pdf.add_page(); pdf.ln(4)
            pdf.set_font("Arial","",10); pdf.set_text_color(40,40,40)
            pdf.set_x(14); pdf.cell(5,6,"-"); pdf.set_x(19)
            pdf.multi_cell(181,6,_bt_text)
        else:
            _line_h = 6
            # Estimer la hauteur nécessaire (approx 45 chars par ligne)
            _est_lines = max(1, len(line) // 45 + 1)
            _needed = _est_lines * _line_h + 4
            if pdf.get_y() + _needed > 272:
                pdf.add_page(); pdf.ln(4)
            pdf.set_font("Arial","",10); pdf.set_text_color(40,40,40)
            cleaned=_s(line.replace("**",""))
            pdf.set_x(10); pdf.multi_cell(190,_line_h,cleaned)

    # ── PAGE 5 : CALL TO ACTION ───────────────────────────────────
    pdf.add_page()
    pdf.set_fill_color(11,37,69); pdf.rect(0,0,210,297,'F')
    pdf.set_fill_color(0,200,150); pdf.rect(0,0,210,6,'F'); pdf.rect(0,291,210,6,'F')
    pdf.set_y(85)
    pdf.set_text_color(0,200,150); pdf.set_font("Arial","B",32)
    pdf.cell(0,16,"LOGIFLO.IO",ln=True,align='C')
    pdf.ln(6)
    pdf.set_draw_color(0,200,150); pdf.set_line_width(0.6)
    pdf.line(50,pdf.get_y(),160,pdf.get_y()); pdf.ln(12)
    if lang=="en":
        cta_lines=[
            ("This report was generated by LOGIFLO.IO",True,255),
            ("","",200),
            ("Designed by a field logistics professional.",False,200),
            ("Not by a consultant.",False,200),
            ("","",200),
            ("Because real margin leaks don't show",False,170),
            ("up in dashboards.",False,170),
            ("","",200),
            ("To go further :",True,255),
            ("contact@logiflo.io",False,150),
            ("logiflo-io.streamlit.app",False,150),
        ]
    else:
        cta_lines=[
            ("Ce rapport a ete genere par LOGIFLO.IO",True,255),
            ("","",200),
            ("Concu par un logisticien terrain.",False,200),
            ("Pas par un consultant.",False,200),
            ("","",200),
            ("Parce que les vraies fuites de marge",False,170),
            ("ne se voient pas dans les tableaux de bord.",False,170),
            ("","",200),
            ("Pour aller plus loin :",True,255),
            ("contact@logiflo.io",False,150),
            ("logiflo-io.streamlit.app",False,150),
        ]
    for (txt,bold,br) in cta_lines:
        if not txt: pdf.ln(5); continue
        pdf.set_font("Arial","B" if bold else "",12 if bold else 11)
        pdf.set_text_color(br,br,br)
        pdf.cell(0,9,_s(txt),ln=True,align='C')

    # Encodage sécurisé : remplace tout caractère non latin-1 avant output
    raw = pdf.output(dest='S')
    if isinstance(raw, str):
        return raw.encode('latin-1', errors='replace')
    return raw


# =========================================
# 8. ROUTING ORS
# =========================================
def calculate_haversine(lon1,lat1,lon2,lat2):
    R=6371.0;dlat=math.radians(lat2-lat1);dlon=math.radians(lon2-lon1)
    a=math.sin(dlat/2)**2+math.cos(math.radians(lat1))*math.cos(math.radians(lat2))*math.sin(dlon/2)**2
    return R*2*math.atan2(math.sqrt(a),math.sqrt(1-a))

def fetch_geo(city,_t=None):
    if not city or str(city).strip() in ("","nan","None"): return city,None
    try:
        r=requests.get("https://nominatim.openstreetmap.org/search",
            params={"q":str(city).strip(),"format":"json","limit":1},
            headers={"User-Agent":"Logiflo.io/2.0"},timeout=5)
        if r.status_code==200:
            d=r.json()
            if d: return city,[float(d[0]["lon"]),float(d[0]["lat"])]
    except: pass
    return city,None

def geocode_cities_mapbox(cities):
    villes=[c for c in set(str(v) for v in cities)
            if c not in st.session_state.geo_cache and c not in ("","nan","None")]
    if villes:
        calc_txt = "Computing..." if st.session_state.get("language","fr")=="en" else "Calcul en cours..."
        bar=st.progress(0,text=calc_txt)
        for i,city in enumerate(villes):
            _discard,coord=fetch_geo(city)
            if coord: st.session_state.geo_cache[city]=coord
            time.sleep(1.1)
            bar.progress((i+1)/len(villes),text=calc_txt)
        bar.empty()
    return {c:st.session_state.geo_cache[c] for c in set(str(v) for v in cities) if c in st.session_state.geo_cache}

@st.cache_data(show_spinner=False)
def _ors_distance(lon1,lat1,lon2,lat2):
    for profile in ["driving-hgv","driving-car"]:
        try:
            r=requests.post(f"https://api.openrouteservice.org/v2/directions/{profile}",
                json={"coordinates":[[lon1,lat1],[lon2,lat2]],"instructions":False},
                headers={"Accept":"application/json","Content-Type":"application/json","Authorization":ORS_API_KEY},
                timeout=6)
            if r.status_code==200: return r.json()["routes"][0]["summary"]["distance"]/1000.0
        except: continue
    return None

def fetch_route(dep,arr,mode,coords,_t=None):
    c1,c2=coords.get(str(dep)),coords.get(str(arr))
    if not c1 or not c2: return (dep,arr,mode),0.0
    lon1,lat1=c1;lon2,lat2=c2
    dv=calculate_haversine(lon1,lat1,lon2,lat2);m=str(mode).lower()
    if any(k in m for k in ["mer","sea","maritime","bateau","port","ferry","conteneur"]): return (dep,arr,mode),dv*1.25
    elif any(k in m for k in ["air","avion","aerien","flight"]): return (dep,arr,mode),dv*1.05
    elif any(k in m for k in ["fer","rail","train","sncf"]): return (dep,arr,mode),dv*1.15
    else:
        d=_ors_distance(lon1,lat1,lon2,lat2)
        return (dep,arr,mode),(d if d and d>0 else dv*1.30)

def smart_multimodal_router(df,dep_col,arr_col,mode_col=None):
    coords=geocode_cities_mapbox(pd.concat([df[dep_col],df[arr_col]]).dropna().unique())
    uniq=[]
    for _r,row in df.iterrows():
        dep=row[dep_col];arr=row[arr_col]
        mode=str(row[mode_col]).lower() if mode_col and pd.notna(row.get(mode_col)) else "route"
        k=(dep,arr,mode)
        if k not in st.session_state.route_cache and k not in uniq: uniq.append(k)
    if uniq:
        with concurrent.futures.ThreadPoolExecutor(max_workers=8) as ex:
            for key,dist in [f.result() for f in concurrent.futures.as_completed(
                [ex.submit(fetch_route,r[0],r[1],r[2],coords) for r in uniq])]:
                st.session_state.route_cache[key]=dist
    df["_DIST_CALCULEE"]=[
        st.session_state.route_cache.get(
            (row[dep_col],row[arr_col],
             str(row[mode_col]).lower() if mode_col and pd.notna(row.get(mode_col)) else "route"),0.0)
        for _r,row in df.iterrows()]
    return df

def detect_transport_mode(df,dep_col=None,arr_col=None,mode_col=None):
    # Ports maritimes — noms complets (matching par inclusion sur tokens)
    PORTS=["havre","marseille","dunkerque","bordeaux","hamburg","rotterdam","antwerp",
           "anvers","amsterdam","barcelona","barcelone","genova","genes","piraeus","piree",
           "istanbul","dakar","casablanca","lagos","mombasa","durban","santos","shanghai",
           "ningbo","guangzhou","shenzhen","hongkong","singapore","singapour","busan",
           "tokyo","yokohama","losangeles","newyork","savannah","miami","sydney","dubai",
           "jeddah","mumbai","nhavasheva","colombo","tanger","tangermed","algier","alger",
           "tunis","tripoli","beyrouth","callao","buenosaires"]

    # Codes IATA aéroports — MATCHING EXACT PAR TOKEN uniquement
    # On split la valeur en tokens et on compare chaque token exactement
    AIRPORT_CODES={"cdg","ory","lyo","mrs","nce","bor","tls","sxb","bod","mlh",
                   "jfk","lax","ord","mia","atl","sfo","sea","bos","iah","dfw","ewr",
                   "lhr","lgw","man","fra","muc","txl","ber","ham","vie","zrh","gva",
                   "ams","bru","cph","arn","hel","mad","bcn","fco","mxp","lin","fcm",
                   "dxb","auh","doh","ist","tlv","bom","del","hkg","nrt","icn","kix",
                   "sin","kul","bkk","syd","mel","per","akl","gru","gig","bog","lim",
                   "mex","yyz","yvr","jnb","nbo","cai","cmn","dkr","los","acc","abv"}

    # Villes clairement routières — boost score road
    ROAD_CITIES={"paris","lyon","toulouse","bordeaux","lille","marseille","nantes",
                 "strasbourg","rennes","nice","grenoble","montpellier","tours","dijon",
                 "metz","nancy","reims","rouen","amiens","clermont","limoges","poitiers",
                 "bruxelles","brussels","amsterdam","berlin","munich","francfort","cologne",
                 "madrid","barcelona","rome","milan","geneve","zurich","vienne","varsovie",
                 "bucarest","budapest","prague","bratislava","ljubljana","zagreb",
                 "london","rotterdam","antwerp","hamburg","dusseldorf","Stuttgart",
                 "birmingham","manchester","edinburgh","glasgow","bristol"}

    KW_AIR  = ["aerien","aérien","air freight","airfreight","awb","air waybill",
               "fret aerien","airline cargo","avion","aerian"]
    KW_SEA  = ["maritime","seafreight","sea freight","ocean freight","bateau","navire",
               "conteneur","container","teu","fcl","lcl","armateur","roro","ro-ro",
               "reefer","vrac","bulk","mer","ocean"]
    KW_RAIL = ["ferroviaire","rail","train","sncf","wagon","fret ferroviaire","railway"]
    KW_ROAD = ["routier","road","camion","truck","ftl","ltl","vl","tir","messagerie",
               "groupage","express","fret routier","road freight","haulage","trucking"]

    scores = {"aerien":0,"maritime":0,"ferroviaire":0,"routier":0}

    # ── 1. Colonne mode explicite (poids fort) ────────────────────
    if mode_col and mode_col in df.columns:
        for v in df[mode_col].dropna().astype(str).str.lower():
            for kw in KW_AIR:
                if kw in v: scores["aerien"] += 3
            for kw in KW_SEA:
                if kw in v: scores["maritime"] += 3
            for kw in KW_RAIL:
                if kw in v: scores["ferroviaire"] += 3
            for kw in KW_ROAD:
                if kw in v: scores["routier"] += 3

    # ── 2. Colonnes dep/arr — MATCHING EXACT PAR TOKEN ───────────
    for col in [dep_col, arr_col]:
        if not col or col not in df.columns:
            continue
        for v in df[col].dropna().astype(str):
            # Split en tokens individuels (espace, tiret, slash)
            raw_tokens = re.split(r'[\s\-/,]+', v.strip())
            tokens_clean = [nettoyer(t) for t in raw_tokens if t.strip()]

            for tok in tokens_clean:
                # Aéroport : code IATA exact (3 lettres)
                if tok in AIRPORT_CODES:
                    scores["aerien"] += 2
                # Port : nom de port inclus dans le token ou vice versa
                if any(p in tok or tok in p for p in PORTS if len(p) >= 5):
                    scores["maritime"] += 1
                # Ville routière identifiée
                if tok in ROAD_CITIES or any(tok in rc for rc in ROAD_CITIES if len(rc) >= 5):
                    scores["routier"] += 1

    # ── 3. Analyse des headers du fichier ────────────────────────
    hdrs = [nettoyer(c) for c in df.columns]

    # Headers forte indication aérien
    if any("awb" in h for h in hdrs):              scores["aerien"] += 6
    if any("airwaybill" in h for h in hdrs):        scores["aerien"] += 6
    if any("chargeableweight" in h for h in hdrs):  scores["aerien"] += 5
    if any("flightdate" in h or "flightno" in h for h in hdrs): scores["aerien"] += 4

    # Headers forte indication maritime
    if any("bl" == h or "billoflading" in h for h in hdrs): scores["maritime"] += 6
    if any("teu" in h for h in hdrs):              scores["maritime"] += 5
    if any("conteneur" in h or "container" in h for h in hdrs): scores["maritime"] += 5
    if any("etd" in h or "eta" in h for h in hdrs): scores["maritime"] += 3
    if any("armateur" in h or "carrier" in h for h in hdrs): scores["maritime"] += 2

    # Headers forte indication routière
    if any("distancekm" in h or "km" in h for h in hdrs): scores["routier"] += 4
    if any("plaque" in h or "immatricul" in h for h in hdrs): scores["routier"] += 3
    if any("orderid" in h or "ordernum" in h for h in hdrs): scores["routier"] += 2

    # Headers forte indication ferroviaire
    if any("wagon" in h or "sncf" in h for h in hdrs): scores["ferroviaire"] += 6

    # ── 4. Analyse du contenu global pour confirmation ────────────
    # Si on a une colonne distance_km avec des valeurs typiques route
    for col in df.columns:
        if "km" in nettoyer(col) or "dist" in nettoyer(col):
            try:
                vals = pd.to_numeric(df[col], errors='coerce').dropna()
                if len(vals) > 3:
                    med = vals.median()
                    # Distance routière typique : 50-3000 km
                    if 50 <= med <= 3000:
                        scores["routier"] += 3
                    # Distance aérienne typique : > 500 km sans doute
            except: pass

    dominant = max(scores, key=scores.get)
    total = sum(scores.values())

    # Pas assez de signal → routier par défaut
    if total == 0 or scores[dominant] < 2:
        return "routier","🚛 Road (default)","🚛"

    # Cas d'égalité → routier gagne (mode le plus courant)
    top_val = scores[dominant]
    rivals = [k for k,v in scores.items() if v == top_val and k != dominant]
    if rivals:
        dominant = "routier"

    lang = st.session_state.get("language","fr")
    if lang == "fr":
        labels = {
            "aerien":     ("✈️ Mode Aérien détecté",      "✈️"),
            "maritime":   ("⚓ Mode Maritime détecté",    "⚓"),
            "ferroviaire":("🚂 Mode Ferroviaire détecté", "🚂"),
            "routier":    ("🚛 Mode Routier détecté",     "🚛"),
        }
    else:
        labels = {
            "aerien":     ("✈️ Air mode detected",        "✈️"),
            "maritime":   ("⚓ Maritime mode detected",   "⚓"),
            "ferroviaire":("🚂 Rail mode detected",       "🚂"),
            "routier":    ("🚛 Road mode detected",       "🚛"),
        }
    label, emoji = labels[dominant]
    return dominant, label, emoji

def super_clean(val):
    if pd.isna(val): return 0.0
    try: return float(str(val).replace('€','').replace('$','').replace('EUR','').replace(' ','').replace('\xa0','').replace(',','.'))
    except: return 0.0

# =========================================
# 9. PAGES
# =========================================
if st.session_state.page=="accueil":
    st.markdown(f"<h1 style='text-align:center;color:#0B2545;font-family:Syne,sans-serif;font-weight:800;letter-spacing:-1px;'>{_('home_title')}</h1>",unsafe_allow_html=True)
    st.markdown(f"<p style='text-align:center;font-size:1.1em;color:#4A6080;'>{_('home_sub')}</p><br>",unsafe_allow_html=True)
    # Sélecteur langue sur la page d'accueil
    _c1,lc,_c2=st.columns([3,1,3])
    with lc:
        lang_choice=st.selectbox("",["🇫🇷 Français","🇬🇧 English"],key="lang_accueil",label_visibility="collapsed")
        st.session_state.language="en" if "English" in lang_choice else "fr"
    st.markdown("<br>",unsafe_allow_html=True)
    c1,c2=st.columns(2)
    with c1:
        st.markdown("<span class='big-emoji'>📦</span>",unsafe_allow_html=True)
        if st.button(_("home_stock"),use_container_width=True):
            st.session_state.module="stock";st.session_state.page="choix_profil_stock";st.rerun()


    with c2:
        st.markdown("<span class='big-emoji'>🌍</span>",unsafe_allow_html=True)
        if st.button(_("home_transport"),use_container_width=True):
            st.session_state.module="transport";st.session_state.page="login";st.rerun()
    st.markdown("<br><br>",unsafe_allow_html=True)
    _c1,cm,_c2=st.columns([1,1,1])
    if cm.button(_("home_access"),use_container_width=True):
        st.session_state.page="contact";st.rerun()
    st.markdown("<br>",unsafe_allow_html=True)
    _ca1,_cf,_ca2=st.columns([1,2,1])
    _free_label="→ Launch my free audit" if st.session_state.get("language","fr")=="en" else "→ Lancer mon audit gratuit"
    if _cf.button(_free_label,use_container_width=True,key="btn_free_home"):
        st.session_state.page="audit_gratuit";st.rerun()

elif st.session_state.page=="contact":
    st.markdown(f"<h2 style='text-align:center;color:#0B2545;font-family:Syne,sans-serif;'>{_('contact_title')}</h2>",unsafe_allow_html=True)
    _c1,cc,_c2=st.columns([1,1.5,1])
    with cc:
        with st.form("vip"):
            st.text_input(_("contact_name"));st.text_input(_("contact_email"));st.text_input(_("contact_company"))
            st.selectbox(_("contact_volume"),[_("vol1"),_("vol2"),_("vol3")])
            st.selectbox(_("contact_issue"),[_("iss1"),_("iss2"),_("iss3")])
            if st.form_submit_button(_("contact_btn"),use_container_width=True):
                st.success(_("contact_ok"))
        if st.button(_("login_back"),use_container_width=True): st.session_state.page="accueil";st.rerun()

elif st.session_state.page=="audit_gratuit":
    lang_ag=st.session_state.get("language","fr")
    st.markdown("<h1 style='text-align:center;color:#0B2545;font-family:Syne,sans-serif;font-weight:800;'>"+("Free Audit" if lang_ag=="en" else "Audit Gratuit")+"</h1>",unsafe_allow_html=True)
    st.markdown("<p style='text-align:center;color:#4A6080;margin-bottom:24px;'>"+("Upload your file — get your diagnosis in 2 minutes. No login required." if lang_ag=="en" else "Uploadez votre fichier — obtenez votre diagnostic en 2 minutes. Sans inscription.")+"</p>",unsafe_allow_html=True)

    if st.session_state.get("audit_gratuit_done"):
        st.warning("You have already used your free audit." if lang_ag=="en" else "Vous avez deja utilise votre audit gratuit.")
        st.info("Create an account for full audits with history and charts." if lang_ag=="en" else "Creez un compte pour les audits complets avec historique et graphiques.")
        if st.button("Back" if lang_ag=="en" else "Retour",use_container_width=True,key="back_done2"):
            st.session_state.page="accueil"; st.rerun()
    else:
        _fmc=st.radio("",["Stock","Transport"],horizontal=True,label_visibility="collapsed",key="fmc")
        _fmod="stock" if "Stock" in _fmc else "transport"
        _upf=st.file_uploader("Fichier Excel ou CSV" if lang_ag=="fr" else "Excel or CSV file",type=["csv","xlsx"],key="free_up2")
        if _upf:
            with st.spinner("Calcul en cours..." if lang_ag=="fr" else "Computing..."):
                try:
                    _dff=pd.read_excel(_upf) if _upf.name.endswith("xlsx") else pd.read_csv(_upf,encoding="utf-8")
                except Exception:
                    _upf.seek(0); _dff=pd.read_csv(_upf,encoding="latin-1")
            if _fmod=="stock":
                _dfok,_st2=smart_ingester_stock_ultime(_dff,client_ai=client)
                if _dfok is None: st.error(_st2)
                else:
                    _sp=bool(_dfok.get("_sans_prix",pd.Series([True])).iloc[0]) if "_sans_prix" in _dfok.columns else True
                    _dfok["valeur_totale"]=_dfok["quantite"]*_dfok["prix_unitaire"]
                    _vt=_dfok["valeur_totale"].sum()
                    _rf=_dfok[_dfok["quantite"]<=0]
                    _txf=(1-len(_rf)/max(len(_dfok),1))*100
                    _fkpis=[_vt if not _sp else float(len(_dfok)),_txf,float(len(_rf))]
                    _flbl=["Capital EUR" if not _sp else "Articles","Service %","Ruptures"]
                    with st.spinner("Analyse IA..." if lang_ag=="fr" else "AI Analysis..."):
                        _fsum=generate_ai_analysis(f"Items:{len(_dfok)}. Service:{_txf:.1f}%. Stockouts:{len(_rf)}. Prices:{'No' if _sp else 'Yes'}. SIMPLIFIED - 3 key points max.")
                    a1,a2,a3=st.columns(3)
                    a1.metric("Capital EUR" if not _sp else "Articles",f"{_fkpis[0]:,.0f}")
                    a2.metric("Service",f"{_txf:.1f}%")
                    a3.metric("Ruptures",str(len(_rf)))
                    st.markdown(render_report(_fsum,"manager"),unsafe_allow_html=True)
                    _fpdf=generate_free_pdf("stock",_fsum,_fkpis,_flbl)
                    st.download_button("Telecharger mon rapport (PDF)" if lang_ag=="fr" else "Download my report (PDF)",_fpdf,"Audit_Gratuit_Logiflo.pdf",use_container_width=True)
                    st.session_state.audit_gratuit_done=True
                    st.info("Audit gratuit utilise. Creez un compte pour l'analyse complete." if lang_ag=="fr" else "Free audit used. Create an account for full analysis.")
            else:
                _mapf=auto_map_columns_with_ai(_dff)
                def _colf(k): return _mapf.get(k) if _mapf.get(k) in _dff.columns else None
                _caf=_colf("ca"); _cof=_colf("co")
                if not _cof:
                    for _cc in _dff.columns:
                        if any(k in str(_cc).lower() for k in ["cout","cost","achat"]): _cof=_cc; break
                if not _caf:
                    for _cc in _dff.columns:
                        if any(k in str(_cc).lower() for k in ["ca","revenue","facture"]): _caf=_cc; break
                if not _cof: st.error("Colonne cout introuvable." if lang_ag=="fr" else "Cost column not found.")
                else:
                    _dff["_CO"]=_dff[_cof].apply(super_clean)
                    _dff["_CA"]=_dff[_caf].apply(super_clean) if _caf else _dff["_CO"]/0.85
                    _dff["_MG"]=_dff["_CA"]-_dff["_CO"]
                    _mgt=_dff["_MG"].sum(); _cat=_dff["_CA"].sum()
                    _txt=(_mgt/_cat*100) if _cat>0 else 0
                    _toxt=len(_dff[_dff["_MG"]<0])
                    _fkpis=[_mgt,_txt,float(_toxt)]
                    _flbl=["Marge EUR","Taux %","Deficitaires"]
                    with st.spinner("Analyse IA..." if lang_ag=="fr" else "AI Analysis..."):
                        _fsum=generate_ai_analysis(f"Routes:{len(_dff)}. Margin:{_mgt:.0f} EUR. Rate:{_txt:.1f}%. Loss:{_toxt}. SIMPLIFIED - 3 key points max.")
                    a1,a2,a3=st.columns(3)
                    a1.metric("Marge",f"{_mgt:,.0f} EUR")
                    a2.metric("Taux",f"{_txt:.1f}%")
                    a3.metric("Deficitaires",str(_toxt))
                    st.markdown(render_report(_fsum,"manager"),unsafe_allow_html=True)
                    _fpdf=generate_free_pdf("transport",_fsum,_fkpis,_flbl)
                    st.download_button("Telecharger mon rapport (PDF)" if lang_ag=="fr" else "Download my report (PDF)",_fpdf,"Audit_Gratuit_Transport.pdf",use_container_width=True)
                    st.session_state.audit_gratuit_done=True
                    st.info("Audit gratuit utilise. Creez un compte pour l'analyse complete." if lang_ag=="fr" else "Free audit used. Create an account for full analysis.")
        st.markdown("<br>",unsafe_allow_html=True)
        if st.button("Retour" if lang_ag=="fr" else "Back",use_container_width=True,key="back_free2"):
            st.session_state.page="accueil"; st.rerun()

elif st.session_state.page=="choix_profil_stock":
    st.markdown(f"<h2 style='text-align:center;color:#0B2545;font-family:Syne,sans-serif;'>{_('profile_title')}</h2>",unsafe_allow_html=True)
    st.markdown(f"<p style='text-align:center;color:#4A6080;'>{_('profile_sub')}</p><br><br>",unsafe_allow_html=True)
    c1,c2=st.columns(2)
    with c1:
        st.markdown("<span class='big-emoji'>📊</span>",unsafe_allow_html=True)
        if st.button(_("profile_mgr"),use_container_width=True):
            st.session_state.stock_view="MANAGER";st.session_state.page="login";st.rerun()
    with c2:
        st.markdown("<span class='big-emoji'>👷</span>",unsafe_allow_html=True)
        if st.button(_("profile_ops"),use_container_width=True):
            st.session_state.stock_view="TERRAIN";st.session_state.page="login";st.rerun()

elif st.session_state.page=="login":
    st.markdown(f"<h2 style='text-align:center;color:#0B2545;font-family:Syne,sans-serif;'>{'Secure Access' if st.session_state.get('language')=='en' else 'Accès Sécurisé'} — {st.session_state.module.upper()}</h2><br>",unsafe_allow_html=True)
    _c1,cl,_c2=st.columns([1,1.2,1])
    with cl:
        with st.form("login_form"):
            u=st.text_input(_("login_id"));p=st.text_input(_("login_pw"),type="password")
            st.markdown("<br>",unsafe_allow_html=True)
            if st.form_submit_button(_("login_btn"),use_container_width=True):
                if u in USERS_DB and USERS_DB[u]==p:
                    st.session_state.auth=True;st.session_state.current_user=u
                    # Charger prefs Supabase
                    try:
                        _prefs_l = load_user_prefs(u)
                        if _prefs_l:
                            if _prefs_l.get("company_name"):
                                st.session_state["company_name"] = _prefs_l["company_name"]
                            if _prefs_l.get("language"):
                                st.session_state["language"] = _prefs_l["language"]
                            if _prefs_l.get("seuil_rupture") is not None:
                                st.session_state["seuil_rupture"] = int(_prefs_l["seuil_rupture"])
                            if _prefs_l.get("preferred_module"):
                                st.session_state["preferred_module"] = _prefs_l["preferred_module"]
                    except Exception:
                        pass
                    st.session_state.page="app";st.rerun()
                else: st.error(_("login_err"))
        if st.button(_("login_back"),use_container_width=True): st.session_state.page="accueil";st.rerun()

elif st.session_state.auth and st.session_state.page=="app":
    with st.sidebar:
        _plan_sb  = get_user_plan(st.session_state.current_user)
        _pinfo_sb = PLAN_LIMITS.get(_plan_sb, PLAN_LIMITS["starter"])
        _can_audit_sb = audit_counter_sidebar(st.session_state.current_user, _plan_sb)
        _plan_bg  = _pinfo_sb["bg"]
        _plan_col = _pinfo_sb["color"]
        _plan_ico = _pinfo_sb["icon"]
        _plan_lbl = _pinfo_sb["label"]
        _cur_user = st.session_state.current_user
        st.markdown(f"""
            <div style="display:flex;align-items:center;gap:10px;margin-bottom:12px;">
                <div class="sidebar-logo">LOGI<span>FLO</span>.IO</div>
            </div>
            <div style="font-size:12px;color:#4A6080;margin-bottom:6px;">
                👤 {_cur_user}
            </div>
            <div style="display:inline-flex;align-items:center;gap:4px;padding:2px 8px;
                        border-radius:20px;margin-bottom:10px;
                        background:{_plan_bg};color:{_plan_col};
                        border:1px solid {_plan_col}40;
                        font-size:10px;font-weight:700;">
                {_plan_ico} {_plan_lbl}
            </div>
        """,unsafe_allow_html=True)
        st.markdown("---")
        _is_terrain = (st.session_state.get("stock_view","") == "TERRAIN")
        _nav_items = ([_("nav_workspace"),_("nav_archives"),_("nav_params"),_("nav_legal")]
                      if _is_terrain else
                      [_("nav_dashboard"),_("nav_workspace"),_("nav_archives"),_("nav_compte"),_("nav_params"),_("nav_legal")])
        nav=st.radio("",_nav_items,
                     label_visibility="collapsed")
        st.markdown("---")
        if st.button(_("nav_logout"),use_container_width=True): st.session_state.clear();st.rerun()
        st.markdown("<div style='margin-top:40px;border-top:1px solid #1e3a5f;padding-top:14px;font-size:11px;color:#4A6080;'>© 2026 Logiflo B2B Enterprise</div>",unsafe_allow_html=True)

    # ── LEGAL ──
    # ── DASHBOARD ──────────────────────────────────────────────────
    if nav==_("nav_dashboard"):
        lang_d     = st.session_state.get("language","fr")
        username_d = st.session_state.current_user
        _df_arch   = load_archives_from_sheets(username_d)

        # Normaliser colonnes
        if _df_arch is not None and not _df_arch.empty:
            if "created_at" in _df_arch.columns and "date" not in _df_arch.columns:
                _df_arch["date"]  = pd.to_datetime(_df_arch["created_at"],errors="coerce").dt.strftime("%d/%m/%Y")
                _df_arch["heure"] = pd.to_datetime(_df_arch["created_at"],errors="coerce").dt.strftime("%H:%M")
            for _c in ["module","date","heure","kpi_1","kpi_2","kpi_3",
                       "kpi_label_1","kpi_label_2","kpi_label_3","resume_ia"]:
                if _c not in _df_arch.columns:
                    _df_arch[_c] = ""
            for _c in ["kpi_1","kpi_2","kpi_3"]:
                _df_arch[_c] = pd.to_numeric(_df_arch[_c],errors="coerce").fillna(0)

        # ── Header salutation ─────────────────────────────────────
        import datetime as _dt_dash
        _hour_d = _dt_dash.datetime.now().hour
        _greet  = ("Good morning" if _hour_d<12 else "Good afternoon" if _hour_d<18 else "Good evening") if lang_d=="en" else ("Bonjour" if _hour_d<18 else "Bonsoir")
        st.markdown(f"""
        <div style="background:linear-gradient(135deg,#0B2545 0%,#0f2f5a 100%);
                    border-radius:14px;padding:22px 28px;margin-bottom:20px;
                    border-left:4px solid #00C896;">
            <div style="font-family:Syne,sans-serif;font-size:20px;font-weight:800;
                        color:white;margin-bottom:3px;">
                {_greet}, <span style="color:#00C896;">{username_d}</span>
            </div>
            <div style="font-size:12px;color:rgba(255,255,255,0.5);">
                {"Your supply chain at a glance" if lang_d=="en" else "Votre supply chain en un coup d\'oeil"}
            </div>
        </div>
        """, unsafe_allow_html=True)

        if _df_arch is None or _df_arch.empty:
            st.info("Aucun audit encore. Lancez votre premier audit." if lang_d=="fr"
                    else "No audit yet. Launch your first audit.")
        else:
            # ── Alertes régression ────────────────────────────────
            for _mod_al in _df_arch["module"].unique():
                _df_al = _df_arch[_df_arch["module"]==_mod_al].tail(2)
                if len(_df_al) >= 2:
                    try:
                        _dv = float(_df_al.iloc[1]["kpi_2"]) - float(_df_al.iloc[0]["kpi_2"])
                        _lbl2_al = str(_df_al.iloc[1].get("kpi_label_2",""))
                        if _dv < -3:
                            _ico_al = "📦" if _mod_al=="stock" else "🚚"
                            st.warning(f"⚠️ {_ico_al} {_lbl2_al} : baisse de **{abs(_dv):.1f} pts** depuis le dernier audit ({str(_df_al.iloc[1].get('date',''))})")
                    except Exception:
                        pass

            # ══ VISUALISATIONS ════════════════════════════════════
            import plotly.graph_objects as _go_d2
            _mod_actif = st.session_state.get("module","stock")
            # Transport → courbe pleine largeur | Stock → camembert + courbe
            if _mod_actif == "transport":
                _col_l, _col_r = None, st  # pleine largeur pour transport
            else:
                _col_l, _col_r = st.columns(2)

            # ── Camembert DERNIER audit sauvegardé (stock uniquement) ──
            _col_l_eff = _col_l if _mod_actif=="stock" and _col_l is not None else None
            if _col_l_eff is not None:
              with _col_l_eff:
                try:
                    _dfs_pie = _df_arch[_df_arch["module"]=="stock"].copy()
                    if not _dfs_pie.empty:
                        # Trier par date pour prendre LE DERNIER audit
                        _sort_c = "created_at" if "created_at" in _dfs_pie.columns else "date"
                        _dfs_pie = _dfs_pie.sort_values(_sort_c, ascending=True)
                        _lp = _dfs_pie.iloc[-1]
                        _k1p  = float(_lp.get("kpi_1",0))
                        _k2p  = float(_lp.get("kpi_2",0))
                        _k3p  = float(_lp.get("kpi_3",0))
                        _l1p  = str(_lp.get("kpi_label_1","Capital"))
                        _l2p  = str(_lp.get("kpi_label_2","Service"))
                        _l3p  = str(_lp.get("kpi_label_3","Ruptures"))
                        _date_pie = str(_lp.get("date","") or str(_lp.get("created_at",""))[:10])
                        # Valeurs normalisées pour le camembert
                        _ok   = max(0.1, 100 - _k2p - min(_k3p*5, 30))
                        _fig_p = _go_d2.Figure(_go_d2.Pie(
                            labels=[_l2p, _l3p, "Sain"],
                            values=[_k2p, max(_k3p, 0.1), max(_ok, 0.1)],
                            hole=0.45,
                            marker=dict(
                                colors=["#00C896","#E8304A","#E2E8F0"],
                                line=dict(color="white",width=2)
                            ),
                            textfont=dict(size=10),
                            hovertemplate="%{label}: <b>%{value:.1f}%</b><extra></extra>",
                        ))
                        _fig_p.update_layout(
                            title=dict(
                                text=f"📦 Stock — {_date_pie}",
                                font=dict(size=12,color="#0B2545"),x=0
                            ),
                            legend=dict(font=dict(size=9),orientation="h",
                                        yanchor="bottom",y=-0.25),
                            margin=dict(t=36,b=50,l=0,r=0),
                            height=230,paper_bgcolor="white",
                        )
                        st.plotly_chart(_fig_p,use_container_width=True,
                                        config={"displayModeBar":False})
                    else:
                        st.markdown(
                            "<div style='height:230px;display:flex;align-items:center;"
                            "justify-content:center;background:white;border-radius:12px;"
                            "border:1px solid #E2E8F0;color:#8FA3BC;font-size:12px;'>"
                            "Lancez un audit Stock pour voir la répartition.</div>",
                            unsafe_allow_html=True
                        )
                except Exception:
                    pass

            # ── Courbe TOUS audits — hover = résumé financier bref ────
            _col_r_eff = _col_r if _col_r is not None else st
            with _col_r_eff:
                try:
                    _plotted_d = False
                    for _mc in ["stock","transport"]:
                        _dfc = _df_arch[_df_arch["module"]==_mc].copy()
                        if len(_dfc) < 2:
                            continue
                        _sort_c2 = "created_at" if "created_at" in _dfc.columns else "date"
                        _dfc = _dfc.sort_values(_sort_c2, ascending=True).reset_index(drop=True)
                        _dfc["kpi_1"] = pd.to_numeric(_dfc["kpi_1"],errors="coerce").fillna(0)
                        _dfc["kpi_2"] = pd.to_numeric(_dfc["kpi_2"],errors="coerce").fillna(0)
                        _dfc["kpi_3"] = pd.to_numeric(_dfc["kpi_3"],errors="coerce").fillna(0)
                        _l2c = str(_dfc["kpi_label_2"].iloc[-1])
                        _l1c = str(_dfc["kpi_label_1"].iloc[-1])
                        _l3c = str(_dfc["kpi_label_3"].iloc[-1])
                        _clr = "#00C896" if _mc=="stock" else "#F39C12"
                        _dates_c = [str(d)[:10] for d in _dfc["date"].tolist()] if "date" in _dfc.columns else [str(i) for i in range(len(_dfc))]
                        # Construire le texte hover : résumé financier bref par audit
                        _hovers = []
                        for _idx_c, _rw_c in _dfc.iterrows():
                            _res_c  = str(_rw_c.get("resume_ia","")).strip()
                            # Extraire 1ère phrase significative du résumé
                            _brief = ""
                            for _sent in _res_c.replace("\n"," ").split("."):
                                _s = _sent.strip()
                                if len(_s) > 20 and not _s.startswith("#") and not _s.startswith("-"):
                                    _brief = _s[:80] + "…"
                                    break
                            if not _brief:
                                _brief = f"{_l2c}: {float(_rw_c.get('kpi_2',0)):.1f}%"
                            _hover_txt = (
                                f"<b>{_dates_c[_idx_c]}</b><br>"
                                f"{_l1c}: {float(_rw_c.get('kpi_1',0)):.0f}<br>"
                                f"{_l2c}: <b>{float(_rw_c.get('kpi_2',0)):.1f}%</b><br>"
                                f"{_l3c}: {float(_rw_c.get('kpi_3',0)):.0f}<br>"
                                f"<i>{_brief}</i>"
                            )
                            _hovers.append(_hover_txt)
                        _fig_c = _go_d2.Figure()
                        _fig_c.add_trace(_go_d2.Scatter(
                            x=list(range(len(_dfc))),
                            y=_dfc["kpi_2"].tolist(),
                            mode="lines+markers",
                            line=dict(color=_clr,width=2.5),
                            marker=dict(size=9,color=_clr,
                                        line=dict(color="white",width=2)),
                            fill="tozeroy",
                            fillcolor="rgba(0,200,150,0.08)" if _mc=="stock" else "rgba(243,156,18,0.08)",
                            hovertemplate="%{customdata}<extra></extra>",
                            customdata=_hovers,
                        ))
                        _fig_c.update_layout(
                            title=dict(
                                text=f"📈 {_l2c} — {len(_dfc)} audits",
                                font=dict(size=12,color="#0B2545"),x=0
                            ),
                            xaxis=dict(
                                tickmode="array",
                                tickvals=list(range(len(_dfc))),
                                ticktext=[d[:5] for d in _dates_c],
                                tickfont=dict(size=8),showgrid=False
                            ),
                            yaxis=dict(
                                tickfont=dict(size=8),
                                gridcolor="rgba(0,0,0,0.04)"
                            ),
                            hoverlabel=dict(
                                bgcolor="white",
                                bordercolor=_clr,
                                font=dict(size=11,color="#0B2545")
                            ),
                            plot_bgcolor="white",paper_bgcolor="white",
                            margin=dict(t=36,b=20,l=30,r=10),
                            height=230,showlegend=False,
                        )
                        st.plotly_chart(_fig_c,use_container_width=True,
                                        config={"displayModeBar":False})
                        _plotted_d = True
                        break
                    if not _plotted_d:
                        st.markdown(
                            "<div style='height:230px;display:flex;align-items:center;"
                            "justify-content:center;background:white;border-radius:12px;"
                            "border:1px solid #E2E8F0;color:#8FA3BC;font-size:12px;'>"
                            "2 audits minimum pour afficher la courbe.</div>",
                            unsafe_allow_html=True
                        )
                except Exception:
                    pass

            # ══ TASKS — extraites du dernier audit ════════════════
            st.markdown("<br>",unsafe_allow_html=True)
            try:
                _sort_col = "created_at" if "created_at" in _df_arch.columns else "date"
                _lr = _df_arch.sort_values(_sort_col,ascending=False).iloc[0]
                _resume_lr = str(_lr.get("resume_ia","")).strip()
                _date_lr   = str(_lr.get("date","") or str(_lr.get("created_at",""))[:10])
                _mod_lr    = str(_lr.get("module",""))
                _k2_lr     = float(_lr.get("kpi_2",0))
                _l2_lr     = str(_lr.get("kpi_label_2",""))
                _ico_lr    = "📦" if _mod_lr=="stock" else "🚚"
                _clr_lr    = "#00C896" if _k2_lr>=90 else ("#F39C12" if _k2_lr>=75 else "#E8304A")

                import re as _re_t
                _tasks = []
                _in_action = False
                for _ln in _resume_lr.split("\n"):
                    _ln = _ln.strip()
                    if not _ln or _ln.lower() in ("none","nan","null","n/a","-","---"):
                        continue
                    # Détecter section actions
                    if any(k in _ln.upper() for k in ["A FAIRE","PRIORIT","ACTION","RECOMMAND","NEXT STEP","TO DO"]):
                        if _ln.startswith("#"):
                            _in_action = True
                            continue
                    if _in_action and _ln.startswith("#"):
                        _in_action = False
                    _ln_c = _ln.replace("**","").replace("*","").strip()
                    # Ignorer les lignes purement numériques ou trop courtes
                    if len(_ln_c) < 15:
                        continue
                    if _re_t.match(r"^[-•*→✅]\s+.{15,}", _ln_c):
                        _t = _ln_c.lstrip("-•*→✅ ").strip()
                        if _t and not _re_t.match(r"^\d+[\.,]\d+", _t):  # éviter "94,2% de service"
                            _tasks.append(_t)
                    elif _in_action and _re_t.match(r"^\d+\.\s+.{15,}", _ln_c):
                        _t = _re_t.sub(r"^\d+\.\s+","",_ln_c).strip()
                        if _t:
                            _tasks.append(_t)
                # Nettoyer : enlever les tâches qui ne sont que des chiffres ou stats
                _tasks = [
                    t[:130] for t in _tasks
                    if t and len(t)>15
                    and not _re_t.match(r"^[\d\.,%€\s]+$", t)
                ][:6]

                _lbl_tasks = f"{_ico_lr} À faire — {_date_lr} · {_l2_lr} : {_k2_lr:.1f}%"
                if lang_d=="en":
                    _lbl_tasks = f"{_ico_lr} To do — {_date_lr} · {_l2_lr}: {_k2_lr:.1f}%"
                st.markdown(f"""
                <div style="font-size:11px;font-weight:700;color:#4A6080;
                            letter-spacing:2px;text-transform:uppercase;margin-bottom:8px;">
                    ✅ {_lbl_tasks}
                </div>""",unsafe_allow_html=True)

                if _tasks:
                    _tk = f"tasks_{username_d}_{_date_lr}"
                    if _tk not in st.session_state:
                        st.session_state[_tk] = [False]*len(_tasks)
                    while len(st.session_state[_tk]) < len(_tasks):
                        st.session_state[_tk].append(False)

                    _done = sum(st.session_state[_tk])
                    _pct  = int((_done/len(_tasks))*100)
                    st.markdown(f"""
                    <div style="background:white;border:1px solid #E2E8F0;border-radius:10px;
                                padding:12px 16px;margin-bottom:8px;">
                        <div style="display:flex;justify-content:space-between;
                                    font-size:11px;margin-bottom:6px;">
                            <span style="color:#4A6080;">{_done}/{len(_tasks)} {"terminée" if lang_d=="fr" else "done"}{"s" if _done>1 else ""}</span>
                            <span style="color:{_clr_lr};font-weight:700;">{_pct}%</span>
                        </div>
                        <div style="height:4px;background:#F0F4F8;border-radius:99px;overflow:hidden;">
                            <div style="height:100%;width:{_pct}%;background:{_clr_lr};border-radius:99px;"></div>
                        </div>
                    </div>""",unsafe_allow_html=True)

                    for _ti, _tt in enumerate(_tasks):
                        _chk = st.session_state[_tk][_ti]
                        _new = st.checkbox(_tt, value=_chk, key=f"task_{_tk}_{_ti}")
                        if _new != _chk:
                            st.session_state[_tk][_ti] = _new
                            st.rerun()
                else:
                    st.info("Lancez une analyse pour générer vos tâches." if lang_d=="fr"
                            else "Run an analysis to generate your tasks.")
            except Exception:
                pass

            # ══ NEWS sectorielles ══════════════════════════════════
            if can_access("news"):
                try:
                    _sec = "generique"
                    if "module" in _df_arch.columns:
                        _lm2 = _df_arch["module"].dropna().iloc[-1] if len(_df_arch)>0 else ""
                        _sec = "transport_routier" if str(_lm2)=="transport" else "stock_distribution"
                        if st.session_state.get("module","") == "transport":
                            _sec = "transport_routier"
                    render_news_widget(_sec, lang=lang_d)
                except Exception:
                    pass
            else:
                show_lock("news")

    # ── MON COMPTE ──────────────────────────────────────────────
    elif nav==_("nav_compte"):
        lang_c = st.session_state.get("language","fr")
        username_c = st.session_state.current_user
        plan_c = get_user_plan(username_c)
        plan_info_c = PLAN_LIMITS.get(plan_c, PLAN_LIMITS["starter"])

        # Titre
        st.markdown(f"""
        <div style="background:linear-gradient(135deg,#0B2545 0%,#0f2f5a 100%);
                    border-radius:14px;padding:22px 24px;margin-bottom:20px;
                    display:flex;align-items:center;gap:16px;">
            <div style="width:52px;height:52px;border-radius:50%;background:#00C896;
                        display:flex;align-items:center;justify-content:center;
                        font-family:Syne,sans-serif;font-size:18px;font-weight:800;
                        color:#0B2545;flex-shrink:0;">
                {username_c[:2].upper()}
            </div>
            <div>
                <div style="font-family:Syne,sans-serif;font-size:18px;font-weight:800;
                            color:white;">{username_c}</div>
                <div style="font-size:11px;color:rgba(255,255,255,0.5);margin-top:2px;">
                    {'My Account' if lang_c=='en' else 'Mon Compte'}
                </div>
            </div>
            <div style="margin-left:auto;padding:4px 12px;border-radius:20px;
                        background:{plan_info_c['bg']};color:{plan_info_c['color']};
                        border:1px solid {plan_info_c['color']}40;
                        font-size:11px;font-weight:700;">
                {plan_info_c['icon']} {plan_info_c['label']}
            </div>
        </div>
        """, unsafe_allow_html=True)

        # ── Section abonnement
        _sect_abo = "Mon abonnement" if lang_c=="fr" else "My subscription"
        st.markdown(f"**{_sect_abo}**")
        st.markdown(f"""
        <div style="background:white;border:1px solid #E2E8F0;border-radius:12px;
                    padding:16px 18px;margin-bottom:14px;">
            <div style="display:flex;justify-content:space-between;align-items:center;">
                <div>
                    <div style="font-size:11px;color:#8FA3BC;margin-bottom:4px;">
                        {'Plan actuel' if lang_c=='fr' else 'Current plan'}
                    </div>
                    <div style="font-size:16px;font-weight:800;
                                font-family:Syne,sans-serif;color:{plan_info_c['color']};">
                        {plan_info_c['icon']} {plan_info_c['label']}
                        <span style="font-size:12px;font-weight:400;color:#4A6080;
                                     margin-left:6px;">{plan_info_c['price']}</span>
                    </div>
                </div>
                <div style="text-align:right;">
                    <div style="font-size:11px;color:#8FA3BC;">
                        {'Modules' if lang_c=='en' else 'Modules'}
                    </div>
                    <div style="font-size:13px;font-weight:600;color:#0B2545;">
                        {'Stock + Transport' if plan_info_c.get('modules',1) >= 2 else '1 module'}
                    </div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        if plan_c != "expert":
            _upgrade_lbl = "Passer à un plan supérieur →" if lang_c=="fr" else "Upgrade your plan →"
            st.markdown(f"""<div style="font-size:12px;color:#059669;margin-bottom:16px;">
                📧 {_upgrade_lbl} contact@logiflo.io</div>""", unsafe_allow_html=True)

        # ── Section entreprise
        _sect_ent = "Mon entreprise" if lang_c=="fr" else "My company"
        st.markdown(f"**{_sect_ent}**")
        with st.form("form_company"):
            _company_curr = st.session_state.get("company_name","")
            _company_new = st.text_input(
                "Nom d'entreprise" if lang_c=="fr" else "Company name",
                value=_company_curr
            )
            _sector_opt = ["Transport","Distribution","Industrie","Agroalimentaire",
                           "Pharma","Retail","BTP","Autre"]
            _saved_sector = st.session_state.get("company_sector","")
            _sector_new = st.selectbox(
                "Secteur" if lang_c=="fr" else "Sector",
                _sector_opt,
                index=_sector_opt.index(_saved_sector) if _saved_sector in _sector_opt else 0
            )
            _save_btn = st.form_submit_button(
                "💾 Sauvegarder" if lang_c=="fr" else "💾 Save",
                use_container_width=True
            )
            if _save_btn:
                st.session_state["company_name"] = _company_new
                st.session_state["company_sector"] = _sector_new
                save_user_prefs(username_c, {
                    "company_name": _company_new,
                    "company_sector": _sector_new,
                })
                st.success("✓ Sauvegardé" if lang_c=="fr" else "✓ Saved")

        # ── Section préférences
        st.markdown("<br>", unsafe_allow_html=True)
        _sect_pref = "Mes préférences" if lang_c=="fr" else "Preferences"
        st.markdown(f"**{_sect_pref}**")
        _col_p1, _col_p2 = st.columns(2)
        with _col_p1:
            _lang_opt = ["🇫🇷 Français","🇬🇧 English"]
            _lang_curr = 1 if lang_c=="en" else 0
            _lang_new = st.selectbox(
                "Langue" if lang_c=="fr" else "Language",
                _lang_opt, index=_lang_curr
            )
            if "English" in _lang_new and lang_c != "en":
                st.session_state["language"] = "en"
                save_user_prefs(username_c, {"language":"en"})
                st.rerun()
            elif "Français" in _lang_new and lang_c != "fr":
                st.session_state["language"] = "fr"
                save_user_prefs(username_c, {"language":"fr"})
                st.rerun()
        with _col_p2:
            _mod_opts = ["📦 Stock","🚚 Transport"]
            _mod_curr = 1 if st.session_state.get("preferred_module")=="transport" else 0
            _mod_new = st.selectbox(
                "Module par défaut" if lang_c=="fr" else "Default module",
                _mod_opts, index=_mod_curr
            )
            _mod_val = "transport" if "Transport" in _mod_new else "stock"
            if _mod_val != st.session_state.get("preferred_module","stock"):
                st.session_state["preferred_module"] = _mod_val
                save_user_prefs(username_c, {"preferred_module": _mod_val})

        # Seuil rupture
        _seuil_lbl = "Seuil alerte rupture (unités)" if lang_c=="fr" else "Stockout alert threshold (units)"
        _seuil_curr = st.session_state.get("seuil_rupture", 5)
        _seuil_new = st.slider(_seuil_lbl, 0, 20, int(_seuil_curr))
        if _seuil_new != _seuil_curr:
            st.session_state["seuil_rupture"] = _seuil_new
            save_user_prefs(username_c, {"seuil_rupture": _seuil_new})

        # ── Scoring dernier audit
        st.markdown("<br>", unsafe_allow_html=True)
        _sect_score = "Mon dernier scoring" if lang_c=="fr" else "My last score"
        st.markdown(f"**{_sect_score}**")
        try:
            _arcs = load_archives_from_sheets(username_c)
            if _arcs is not None and not _arcs.empty:
                _last_arc = _arcs.sort_values("created_at" if "created_at" in _arcs.columns else "date",
                                               ascending=False).iloc[0]
                _k2 = float(_last_arc.get("kpi_2",0))
                _k2c = "#00C896" if _k2>=90 else ("#F39C12" if _k2>=75 else "#E8304A")
                _lbl2 = str(_last_arc.get("kpi_label_2",""))
                _mod_arc = str(_last_arc.get("module",""))
                _date_arc = str(_last_arc.get("date",""))
                st.markdown(f"""
                <div style="background:white;border:1px solid #E2E8F0;border-radius:12px;
                            padding:14px 18px;margin-bottom:8px;">
                    <div style="font-size:11px;color:#8FA3BC;margin-bottom:4px;">
                        {'📦' if _mod_arc=='stock' else '🚚'} {_mod_arc.upper()} — {_date_arc}
                    </div>
                    <div style="font-size:28px;font-weight:800;
                                font-family:Syne,sans-serif;color:{_k2c};">
                        {_k2:.1f}<span style="font-size:13px;font-weight:400;
                        color:#4A6080;">%</span>
                    </div>
                    <div style="font-size:11px;color:#4A6080;">{_lbl2}</div>
                </div>
                """, unsafe_allow_html=True)
        except Exception:
            pass


    elif nav==_("nav_params"):
        _lang_p = st.session_state.get("language","fr")
        _title_p = "Paramètres" if _lang_p=="fr" else "Settings"
        st.markdown(f"<h2 style='font-family:Syne,sans-serif;color:#0B2545;'>{_title_p}</h2>",
                    unsafe_allow_html=True)

        # ── Langue
        st.markdown("**🌐 Langue / Language**")
        _lang_opts = ["🇫🇷 Français","🇬🇧 English"]
        _lang_idx = 1 if _lang_p=="en" else 0
        _lang_sel = st.radio("", _lang_opts, index=_lang_idx, horizontal=True,
                              label_visibility="collapsed", key="params_lang_radio")
        _new_lang = "en" if "English" in _lang_sel else "fr"
        if _new_lang != _lang_p:
            st.session_state["language"] = _new_lang
            try:
                save_user_prefs(st.session_state.current_user, {"language": _new_lang})
            except Exception: pass
            st.rerun()

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Mode sombre
        st.markdown("**🌙 Thème / Theme**")
        _dark = st.session_state.get("dark_mode", False)
        _dark_lbl = ("Mode sombre" if _lang_p=="fr" else "Dark mode")
        _dark_new = st.toggle(_dark_lbl, value=_dark, key="params_dark_toggle")
        if _dark_new != _dark:
            st.session_state["dark_mode"] = _dark_new
            if _dark_new:
                st.markdown("""<style>
                .stApp { background-color: #0d1117 !important; color: #e6edf3 !important; }
                .stSidebar { background-color: #161b22 !important; }
                </style>""", unsafe_allow_html=True)
            st.info("Rechargez la page pour appliquer le thème." if _lang_p=="fr"
                    else "Reload the page to apply the theme.")

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Lien vers Mon Compte pour les autres params
        _compte_lbl = ("⚙️ Les autres paramètres (seuils, module par défaut, entreprise) "
                        "sont dans **Mon Compte**." if _lang_p=="fr" else
                        "⚙️ Other settings (thresholds, default module, company) "
                        "are in **My Account**.")
        st.info(_compte_lbl)

        st.markdown("<br>", unsafe_allow_html=True)
        # ── Fichier exemple
        st.markdown("**📋 Fichier modèle**" if _lang_p=="fr" else "**📋 Sample file**")
        try:
            _ex_bytes_p = generate_exemple_excel()
            if _ex_bytes_p:
                _ex_lbl_p = "📥 Télécharger le fichier exemple" if _lang_p=="fr" else "📥 Download sample file"
                st.download_button(
                    _ex_lbl_p, _ex_bytes_p,
                    "logiflo_modele.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="dl_exemple_params"
                )
                _hint = ("Importez ce fichier dans l'Espace de Travail pour voir Logiflo en action."
                         if _lang_p=="fr" else
                         "Import this file in the Workspace to see Logiflo in action.")
                st.caption(_hint)
        except Exception:
            pass

    elif nav==_("nav_workspace"):

        # ══ MODULE STOCK ══
        if st.session_state.module=="stock":
            st.title(_("stock_title"))
            ci,cb=st.columns([4,1])
            ci.markdown(f"**{_('active_profile')} : {st.session_state.stock_view}**")
            if cb.button(_("change_profile")): st.session_state.page="choix_profil_stock";st.rerun()
            st.markdown("<br>",unsafe_allow_html=True)
            st.markdown(f"""<div class='import-card'><h3>{_('stock_import')}</h3>
                <p>{_('stock_import_sub')}</p></div>""",unsafe_allow_html=True)
            up=st.file_uploader("",type=["csv","xlsx"],key="stock_upload")
            st.markdown("---")
            if up:
                pg=StepProgress([_("step_read"),_("step_detect"),_("step_calc")])
                pg.step(_("step_read"))
                try:
                    df_brut=pd.read_excel(up) if up.name.endswith("xlsx") else pd.read_csv(up,encoding='utf-8')
                except UnicodeDecodeError:
                    up.seek(0);df_brut=pd.read_csv(up,encoding='latin-1')
                except:
                    up.seek(0);df_brut=pd.read_csv(up,sep=';',encoding='latin-1')
                pg.step(_("step_detect"))
                df_propre,statut=smart_ingester_stock_ultime(df_brut,client_ai=client)
                pg.step(_("step_calc"));pg.done()
                if df_propre is None: st.error(statut)
                else:
                    if st.session_state.stock_view=="MANAGER":
                        st.session_state.df_stock_manager=df_propre
                    else:
                        st.session_state.df_stock_terrain=df_propre

            _df_key = "df_stock_manager" if st.session_state.stock_view=="MANAGER" else "df_stock_terrain"
            if st.session_state.get(_df_key) is not None:
                df=st.session_state[_df_key].copy()
                sans_prix=bool(df.get("_sans_prix",pd.Series([True])).iloc[0]) if "_sans_prix" in df.columns else True
                has_conso=bool(df.get("_has_conso",pd.Series([False])).iloc[0]) if "_has_conso" in df.columns else False
                if sans_prix: st.markdown(f"<span class='sans-prix-badge'>{_('stock_badge_no_price')}</span>",unsafe_allow_html=True)
                if has_conso: st.markdown(f"<span class='sans-prix-badge'>{_('stock_badge_conso')}</span>",unsafe_allow_html=True)
                else: st.markdown(f"<span class='sans-prix-badge'>{_('stock_badge_no_conso')}</span>",unsafe_allow_html=True)

                # Statuts uniformes — toujours présents quel que soit has_conso
                if has_conso:
                    df["_conso_moy"]=df["_conso_moy"].fillna(0)
                    df["Couverture_mois"]=np.where(df["_conso_moy"]>0,df["quantite"]/df["_conso_moy"],9999)
                    df["Statut"]=np.select(
                        [(df["quantite"]<=st.session_state.seuil_rupture),
                         (df["quantite"]>0)&(df["_conso_moy"]==0),
                         (df["quantite"]>0)&(df["Couverture_mois"]>6)],
                        ["🔴 Rupture","🔴 Dormant","🟠 Surstock"],default="🟢 OK")
                else:
                    df["Statut"]=np.where(
                        df["quantite"]<=st.session_state.seuil_rupture,
                        "🔴 Rupture","🟢 OK"
                    )

                df["valeur_totale"]=df["quantite"]*df["prix_unitaire"]
                val_totale=df["valeur_totale"].sum()
                ruptures=df[df["Statut"]=="🔴 Rupture"]
                tx_serv=(1-len(ruptures)/len(df))*100 if len(df)>0 else 100

                if not st.session_state.history_stock or st.session_state.history_stock[-1].get("valeur")!=val_totale:
                    st.session_state.history_stock.append({"date":datetime.datetime.now().strftime("%H:%M:%S"),"valeur":val_totale})

                if st.session_state.stock_view=="MANAGER":  # ── MANAGER UNIQUEMENT
                    c1,c2,c3=st.columns(3)
                    kpi1_label=_("stock_kpi_capital") if not sans_prix else _("stock_kpi_articles")
                    kpi1_val=f"{val_totale:,.0f} €" if not sans_prix else str(len(df))
                    kpi1_color="#0B2545"

                    # ── Tooltips KPI (surligné desktop, ❓ mobile) ──
                    _TIP_CSS = ("<style>.kpi-tip{border-bottom:1.5px dashed #4A6080;cursor:help;}"
                                "@media(max-width:600px){.kpi-tip{border-bottom:none;}"
                                ".kpi-q{display:inline!important;}}.kpi-q{display:none;}</style>")
                    def _mk(col,title,value,color,tip_d,tip_b):
                        col.markdown(
                            _TIP_CSS+"<div class='kpi-card'>"
                            "<span class='kpi-tip' title='"+tip_d+" | "+tip_b+"'>"+title+"</span>"
                            "<span class='kpi-q'> ❓</span>"
                            "<h2 style='color:"+color+";'>"+value+"</h2></div>",
                            unsafe_allow_html=True)
                    _lkpi = st.session_state.get("language","fr")
                    _tips = (
                        [("Capital immobilisé","< 60 jours de CA"),
                         ("Taux de service","Benchmark : > 93% B2B"),
                         ("Références en rupture","Objectif : < 2% des références")]
                        if _lkpi=="fr" else
                        [("Tied-up capital","Norm: < 60 days revenue"),
                         ("Service level","Benchmark: > 93% B2B"),
                         ("Stockouts","Target: < 2% of references")]
                    )
                    _mk(c1,_tips[0][0],kpi1_val,kpi1_color,
                         'Capital immob. en EUR. Trop élevé = trésorerie sous pression.' if _lkpi=='fr' else 'Tied-up capital. Excess stock strains cash flow.',_tips[0][1])
                    _mk(c2,_tips[1][0],f"{tx_serv:.1f} %","#00C896",
                         '% commandes livrées sans rupture.' if _lkpi=='fr' else '% orders fulfilled without stockout.',_tips[1][1])
                    _mk(c3,_tips[2][0],str(len(ruptures)),"#E8304A",
                         'Références à zéro ou sous le seuil d alerte.' if _lkpi=='fr' else 'References at zero or below alert threshold.',_tips[2][1])
                    st.markdown("<br>",unsafe_allow_html=True)
                    cp,cl2=st.columns(2)
                    cmap={"🔴 Rupture":"#E8304A","🟢 OK":"#00C896","🟢 OK":"#00C896",
                          "🔴 Dormant":"#c0392b","🟠 Surstock":"#f39c12"}
                    with cp:
                        fig_pie=px.pie(df,names="Statut",hole=0.4,color="Statut",color_discrete_map=cmap)
                        fig_pie.update_layout(paper_bgcolor="rgba(0,0,0,0)",font=dict(family="DM Sans"))
                        st.plotly_chart(fig_pie,use_container_width=True)
                    with cl2:
                        if has_conso:
                            top15=df.nlargest(15,"_conso_moy")[["reference","_conso_moy","quantite"]].copy()
                            fig_conso=px.bar(top15,x="reference",y=["quantite","_conso_moy"],barmode="group",
                                color_discrete_map={"quantite":"#0B2545","_conso_moy":"#00C896"})
                            fig_conso.update_layout(paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)")
                            st.plotly_chart(fig_conso,use_container_width=True)
                        else:
                            fig_line=px.line(pd.DataFrame(st.session_state.history_stock),x="date",y="valeur")
                            fig_line.update_traces(line_color="#00C896")
                            fig_line.update_layout(paper_bgcolor="rgba(0,0,0,0)")
                            st.plotly_chart(fig_line,use_container_width=True)

                    col_audit,col_save=st.columns([3,1])
                    with col_audit: run_ia=st.button(_("stock_btn_ia"),use_container_width=True)
                    with col_save:
                        if st.button(_("stock_btn_save"),use_container_width=True,key="save_stock_early"):
                            kpi1=val_totale if not sans_prix else float(len(df))
                            label1=_("stock_kpi_capital") if not sans_prix else _("stock_kpi_articles")
                            ok=save_audit_to_sheets(st.session_state.current_user,"stock",len(df),
                                [kpi1,tx_serv,len(ruptures)],[label1,_("stock_kpi_service"),_("stock_kpi_rupture")],
                                st.session_state.analysis_stock_manager or "",st.session_state.last_pdf or b"")
                            if ok: st.success(_("stock_saved"))
                            else: st.warning(_("stock_save_err"))

                    # ── Prédiction rupture (visible à partir du 3e audit) ──
                    _lang_pred = st.session_state.get("language","fr")
                    if can_access("prediction"):
                        try:
                            _nb_pred = 0
                            _arc_pred = load_archives_from_sheets(st.session_state.current_user)
                            if _arc_pred is not None and not _arc_pred.empty:
                                _col_mp = "module" if "module" in _arc_pred.columns else None
                                _nb_pred = len(_arc_pred[_arc_pred[_col_mp]=="stock"]) if _col_mp else len(_arc_pred)
                            if _nb_pred >= 2:
                                _preds_ui = predict_ruptures(df, lang=_lang_pred)
                                st.session_state["_last_predictions"] = _preds_ui
                                render_prediction_rupture(df, lang=_lang_pred)
                            else:
                                _rem = max(0, 2 - _nb_pred)
                                st.info(
                                    f"💡 La prédiction de rupture apparaîtra à partir du 3e audit ({_rem} audit(s) restant(s))."
                                    if _lang_pred=="fr" else
                                    f"💡 Stockout prediction available from the 3rd audit ({_rem} remaining)."
                                )
                        except Exception:
                            pass
                    else:
                        show_lock("prediction")

                    if run_ia:
                        _ia_txt = "Deep AI Analysis in progress..." if st.session_state.get("language","fr")=="en" else "Analyse approfondie IA en cours..."
                        pg2=StepProgress([_("step_read"),_("step_ia"),_("step_report")],text=_ia_txt)
                        pg2.step(_("step_read"))
                        df_tox=df[df["Statut"].isin(["🔴 Dormant","🟠 Surstock"])]
                        pires=df_tox.nlargest(3,"quantite") if not df_tox.empty else df.nlargest(3,"quantite")
                        top_str=", ".join([f"{r['reference']} (qty:{r['quantite']:.0f})" for _ii,r in pires.iterrows()])
                        rupt_l=ruptures.nlargest(3,"quantite")["reference"].astype(str).tolist() if not ruptures.empty else "None"
                        med_info=""
                        if not has_conso: med_info=" BLIND SPOT: no consumption history. Sector benchmark: 2-4 months healthy coverage."
                        else:
                            cm_glob=df["_conso_moy"].mean()
                            cv_moy=df["Couverture_mois"].replace(9999,np.nan).mean()
                            med_info=f" Avg consumption: {cm_glob:.1f}/period. Avg coverage: {cv_moy:.1f} months."
                        prix_info="" if sans_prix else f" Tied-up capital: {val_totale:.0f} EUR."
                        pg2.step(_("step_ia"))
                        # Chargement historique avec KPIs courants
                        _kpis_curr_s=[val_totale if not sans_prix else float(len(df)),tx_serv,float(len(ruptures))]
                        _labels_curr_s=[_("stock_kpi_capital") if not sans_prix else _("stock_kpi_articles"),_("stock_kpi_service"),_("stock_kpi_rupture")]
                        _hist_s=get_historique_audits(st.session_state.current_user,"stock",
                                                      current_kpis=_kpis_curr_s,current_labels=_labels_curr_s)
                        _hist_txt_s=format_historique_pour_prompt(_hist_s,"stock",st.session_state.get("language","fr"))
                        _sector_s = detect_sector(df=df, module="stock")
                        st.session_state.analysis_stock_manager=generate_ai_analysis(
                            f"Items: {len(df)}. Service level: {tx_serv:.1f}%. Stock-outs: {len(ruptures)}. "
                            f"Top dormant: {top_str}. Top stock-outs: {rupt_l}.{prix_info}{med_info} "
                            f"Prices: {'No' if sans_prix else 'Yes'}. Consumption history: {'Yes' if has_conso else 'No'}.",
                            historique_txt=_hist_txt_s,
                            df_raw=df,
                            sector_key=_sector_s)
                        # KPIs calculés AVANT generate_expert_pdf
                        kpi1=val_totale if not sans_prix else float(len(df))
                        label1=_("stock_kpi_capital") if not sans_prix else _("stock_kpi_articles")
                        _kpis_final=[kpi1, tx_serv, float(len(ruptures))]
                        _labels_final=[label1, _("stock_kpi_service"), _("stock_kpi_rupture")]
                        st.session_state.last_kpis=_kpis_final
                        st.session_state.last_labels=_labels_final
                        # Courbe évolution historique pour le PDF
                        _arc_pdf_s = load_archives_from_sheets(st.session_state.current_user)
                        fig_hist_pdf = None
                        try:
                            if _arc_pdf_s is not None and not _arc_pdf_s.empty:
                                _df_hs = _arc_pdf_s[_arc_pdf_s["module"]=="stock"].copy()
                                if len(_df_hs) >= 2:
                                    for _c in ["kpi_2"]:
                                        _df_hs[_c] = pd.to_numeric(_df_hs[_c], errors="coerce").fillna(0)
                                    import plotly.graph_objects as _go_pdfh
                                    fig_hist_pdf = _go_pdfh.Figure()
                                    fig_hist_pdf.add_trace(_go_pdfh.Scatter(
                                        x=list(range(len(_df_hs))),
                                        y=_df_hs["kpi_2"].tolist(),
                                        mode="lines+markers",
                                        line=dict(color="#00C896", width=2),
                                        marker=dict(size=7, color="#00C896"),
                                        fill="tozeroy",
                                        fillcolor="rgba(0,200,150,0.1)",
                                    ))
                                    _lbl2_pdf = str(_df_hs["kpi_label_2"].iloc[-1])
                                    fig_hist_pdf.update_layout(
                                        title=dict(text=_lbl2_pdf, font=dict(size=11,color="#0B2545"), x=0),
                                        xaxis=dict(tickmode="array",tickvals=list(range(len(_df_hs))),
                                                   ticktext=[str(d) for d in _df_hs["date"].tolist()],
                                                   tickfont=dict(size=8),showgrid=False),
                                        yaxis=dict(tickfont=dict(size=8),gridcolor="rgba(0,0,0,0.05)"),
                                        plot_bgcolor="white",paper_bgcolor="white",
                                        margin=dict(t=30,b=20,l=30,r=10),
                                        height=260,showlegend=False,
                                    )
                        except Exception:
                            fig_hist_pdf = None
                        figs_pdf=[fig_pie]
                        if has_conso: figs_pdf.append(fig_conso)
                        if fig_hist_pdf is not None: figs_pdf.append(fig_hist_pdf)
                        st.session_state.last_pdf=generate_expert_pdf(_("pdf_title_stock"),st.session_state.analysis_stock_manager,figs_pdf,kpis=_kpis_final,labels=_labels_final,module="stock")
                        pg2.done()

                    if st.session_state.analysis_stock_manager:
                        st.markdown(render_report(st.session_state.analysis_stock_manager,"manager"),unsafe_allow_html=True)
                        st.markdown("<br>",unsafe_allow_html=True)
                        if st.session_state.last_pdf:
                            st.download_button(_("stock_btn_dl"),st.session_state.last_pdf,"Audit_Stock_Logiflo.pdf",use_container_width=True)

                elif st.session_state.stock_view=="TERRAIN":  # ── TERRAIN UNIQUEMENT
                    c1,c2=st.columns(2)
                    c1.markdown(f"<div class='kpi-card'><h4>{_('stock_kpi_rupture')}</h4><h2 style='color:#E8304A;'>{len(ruptures)}</h2></div>",unsafe_allow_html=True)
                    c2.markdown(f"<div class='kpi-card'><h4>{_('stock_kpi_service')}</h4><h2 style='color:#00C896;'>{tx_serv:.1f} %</h2></div>",unsafe_allow_html=True)
                    st.markdown(f"### {_('stock_urgent')}")
                    if len(ruptures)>0:
                        cols_s=["reference","quantite","Statut"]
                        if has_conso: cols_s.append("_conso_moy")
                        st.dataframe(ruptures[cols_s],use_container_width=True)
                    else: st.success(_("stock_no_rupture"))
                    run_ops=st.button(_("stock_btn_ia_terrain"),use_container_width=True,key="terrain_ia")
                    if run_ops:
                        _ia_txt_t = "Deep AI Analysis in progress..." if st.session_state.get("language","fr")=="en" else "Analyse approfondie IA en cours..."
                        pg3=StepProgress([_("step_read"),_("step_ia"),_("step_report")],text=_ia_txt_t)
                        pg3.step(_("step_read"))
                        top_c=df.nsmallest(5,"quantite")
                        top_s=", ".join([f"{r['reference']} ({r['quantite']:.0f})" for _ii,r in top_c.iterrows()])
                        dorm_s="No history" if not has_conso else f"{len(df[df['_conso_moy']==0])} items no movement"
                        pg3.step(_("step_ia"))
                        # Chargement historique terrain
                        # Terrain : passer aussi les KPIs courants pour l'historique
                        _kpis_curr_t=[float(len(df)),float(len(ruptures)),(1-len(ruptures)/max(len(df),1))*100]
                        _labels_curr_t=["Articles","Ruptures","Service %"]
                        _hist_t = get_historique_audits(st.session_state.current_user,"stock",
                                                        current_kpis=_kpis_curr_t,current_labels=_labels_curr_t)
                        _hist_txt_t = format_historique_pour_prompt(_hist_t,"terrain",st.session_state.get("language","fr"))
                        _sector_t = detect_sector(df=df, module="stock")
                        st.session_state.analysis_stock_terrain=generate_ai_analysis(
                            f"Field stock: {len(df)} refs. Stock-outs: {len(ruptures)}. "
                            f"Lowest stocks: {top_s}. Dormant: {dorm_s}. "
                            f"Prices: {'No' if sans_prix else 'Yes'}.",
                            historique_txt=_hist_txt_t,
                            df_raw=df,
                            sector_key=_sector_t)
                        pg3.done()
                    if st.session_state.analysis_stock_terrain:
                        st.markdown(render_report(st.session_state.analysis_stock_terrain,"terrain"),unsafe_allow_html=True)
                        st.markdown(f"### {_('stock_full')}")
                        cols_s=["reference","quantite","Statut"]
                        if has_conso: cols_s.append("_conso_moy")
                        st.dataframe(df[cols_s],use_container_width=True,height=400)

        # ══ MODULE TRANSPORT ══
        elif st.session_state.module=="transport":
            st.title(_("trans_title"))
            st.markdown("<br>",unsafe_allow_html=True)
            st.markdown(f"""<div class='import-card'><h3>{_('trans_import')}</h3>
                <p>{_('trans_import_sub')}</p></div>""",unsafe_allow_html=True)
            up_t=st.file_uploader("",type=["csv","xlsx"],key="trans_upload")
            st.markdown("---")

            if up_t and st.session_state.trans_filename!=up_t.name:
                _lbl_calc = "Calcul en cours..." if st.session_state.get("language","fr")=="fr" else "Computing..."
                with st.spinner(_lbl_calc):
                    try: df_t=pd.read_excel(up_t) if up_t.name.endswith("xlsx") else pd.read_csv(up_t,encoding="utf-8")
                    except UnicodeDecodeError:
                        up_t.seek(0);df_t=pd.read_csv(up_t,encoding="latin-1")
                    # Supprimer lignes totalement vides ou avec client None/nan
                    df_t = df_t.dropna(how="all")
                    for _chk_col in df_t.columns[:3]:
                        _bad = df_t[_chk_col].astype(str).str.strip().str.lower().isin(["nan","none","null","","n/a","-"])
                        df_t = df_t[~_bad].copy()
                        break
                    mapping=auto_map_columns_with_ai(df_t)
                    dep_c_tmp=mapping.get("dep") if mapping.get("dep") in df_t.columns else None
                    arr_c_tmp=mapping.get("arr") if mapping.get("arr") in df_t.columns else None
                    mode_c_tmp=mapping.get("mode") if mapping.get("mode") in df_t.columns else None
                    mode_det,mode_label,mode_emoji=detect_transport_mode(df_t,dep_c_tmp,arr_c_tmp,mode_c_tmp)
                    st.session_state.trans_mapping=mapping
                    st.session_state.df_trans=df_t
                    st.session_state.trans_filename=up_t.name
                    st.session_state.trans_mode_detected=(mode_det,mode_label,mode_emoji)
                    if dep_c_tmp and arr_c_tmp:
                        df_t=smart_multimodal_router(df_t,dep_c_tmp,arr_c_tmp,mode_c_tmp)
                        st.session_state.df_trans=df_t
                        st.session_state.df_trans["_DIST_CALCULEE_DONE"]=True

            if st.session_state.df_trans is not None:
                df_t=st.session_state.df_trans
                mapping=st.session_state.trans_mapping
                if st.session_state.trans_mode_detected:
                    mode_det,mode_label,mode_emoji=st.session_state.trans_mode_detected
                    st.markdown(f"<div class='mode-badge'>{mode_label} {_('mode_detected')}</div>",unsafe_allow_html=True)

                def col(k): return mapping.get(k) if mapping.get(k) in df_t.columns else None
                tour_c=col("client") or df_t.columns[0]
                dep_c=col("dep");arr_c=col("arr");dist_c=col("dist")
                mode_c=col("mode");ca_c=col("ca");co_c=col("co");poids_c=col("poids")

                if not co_c:
                    for c in df_t.columns:
                        if any(k in str(c).lower() for k in ["cout","cost","achat","charge"]): co_c=c;break
                if not ca_c:
                    for c in df_t.columns:
                        if any(k in str(c).lower() for k in ["ca","revenue","revenu","facture"]): ca_c=c;break
                if not co_c: st.error(_("trans_no_cost"));st.stop()

                df_t["_CO"]=df_t[co_c].apply(super_clean)
                if ca_c: df_t["_CA"]=df_t[ca_c].apply(super_clean)
                else: df_t["_CA"]=df_t["_CO"]/0.85;st.warning(_("trans_ca_miss"))
                df_t["Marge_Nette"]=df_t["_CA"]-df_t["_CO"]

                # ORS déjà fait pendant l'import si dep/arr présents
                if dep_c and arr_c and "_DIST_CALCULEE" not in df_t.columns and "_DIST_CALCULEE_DONE" not in df_t.columns:
                    df_t=smart_multimodal_router(df_t,dep_c,arr_c,mode_c)
                    st.session_state.df_trans=df_t

                df_t["_DIST_FINALE"]=(df_t["_DIST_CALCULEE"] if "_DIST_CALCULEE" in df_t.columns and df_t["_DIST_CALCULEE"].sum()>0
                                      else (df_t[dist_c].apply(super_clean) if dist_c else 0))
                df_t["Rentabilité_%"]=np.where(df_t["_CA"]>0,df_t["Marge_Nette"]/df_t["_CA"]*100,0)
                df_t["_DS"]=df_t["_DIST_FINALE"].replace(0,1)
                df_t["Cout_KM"]=np.where(df_t["_DIST_FINALE"]>0,df_t["_CO"]/df_t["_DS"],0)

                poids_info=""
                if poids_c:
                    df_t["_POIDS"]=df_t[poids_c].apply(super_clean)
                    df_t["Cout_kg"]=np.where(df_t["_POIDS"]>0,df_t["_CO"]/df_t["_POIDS"].replace(0,1),0)
                    poids_info=f" Total weight: {df_t['_POIDS'].sum():,.0f} kg. Avg cost/kg: {df_t['Cout_kg'].mean():.3f} EUR."

                marge_tot=df_t["Marge_Nette"].sum(); ca_tot=df_t["_CA"].sum()
                taux=(marge_tot/ca_tot*100) if ca_tot>0 else 0
                traj_def=len(df_t[df_t["Marge_Nette"]<0]); cout_km=df_t["Cout_KM"].mean()
                toxiques=df_t[df_t["Marge_Nette"]<(df_t["_CA"]*0.05)]
                fuite=toxiques["_CO"].sum()-toxiques["_CA"].sum(); nb_tox=len(toxiques)

                c1,c2,c3=st.columns(3)
                c1.markdown(f"<div class='kpi-card'><h4>{_('trans_kpi_marge')}</h4><h2 style='color:#0B2545;'>{marge_tot:,.0f} €</h2></div>",unsafe_allow_html=True)
                c2.markdown(f"<div class='kpi-card'><h4>{_('trans_kpi_taux')}</h4><h2 style='color:#00C896;'>{taux:.1f} %</h2></div>",unsafe_allow_html=True)
                if fuite>0:
                    c3.markdown(f"<div class='kpi-card'><h4>{_('trans_kpi_fuite')}</h4><h2 style='color:#E8304A;'>-{fuite:,.0f} €</h2><p>{nb_tox} toxic routes</p></div>",unsafe_allow_html=True)
                else:
                    c3.markdown(f"<div class='kpi-card'><h4>{_('trans_kpi_sain')}</h4><h2 style='color:#00C896;'>OK</h2></div>",unsafe_allow_html=True)

                if poids_c: st.info(f"⚖️ Avg cost: **{df_t['Cout_kg'].mean():.3f} €/kg** | Total: **{df_t['_POIDS'].sum():,.0f} kg**")

                # Bouton save avant IA
                col_audit2,col_save2=st.columns([3,1])
                with col_audit2: run_ia_t=st.button(_("trans_btn_ia"),use_container_width=True)
                with col_save2:
                    if st.button(_("trans_btn_save"),use_container_width=True,key="save_trans_early"):
                        ok=save_audit_to_sheets(st.session_state.current_user,"transport",len(df_t),
                            [marge_tot,taux,nb_tox],[_("trans_kpi_marge"),_("trans_kpi_taux"),"Toxic routes"],
                            st.session_state.analysis_trans or "",st.session_state.last_pdf or b"")
                        if ok: st.success(_("stock_saved"))
                        else: st.warning(_("stock_save_err"))

                # ── GRAPHIQUES REFAITS ──
                st.markdown("<br>",unsafe_allow_html=True)
                df_plot=df_t.copy()
                df_plot["Statut"]=np.where(
                    df_plot["Rentabilité_%"]<0,"🔴 Loss / Perte",
                    np.where(df_plot["Rentabilité_%"]<10,"🟠 Alert / Alerte","🟢 Healthy / Sain"))
                CMAP={"🔴 Loss / Perte":"#E8304A","🟠 Alert / Alerte":"#f39c12","🟢 Healthy / Sain":"#00C896"}

                tab_top,tab_global=st.tabs([_("trans_tab_top"),_("trans_tab_all")])

                with tab_top:
                    top_n=df_plot.nsmallest(15,"Marge_Nette").sort_values("Marge_Nette")
                    top_n["label"]=top_n[tour_c].astype(str).str[:35]
                    top_n["pct_label"]=top_n["Rentabilité_%"].apply(lambda x:f"{x:.1f}%")
                    fig_top=px.bar(top_n,x="Marge_Nette",y="label",orientation="h",
                        color="Statut",color_discrete_map=CMAP,text="pct_label",
                        custom_data=["_CA","_CO","Rentabilité_%"],
                        title=_("trans_top15_title"),
                        labels={"Marge_Nette":"Margin / Marge (€)","label":""})
                    fig_top.update_traces(
                        textposition="outside",
                        hovertemplate=(
                            "<b>%{y}</b><br>"
                            "Margin: <b>%{x:,.0f} €</b><br>"
                            "Revenue: %{customdata[0]:,.0f} €<br>"
                            "Cost: %{customdata[1]:,.0f} €<br>"
                            "Rate: %{customdata[2]:.1f}%<extra></extra>"))
                    fig_top.update_layout(
                        paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",
                        height=520,showlegend=False,
                        margin=dict(l=20,r=90,t=50,b=20),
                        xaxis=dict(title="Margin / Marge (€)",tickformat=",.0f",
                                   gridcolor="#f0f4f8",zerolinecolor="#0B2545",zerolinewidth=2),
                        yaxis=dict(title="",tickfont=dict(size=11)),
                        font=dict(family="DM Sans",size=12,color="#0B2545"),
                        title=dict(font=dict(family="Syne",size=14,color="#0B2545")))
                    st.plotly_chart(fig_top,use_container_width=True)
                    # Tableau détail
                    st.markdown(f"**{_('trans_detail')}**")
                    cols_show=[tour_c,"_CA","_CO","Marge_Nette","Rentabilité_%","Statut"]
                    cols_show=[c for c in cols_show if c in df_t.columns]
                    rename_map={tour_c:_("trans_col_client"),"_CA":_("trans_col_ca"),
                                "_CO":_("trans_col_co"),"Marge_Nette":_("trans_col_marge"),
                                "Rentabilité_%":_("trans_col_pct")}
                    display_df=top_n[cols_show].rename(columns=rename_map)
                    num_cols=[_("trans_col_ca"),_("trans_col_co"),_("trans_col_marge")]
                    fmt={c:"{:,.0f}" for c in num_cols if c in display_df.columns}
                    if _("trans_col_pct") in display_df.columns: fmt[_("trans_col_pct")]="{:.1f}%"
                    # pandas >= 2.1 : map() remplace applymap()
                    _style = display_df.style.format(fmt)
                    _neg_cols = [c for c in [_("trans_col_marge"),_("trans_col_pct")] if c in display_df.columns]
                    if _neg_cols:
                        try:
                            _style = _style.map(
                                lambda v:"color:#E8304A;font-weight:600" if isinstance(v,(int,float)) and v<0 else "",
                                subset=_neg_cols)
                        except AttributeError:
                            _style = _style.applymap(
                                lambda v:"color:#E8304A;font-weight:600" if isinstance(v,(int,float)) and v<0 else "",
                                subset=_neg_cols)
                    st.dataframe(_style, use_container_width=True, height=380)

                with tab_global:
                    fig_scatter=px.scatter(df_plot,x="_CA",y="Rentabilité_%",
                        color="Statut",color_discrete_map=CMAP,
                        size=df_plot["_CO"].clip(lower=1),size_max=40,
                        hover_name=tour_c,custom_data=["Marge_Nette","_CO"],
                        title=_("trans_scatter_title"),
                        labels={"_CA":"Revenue / CA (€)","Rentabilité_%":"Margin Rate / Taux Marge (%)"})
                    fig_scatter.update_traces(
                        hovertemplate=(
                            "<b>%{hovertext}</b><br>"
                            "Revenue: %{x:,.0f} €<br>"
                            "Margin: %{customdata[0]:,.0f} €<br>"
                            "Cost: %{customdata[1]:,.0f} €<br>"
                            "Rate: %{y:.1f}%<extra></extra>"))
                    fig_scatter.add_hline(y=0,line_dash="solid",line_color="#E8304A",line_width=2,
                        annotation_text=_("trans_seuil_zero"),annotation_position="right")
                    fig_scatter.add_hline(y=10,line_dash="dot",line_color="#f39c12",line_width=1.5,
                        annotation_text=_("trans_seuil_alert"),annotation_position="right")
                    fig_scatter.update_layout(
                        paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="#fafbfc",height=500,
                        margin=dict(l=20,r=130,t=50,b=20),
                        xaxis=dict(title="Revenue / CA (€)",tickformat=",.0f",gridcolor="#f0f4f8"),
                        yaxis=dict(title="Margin Rate (%)",ticksuffix="%",gridcolor="#f0f4f8",
                                   zerolinecolor="#E8304A",zerolinewidth=1.5),
                        font=dict(family="DM Sans",size=12,color="#0B2545"),
                        title=dict(font=dict(family="Syne",size=14,color="#0B2545")),
                        legend=dict(title="",orientation="h",yanchor="bottom",y=1.02))
                    st.plotly_chart(fig_scatter,use_container_width=True)
                    n_loss=len(df_plot[df_plot["Statut"]=="🔴 Loss / Perte"])
                    n_alert=len(df_plot[df_plot["Statut"]=="🟠 Alert / Alerte"])
                    n_ok=len(df_plot[df_plot["Statut"]=="🟢 Healthy / Sain"])
                    st.caption(f"📊 {len(df_plot)} routes — {n_loss} loss | {n_alert} alert | {n_ok} healthy")

                fig_trans=fig_top  # pour le PDF

                if run_ia_t:
                    _spin_lbl = "Analyse IA en cours..." if st.session_state.get("language","fr")=="fr" else "AI Analysis in progress..."
                    with st.spinner(_spin_lbl):
                        # Filtrer les NaN dans les données transport avant l'analyse
                        top3=df_t[df_t["Marge_Nette"].notna() & df_t[tour_c].notna()].nsmallest(3,"Marge_Nette")
                        pires_s=", ".join([
                            f"{str(r[tour_c]).strip()} ({r['Marge_Nette']:.0f} EUR)"
                            for _ii,r in top3.iterrows()
                            if str(r[tour_c]).strip() not in ("","nan","None","NaN")
                        ]) if not top3.empty else "None"
                        mode_info=f" Dominant transport mode: {st.session_state.trans_mode_detected[0] if st.session_state.trans_mode_detected else 'road'}."
                        _kpis_tr=[marge_tot,taux,nb_tox]
                        _labels_tr=[_("trans_kpi_marge"),_("trans_kpi_taux"),"Toxic"]
                        _hist_tr=get_historique_audits(st.session_state.current_user,"transport",
                                                       current_kpis=_kpis_tr,current_labels=_labels_tr)
                        _hist_txt_tr=format_historique_pour_prompt(_hist_tr,"transport",st.session_state.get("language","fr"))
                        _mode_k = st.session_state.trans_mode_detected[0] if st.session_state.get("trans_mode_detected") else "routier"
                        _sector_tr = detect_sector(df=df_t, module="transport", mode_detected=_mode_k)
                        st.session_state.analysis_trans=generate_ai_analysis(
                            f"Routes: {len(df_t)}. Total margin: {marge_tot:.0f} EUR. Rate: {taux:.1f}%. "
                            f"Loss routes: {traj_def}. Top 3 worst: {pires_s}. Avg cost/km: {cout_km:.2f} EUR.{poids_info}{mode_info}",
                            historique_txt=_hist_txt_tr,
                            df_raw=df_t,
                            sector_key=_sector_tr,
                            mode_detected=_mode_k)
                        st.session_state.last_kpis=_kpis_tr
                        st.session_state.last_labels=_labels_tr
                        st.session_state.last_pdf=generate_expert_pdf(_("pdf_title_trans"),st.session_state.analysis_trans,[fig_trans],kpis=_kpis_tr,labels=_labels_tr,module="transport")

                if st.session_state.analysis_trans:
                    st.markdown(render_report(st.session_state.analysis_trans,"manager"),unsafe_allow_html=True)
                    st.markdown("<br>",unsafe_allow_html=True)
                    if st.session_state.last_pdf:
                        st.download_button(_("trans_btn_dl"),st.session_state.last_pdf,"Transport_Logiflo.pdf",use_container_width=True)
